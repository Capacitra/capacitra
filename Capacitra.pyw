"""
Capacitra — Storage capacity intelligence for Windows.

A single-file, offline, no-telemetry disk-space analyzer. Pure Python +
tkinter, zero external runtime dependencies. Website: capacitra.com

Copyright (C) 2026 Samet Ozcan

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful, but
WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program. If not, see <https://www.gnu.org/licenses/>.

Run with:
    pythonw Capacitra.pyw          (no console)
or just double-click on Windows.

Features
--------
  * Fast multi-threaded scan with live progress
  * Explorer-style folder tree with inline size bars
  * Synced Treemap, Bar, Pie and Extensions panels
  * File-age distribution and Top-N largest files
  * SHA1-based duplicate finder (size pre-grouped)
  * Recycle Bin, Open, Show in Explorer, Copy Path
  * CSV and rich HTML report export
  * Light & Dark themes (Fluent-inspired)
  * Headless CLI (--scan / --export) for scripting & Task Scheduler
"""

import os
import sys
import threading
import queue
import time
import csv
import math
import html as html_lib
import hashlib
import subprocess
import argparse
import json
import heapq
from collections import defaultdict, OrderedDict

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ---------------------------------------------------------------------------
# Constants & helpers
# ---------------------------------------------------------------------------

APP_NAME = "Capacitra"
APP_TAGLINE = "Storage capacity intelligence"
APP_VERSION = "4.3.2"


# Optional dependency probes — done once at import so the UI can hide
# menu items that wouldn't work in this build.
def _has_module(name):
    try:
        __import__(name)
        return True
    except ImportError:
        return False


HAS_REPORTLAB  = _has_module("reportlab")
HAS_OPENPYXL   = _has_module("openpyxl")
HAS_SEND2TRASH = _has_module("send2trash") or sys.platform == "win32"


UI_FONT = ("Segoe UI", 9)
UI_FONT_BOLD = ("Segoe UI", 9, "bold")
UI_FONT_BIG = ("Segoe UI Semibold", 11)
UI_FONT_TITLE = ("Segoe UI Semibold", 14)
ICON_FONT = ("Segoe UI Emoji", 16)
ICON_FONT_SM = ("Segoe UI Emoji", 11)


def human_size(num):
    if num is None:
        return ""
    if num < 1024:
        return f"{num} B"
    val = float(num)
    for u in ("KB", "MB", "GB", "TB", "PB"):
        val /= 1024.0
        if val < 1024 or u == "PB":
            if val >= 100:
                return f"{val:.0f} {u}"
            if val >= 10:
                return f"{val:.1f} {u}"
            return f"{val:.2f} {u}"
    return f"{num} B"


def format_age(seconds):
    days = seconds / 86400.0
    if days < 1:
        h = int(seconds / 3600)
        return f"{max(h, 0)}h" if h < 24 else "1d"
    if days < 30:
        return f"{int(days)}d"
    if days < 365:
        return f"{int(days / 30)}mo"
    return f"{days / 365:.1f}y"


def bar_string(percent, width=10):
    """Unicode block bar like '████▌░░░░░' for percent (0–100)."""
    if percent < 0:
        percent = 0
    if percent > 100:
        percent = 100
    full = percent / 100.0 * width
    whole = int(full)
    frac = full - whole
    bar = "█" * whole
    if whole < width:
        if frac >= 0.5:
            bar += "▌"
            bar += "░" * (width - whole - 1)
        else:
            bar += "░" * (width - whole)
    return bar


def open_in_explorer(path):
    """Open the system file manager and highlight `path` (not just its
    parent folder). On Windows, Explorer.exe has nonstandard cmdline
    parsing; we use Win32 ShellExecuteW directly for reliability."""
    if sys.platform == "win32":
        try:
            norm = os.path.normpath(path)
            # Strip the Win32 long-path prefix Explorer does not understand
            if norm.startswith("\\\\?\\"):
                norm = norm[4:]
            # ShellExecuteW is the only reliable way to invoke
            # "explorer /select,<path>" on Windows. subprocess.Popen
            # always wraps argv items containing colons in quotes which
            # Explorer then refuses to parse, opening the wrong folder
            # (usually Documents) instead of selecting the file.
            from ctypes import windll
            # 1 = SW_SHOWNORMAL
            rc = windll.shell32.ShellExecuteW(
                None, "open", "explorer.exe",
                f'/select,"{norm}"', None, 1)
            if rc <= 32:
                raise OSError(f"ShellExecuteW failed (code {rc})")
        except Exception:
            # Last-ditch: open the parent folder so the user is at
            # least near the file.
            try:
                os.startfile(os.path.dirname(path) or path)
            except Exception:
                pass
    elif sys.platform == "darwin":
        subprocess.Popen(["open", "-R", path])
    else:
        # xdg-open has no "reveal" semantics; try nautilus/dolphin
        # selection helpers first, fall back to opening the folder.
        for cmd in (["nautilus", "--select", path],
                    ["dolphin", "--select", path],
                    ["nemo", path]):
            try:
                subprocess.Popen(cmd)
                return
            except FileNotFoundError:
                continue
        subprocess.Popen(["xdg-open", os.path.dirname(path) or "."])


def _sanitize_cell(v):
    """Prefix leading spreadsheet-formula-trigger characters with a
    single quote. Prevents CSV/XLSX injection attacks where a filename
    like `=cmd|'/c calc'!A1` would auto-execute when the report is
    opened in Excel / LibreOffice / Numbers."""
    if isinstance(v, str) and v and v[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + v
    return v


def open_path(path):
    try:
        if sys.platform == "win32":
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception as e:
        # No CapacitraApp instance available here; surface to stderr.
        sys.stderr.write(f"Cannot open: {e}\n")


def long_path(path):
    """Return a Win32 long-path-safe form of *path* (>=260 chars).

    On Windows we add the special ``\\\\?\\`` prefix so the file system
    APIs accept paths longer than MAX_PATH. No-op on other platforms.
    """
    if sys.platform != "win32" or not path:
        return path
    if path.startswith("\\\\?\\") or path.startswith("\\\\.\\"):
        return path
    if len(path) < 240:
        return path
    if path.startswith("\\\\"):
        # UNC path → \\?\UNC\server\share\…
        return "\\\\?\\UNC\\" + path[2:]
    return "\\\\?\\" + path


def get_file_owner(path):
    """Return the owner of *path* as 'DOMAIN\\user' on Windows, the
    username on POSIX, or None if it cannot be determined.

    Uses GetFileSecurityW + LookupAccountSidW via ctypes on Windows so
    no extra dependency is needed. Each call costs ~5-30 ms, so the
    UI is expected to call this lazily for visible rows only, never
    inline during a scan.
    """
    if sys.platform == "win32":
        try:
            from ctypes import (windll, c_void_p, c_ulong, c_int,
                                byref, create_unicode_buffer, c_char)
            OWNER_INFO = 0x00000001
            adv = windll.advapi32
            needed = c_ulong(0)
            adv.GetFileSecurityW(path, OWNER_INFO, None, 0, byref(needed))
            if not needed.value:
                return None
            buf = (c_char * needed.value)()
            if not adv.GetFileSecurityW(path, OWNER_INFO, buf,
                                        needed.value, byref(needed)):
                return None
            owner_sid = c_void_p()
            defaulted = c_int(0)
            if not adv.GetSecurityDescriptorOwner(
                    buf, byref(owner_sid), byref(defaulted)):
                return None
            name = create_unicode_buffer(256)
            domain = create_unicode_buffer(256)
            name_sz = c_ulong(256)
            domain_sz = c_ulong(256)
            sid_type = c_ulong()
            if not adv.LookupAccountSidW(
                    None, owner_sid, name, byref(name_sz),
                    domain, byref(domain_sz), byref(sid_type)):
                return None
            if domain.value:
                return f"{domain.value}\\{name.value}"
            return name.value
        except Exception:
            return None
    # POSIX fallback
    try:
        import pwd
        st = os.stat(path)
        return pwd.getpwuid(st.st_uid).pw_name
    except Exception:
        return None


def send_to_recycle_bin(path):
    """Move *path* to the OS Recycle Bin / Trash.

    On Windows we use the SHFileOperationW shell call (always undoable).
    On macOS and Linux we look for the `send2trash` package — if it is
    not installed we *refuse* to delete instead of silently doing a
    permanent removal, so users don't lose data they thought they could
    restore.
    Returns True on success, False otherwise.
    """
    if sys.platform == "win32":
        try:
            from ctypes import (wintypes, windll, Structure,
                                c_int, byref, c_wchar_p)
            FO_DELETE = 0x0003
            FOF_ALLOWUNDO = 0x0040
            FOF_NOCONFIRMATION = 0x0010
            FOF_SILENT = 0x0004
            FOF_NOERRORUI = 0x0400

            class SHFILEOPSTRUCTW(Structure):
                _fields_ = [
                    ("hwnd", wintypes.HWND),
                    ("wFunc", wintypes.UINT),
                    ("pFrom", c_wchar_p),
                    ("pTo", c_wchar_p),
                    ("fFlags", c_int),
                    ("fAnyOperationsAborted", wintypes.BOOL),
                    ("hNameMappings", wintypes.LPVOID),
                    ("lpszProgressTitle", c_wchar_p),
                ]
            op = SHFILEOPSTRUCTW()
            op.wFunc = FO_DELETE
            op.pFrom = path + "\0\0"
            op.fFlags = (FOF_ALLOWUNDO | FOF_NOCONFIRMATION
                         | FOF_SILENT | FOF_NOERRORUI)
            return windll.shell32.SHFileOperationW(byref(op)) == 0
        except Exception:
            return False
    # macOS / Linux: require send2trash for a real (undoable) trash op.
    try:
        import send2trash
        send2trash.send2trash(path)
        return True
    except ImportError:
        # No send2trash → refuse to delete to avoid data loss.
        return False
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Themes
# ---------------------------------------------------------------------------

THEMES = {
    "light": {
        "bg":            "#F8FAFC",
        "panel":         "#FFFFFF",
        "panel_alt":     "#F1F5F9",
        "panel_hover":   "#E2E8F0",
        "row_alt":       "#F8FAFC",
        "fg":            "#0F172A",
        "fg_subtle":     "#374151",
        "muted":         "#6B7280",
        "accent":        "#2563EB",
        "accent_dark":   "#1D4ED8",
        "accent_soft":   "#EFF6FF",
        "bar_fill":      "#2563EB",
        "border":        "#E2E8F0",
        "border_strong": "#CBD5E1",
        "select":        "#DBEAFE",
        "header_bg":     "#FFFFFF",
        "ribbon_bg":     "#F1F5F9",
        "ribbon_label":  "#6B7280",
        "chart_grid":    "#E2E8F0",
        "chart_text":    "#0F172A",
        "danger":        "#DC2626",
        "success":       "#16A34A",
        "warn":          "#EA580C",
        # Sidebar (always dark, even in light mode — matches mock)
        "sb_bg":         "#161D2D",
        "sb_bg_active":  "#1F2C44",
        "sb_bg_hover":   "#1B2336",
        "sb_fg":         "#9CA3AF",
        "sb_fg_active":  "#FFFFFF",
        "sb_label":      "#6B7383",
        "sb_border":     "#21283A",
        "logo_a":        "#5B7CFA",
        "logo_b":        "#9F7AEA",
        # Card shadow tone
        "shadow":        "#E2E8F0",
    },
    "dark": {
        "bg":            "#0B1220",
        "panel":         "#111827",
        "panel_alt":     "#1F2937",
        "panel_hover":   "#1F2937",
        "row_alt":       "#111827",
        "fg":            "#E8EAED",
        "fg_subtle":     "#C8CCD3",
        "muted":         "#8B92A0",
        "accent":        "#3B82F6",
        "accent_dark":   "#2563EB",
        "accent_soft":   "#1B3157",
        "bar_fill":      "#3B82F6",
        "border":        "#1F2937",
        "border_strong": "#374151",
        "select":        "#1F3157",
        "header_bg":     "#0F172A",
        "ribbon_bg":     "#0F172A",
        "ribbon_label":  "#8B92A0",
        "chart_grid":    "#1F2937",
        "chart_text":    "#E8EAED",
        "danger":        "#EF4444",
        "success":       "#22C55E",
        "warn":          "#F97316",
        "sb_bg":         "#0B111C",
        "sb_bg_active":  "#162038",
        "sb_bg_hover":   "#13192A",
        "sb_fg":         "#9CA3AF",
        "sb_fg_active":  "#FFFFFF",
        "sb_label":      "#6B7383",
        "sb_border":     "#1A2030",
        "logo_a":        "#5B7CFA",
        "logo_b":        "#9F7AEA",
        "shadow":        "#040608",
    },
}

# A modern palette for charts/treemap
PALETTE = [
    "#0078D4", "#107C10", "#D83B01", "#5C2D91", "#E81123",
    "#008272", "#B4009E", "#FFAA44", "#018574", "#7A7574",
    "#4F6BED", "#C239B3", "#0F548B", "#498205", "#A4262C",
    "#00BCF2", "#FFA500", "#A0AEC0",
]


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

class Node:
    __slots__ = ("name", "path", "size", "allocated", "is_dir",
                 "children", "file_count", "folder_count",
                 "parent", "mtime", "atime")

    def __init__(self, name, path, is_dir=False, parent=None):
        self.name = name
        self.path = path
        self.is_dir = is_dir
        self.size = 0           # logical size (sum of bytes)
        self.allocated = 0      # cluster-rounded disk footprint
        self.children = []
        self.file_count = 0     # recursive descendant file count
        self.folder_count = 0   # recursive descendant folder count
        self.parent = parent
        self.mtime = 0          # last modification time
        self.atime = 0          # last access time


# ---------------------------------------------------------------------------
# Scan worker
# ---------------------------------------------------------------------------

class ScanWorker(threading.Thread):
    DUP_THRESHOLD = 256 * 1024  # 256 KB (catches more candidates by default)

    def __init__(self, root_path, result_queue, stop_event, excludes=None):
        super().__init__(daemon=True)
        self.root_path = root_path
        self.result_queue = result_queue
        self.stop_event = stop_event
        # Folder names (case-insensitive) to skip during the scan.
        self.excludes = {e.lower() for e in (excludes or [])}
        self.scanned_files = 0
        self.total_bytes = 0
        self.total_dirs = 0
        self.denied_count = 0
        self.last_progress = 0
        self.ext_stats = defaultdict(lambda: [0, 0])
        self.largest_files = []
        self.size_groups = defaultdict(list)
        self.age_buckets = [
            ["< 7 days",      0,     7,        0, 0],
            ["7–30 days",     7,     30,       0, 0],
            ["30–90 days",    30,    90,       0, 0],
            ["3–12 months",   90,    365,      0, 0],
            ["1–3 years",     365,   365 * 3,  0, 0],
            ["> 3 years",     365*3, 10**9,    0, 0],
        ]
        self.now = time.time()


    def _push_largest(self, size, path, mtime, cap=1000):
        """Bounded min-heap for the top-N largest files (Perf-3)."""
        # First cap items: heappush (turns list into heap on first push).
        h = self.largest_files
        if len(h) < cap:
            heapq.heappush(h, (size, path, mtime))
        elif h and size > h[0][0]:
            heapq.heapreplace(h, (size, path, mtime))

    def run(self):
        try:
            # Auto-detect the real filesystem cluster size for this
            # scan root so the "Allocated" column is correct on 64K
            # NTFS / ReFS / exFAT volumes.
            try:
                self.__class__._CLUSTER_SIZE = \
                    self._detect_cluster_size(self.root_path)
            except Exception:
                pass
            root = Node(self.root_path, self.root_path, is_dir=True)
            self._scan(root)
            if self.stop_event.is_set():
                self.result_queue.put(("cancelled", None))
                return
            self.largest_files.sort(key=lambda x: x[0], reverse=True)
            exts = [(e, sz, cnt) for e, (sz, cnt) in self.ext_stats.items()]
            exts.sort(key=lambda x: x[1], reverse=True)
            self.result_queue.put(("done", {
                "root": root,
                "total_files": self.scanned_files,
                "total_bytes": self.total_bytes,
                "total_dirs": self.total_dirs,
                "denied": self.denied_count,
                "extensions": exts,
                "largest_files": self.largest_files[:1000],
                "age_buckets": self.age_buckets,
                "size_groups": dict(self.size_groups),
            }))
        except Exception as exc:
            self.result_queue.put(("error", str(exc)))

    def _bump_age(self, age_days, size):
        for b in self.age_buckets:
            if b[1] <= age_days < b[2]:
                b[3] += 1
                b[4] += size
                return

    # Cluster size is auto-detected per scan root via GetDiskFreeSpaceW
    # on Windows. On other platforms (or if the call fails) we fall
    # back to 4 KB, which matches typical NTFS/ext4 defaults.
    _CLUSTER_SIZE = 4096

    @staticmethod
    def _detect_cluster_size(path):
        """Return the filesystem's real allocation unit size in bytes
        for the drive containing `path`, or 4096 as fallback. Matters
        for 64K NTFS, ReFS, exFAT and older FAT32 volumes."""
        if sys.platform == "win32":
            try:
                from ctypes import (windll, c_ulong, byref, c_wchar_p)
                sectors = c_ulong()
                bps = c_ulong()
                freec = c_ulong()
                totalc = c_ulong()
                drive = os.path.splitdrive(path)[0]
                if drive:
                    root_path = drive + "\\"
                    if windll.kernel32.GetDiskFreeSpaceW(
                            c_wchar_p(root_path),
                            byref(sectors), byref(bps),
                            byref(freec), byref(totalc)):
                        cluster = sectors.value * bps.value
                        if cluster > 0:
                            return cluster
            except Exception:
                pass
        return 4096

    @classmethod
    def _allocated_for(cls, size):
        """Round file size up to the next cluster boundary."""
        if size <= 0:
            return 0
        return ((size + cls._CLUSTER_SIZE - 1) // cls._CLUSTER_SIZE) * cls._CLUSTER_SIZE

    def _scan(self, node):
        if self.stop_event.is_set():
            return 0
        total = 0
        allocated_total = 0
        file_count = 0
        folder_count = 0
        # Folder's own mtime/atime (for the Modified/Accessed columns)
        try:
            dst = os.stat(long_path(node.path))
            node.mtime = dst.st_mtime
            node.atime = dst.st_atime
        except (PermissionError, OSError):
            pass
        try:
            with os.scandir(long_path(node.path)) as it:
                entries = list(it)
        except (PermissionError, OSError):
            self.denied_count += 1
            return 0

        for entry in entries:
            if self.stop_event.is_set():
                return total
            try:
                if entry.is_symlink():
                    continue
                if entry.is_dir(follow_symlinks=False):
                    # Skip excluded folder names (case-insensitive)
                    if entry.name.lower() in self.excludes:
                        continue
                    self.total_dirs += 1
                    child = Node(entry.name, entry.path, is_dir=True, parent=node)
                    sub = self._scan(child)
                    child.size = sub
                    node.children.append(child)
                    total += sub
                    allocated_total += child.allocated
                    file_count += child.file_count
                    folder_count += 1 + child.folder_count
                else:
                    try:
                        st = entry.stat(follow_symlinks=False)
                        sz = st.st_size
                        mt = st.st_mtime
                        at = st.st_atime
                    except (PermissionError, OSError):
                        self.denied_count += 1
                        continue
                    f = Node(entry.name, entry.path, is_dir=False, parent=node)
                    f.size = sz
                    f.allocated = self._allocated_for(sz)
                    f.mtime = mt
                    f.atime = at
                    allocated_total += f.allocated
                    node.children.append(f)
                    total += sz
                    file_count += 1

                    ext = os.path.splitext(entry.name)[1].lower() or "(no ext)"
                    self.ext_stats[ext][0] += sz
                    self.ext_stats[ext][1] += 1
                    self._bump_age((self.now - mt) / 86400.0, sz)

                    if len(self.largest_files) < 1000:
                        self._push_largest(sz, entry.path, mt)
                        if len(self.largest_files) == 1000:
                            self.largest_files.sort(key=lambda x: x[0])
                    elif sz > self.largest_files[0][0]:
                        self.largest_files[0] = (sz, entry.path, mt)
                        self.largest_files.sort(key=lambda x: x[0])

                    if sz >= self.DUP_THRESHOLD:
                        self.size_groups[sz].append(entry.path)

                    self.scanned_files += 1
                    self.total_bytes += sz

                    now = time.time()
                    if now - self.last_progress > 0.12:
                        self.last_progress = now
                        self.result_queue.put(("progress", {
                            "files": self.scanned_files,
                            "bytes": self.total_bytes,
                            "current": entry.path,
                        }))
            except (PermissionError, OSError):
                self.denied_count += 1
                continue

        node.file_count = file_count
        node.folder_count = folder_count
        node.size = total
        node.allocated = allocated_total
        node.children.sort(key=lambda n: n.size, reverse=True)
        return total


# ---------------------------------------------------------------------------
# Duplicate worker (size groups -> SHA1 confirmation)
# ---------------------------------------------------------------------------

class DuplicateWorker(threading.Thread):
    _HEAD_HASH_BYTES = 65536  # 64 KB head-hash prefilter (Perf-5)

    @staticmethod
    def _head_hash(path):
        """Return a fast SHA1 of the first 64 KB. Files that disagree
        here can never be full-file duplicates."""
        try:
            import hashlib as _h
            h = _h.sha1()
            with open(path, "rb") as f:
                h.update(f.read(65536))
            return h.digest()
        except Exception:
            return None

    # Windows FILE_ATTRIBUTE bits that mark a cloud placeholder whose
    # bytes are NOT actually on disk (OneDrive, SharePoint, Dropbox
    # Smart Sync, etc). Hashing those would trigger a silent download.
    _ATTR_OFFLINE               = 0x00001000
    _ATTR_RECALL_ON_OPEN        = 0x00040000
    _ATTR_RECALL_ON_DATA_ACCESS = 0x00400000
    _PLACEHOLDER_MASK = (_ATTR_OFFLINE
                         | _ATTR_RECALL_ON_OPEN
                         | _ATTR_RECALL_ON_DATA_ACCESS)

    def __init__(self, size_groups, result_queue, stop_event):
        super().__init__(daemon=True)
        self.size_groups = size_groups
        self.result_queue = result_queue
        self.stop_event = stop_event
        self.skipped_cloud = 0

    @classmethod
    def _is_cloud_placeholder(cls, path):
        if sys.platform != "win32":
            return False
        try:
            from ctypes import windll
            attrs = windll.kernel32.GetFileAttributesW(path)
            if attrs == 0xFFFFFFFF:
                return False
            return bool(attrs & cls._PLACEHOLDER_MASK)
        except Exception:
            return False

    def run(self):
        candidates = [(sz, ps) for sz, ps in self.size_groups.items() if len(ps) > 1]
        candidates.sort(key=lambda x: -x[0])
        groups = []
        total_g = len(candidates)
        for i, (sz, paths) in enumerate(candidates):
            if self.stop_event.is_set():
                break
            by_hash = defaultdict(list)
            for p in paths:
                if self.stop_event.is_set():
                    break
                # Skip cloud placeholders to avoid triggering downloads
                if self._is_cloud_placeholder(p):
                    self.skipped_cloud += 1
                    continue
                h = self._hash(p)
                if h is None:
                    continue
                by_hash[h].append(p)
            for h, plist in by_hash.items():
                if len(plist) > 1:
                    groups.append((sz, h, plist))
            if (i + 1) % 4 == 0:
                self.result_queue.put(("dup_progress",
                                       {"current": i + 1, "total": total_g}))
        groups.sort(key=lambda x: -x[0] * (len(x[2]) - 1))
        self.result_queue.put(("dup_done",
                              {"groups": groups,
                               "skipped_cloud": self.skipped_cloud}))

    def _hash(self, path):
        try:
            h = hashlib.sha1()
            with open(path, "rb") as f:
                while True:
                    if self.stop_event.is_set():
                        return None
                    b = f.read(65536)
                    if not b:
                        break
                    h.update(b)
            return h.hexdigest()
        except (OSError, PermissionError):
            return None


# ---------------------------------------------------------------------------
# Charts (Canvas-based)
# ---------------------------------------------------------------------------

class BaseChart(tk.Canvas):
    def __init__(self, master, theme, **kw):
        kw.setdefault("highlightthickness", 0)
        kw.setdefault("bg", theme["panel"])
        super().__init__(master, **kw)
        self.theme = theme
        self._data = []
        self._tooltip = None
        self.bind("<Configure>", lambda e: self.redraw())
        self.bind("<Motion>", self._on_motion)
        self.bind("<Leave>", lambda e: self._hide_tip())

    def set_theme(self, theme):
        self.theme = theme
        self.configure(bg=theme["panel"])
        self.redraw()

    def set_data(self, data):
        self._data = data
        self.redraw()

    def redraw(self):
        self.delete("all")

    def _on_motion(self, event):
        pass

    def _show_tip(self, sx, sy, text):
        if self._tooltip is None:
            self._tooltip = tk.Toplevel(self)
            self._tooltip.wm_overrideredirect(True)
            self._tip_lbl = tk.Label(
                self._tooltip, text=text, bg="#222", fg="white",
                font=UI_FONT, padx=10, pady=6, justify="left",
            )
            self._tip_lbl.pack()
        else:
            self._tip_lbl.config(text=text)
        self._tooltip.geometry(f"+{sx+15}+{sy+15}")
        self._tooltip.deiconify()

    def _hide_tip(self):
        if self._tooltip:
            self._tooltip.withdraw()


class TreemapChart(BaseChart):
    def __init__(self, master, theme, on_click=None, **kw):
        super().__init__(master, theme, **kw)
        self._items = []
        self.on_click = on_click
        self.bind("<Button-1>", self._on_click)

    def redraw(self):
        self.delete("all")
        self._items = []
        nodes = self._data
        if not nodes:
            self._draw_empty()
            return
        w = max(self.winfo_width(), 50)
        h = max(self.winfo_height(), 50)
        total = sum(max(n.size, 1) for n in nodes)
        if total <= 0:
            return
        values = [max(n.size, 1) for n in nodes]
        self._squarify(nodes, values, 4, 4, w - 8, h - 8, total)

    def _draw_empty(self):
        self.create_text(self.winfo_width() / 2, self.winfo_height() / 2,
                         text="No data — run a scan",
                         fill=self.theme["muted"], font=UI_FONT)

    def _squarify(self, nodes, values, x, y, w, h, total):
        if not nodes or w <= 0 or h <= 0:
            return
        if len(nodes) == 1:
            self._draw_rect(nodes[0], x, y, w, h, 0)
            return
        area = w * h
        scaled = [v / total * area for v in values]
        cx, cy, cw, ch = x, y, w, h
        n = len(nodes)
        i = 0
        while i < n:
            row = [scaled[i]]
            j = i + 1
            best = self._worst(row, min(cw, ch))
            while j < n:
                new = row + [scaled[j]]
                nr = self._worst(new, min(cw, ch))
                if nr > best:
                    break
                row = new
                best = nr
                j += 1
            s = sum(row)
            if cw >= ch:
                sw = s / ch if ch else 0
                ry = cy
                for k, v in enumerate(row):
                    rh = v / sw if sw else 0
                    self._draw_rect(nodes[i + k], cx, ry, sw, rh, i + k)
                    ry += rh
                cx += sw
                cw -= sw
            else:
                sh = s / cw if cw else 0
                rx = cx
                for k, v in enumerate(row):
                    rw = v / sh if sh else 0
                    self._draw_rect(nodes[i + k], rx, cy, rw, sh, i + k)
                    rx += rw
                cy += sh
                ch -= sh
            i = j

    def _worst(self, row, side):
        if not row or side <= 0:
            return float("inf")
        s = sum(row)
        if s <= 0:
            return float("inf")
        return max((side * side * max(row)) / (s * s),
                   (s * s) / (side * side * min(row)))

    TILE_COLORS = [
        "#3B82F6",  # blue
        "#22C55E",  # green
        "#F97316",  # orange
        "#9333EA",  # purple
        "#EF4444",  # red
        "#14B8A6",  # teal
        "#C026D3",  # magenta
        "#F59E0B",  # amber
        "#475569",  # slate
        "#0EA5E9",  # sky
        "#65A30D",  # lime
        "#DB2777",  # rose
    ]

    # WinDirStat-style colour mapping by extension category.
    EXT_COLOR_MAP = {
        # Images
        ".jpg": "#10B981", ".jpeg": "#10B981", ".png": "#10B981",
        ".gif": "#10B981", ".bmp": "#10B981", ".webp": "#10B981",
        ".svg": "#10B981", ".tiff": "#10B981", ".heic": "#10B981",
        # Videos
        ".mp4": "#EF4444", ".mov": "#EF4444", ".avi": "#EF4444",
        ".mkv": "#EF4444", ".wmv": "#EF4444", ".webm": "#EF4444",
        ".m4v": "#EF4444",
        # Audio
        ".mp3": "#8B5CF6", ".wav": "#8B5CF6", ".flac": "#8B5CF6",
        ".aac": "#8B5CF6", ".ogg": "#8B5CF6", ".m4a": "#8B5CF6",
        # Documents
        ".pdf": "#3B82F6", ".doc": "#3B82F6", ".docx": "#3B82F6",
        ".xls": "#3B82F6", ".xlsx": "#3B82F6", ".ppt": "#3B82F6",
        ".pptx": "#3B82F6", ".txt": "#3B82F6", ".rtf": "#3B82F6",
        ".csv": "#3B82F6", ".md": "#3B82F6",
        # Archives
        ".zip": "#A855F7", ".rar": "#A855F7", ".7z": "#A855F7",
        ".tar": "#A855F7", ".gz": "#A855F7", ".iso": "#A855F7",
        # Executables
        ".exe": "#0EA5E9", ".msi": "#0EA5E9", ".dll": "#0EA5E9",
        ".sys": "#0EA5E9", ".bat": "#0EA5E9",
        # Code
        ".py": "#F59E0B", ".js": "#F59E0B", ".ts": "#F59E0B",
        ".html": "#F59E0B", ".css": "#F59E0B", ".json": "#F59E0B",
    }
    # color_mode: "by_depth" (default) or "by_extension"
    color_mode = "by_depth"

    @staticmethod
    def _tile_color(node, idx):
        if node.path == "(free)":
            return "#3B82F6"
        if node.path == "(gap)":
            return "#EF4444"
        if TreemapChart.color_mode == "by_extension" and not node.is_dir:
            import os as _os
            ext = _os.path.splitext(node.name)[1].lower()
            if ext in TreemapChart.EXT_COLOR_MAP:
                return TreemapChart.EXT_COLOR_MAP[ext]
            return "#94A3B8"  # slate-grey for "other"
        return TreemapChart.TILE_COLORS[idx % len(TreemapChart.TILE_COLORS)]

    @classmethod
    def set_color_mode(cls, mode):
        """Switch tile colouring between 'by_depth' and 'by_extension'."""
        cls.color_mode = mode if mode in ("by_depth", "by_extension") else "by_depth"

    @staticmethod
    def _tile_icon(node):
        if node.path == "(free)":
            return "🖴"
        if node.path == "(gap)":
            return "🔒"
        if node.is_dir:
            n = node.name.lower()
            if n.startswith("user"):
                return "👤"
            if n in ("windows", "winnt", "system32"):
                return "🪟"
            if n in ("recovery", "system volume information"):
                return "💾"
            if "program" in n:
                return "📁"
            if n.startswith("$"):
                return "🗑"
            return "📁"
        ext = os.path.splitext(node.name)[1].lower()
        if ext in (".jpg", ".jpeg", ".png", ".gif", ".webp", ".svg"):
            return "🖼"
        if ext in (".mp4", ".mov", ".avi", ".mkv", ".webm"):
            return "🎬"
        if ext in (".mp3", ".wav", ".flac"):
            return "🎵"
        if ext in (".pdf", ".doc", ".docx", ".xls", ".xlsx"):
            return "📄"
        if ext in (".zip", ".rar", ".7z"):
            return "📦"
        if ext in (".exe", ".msi", ".dll"):
            return "⚡"
        if ext == ".sys":
            return "🛠"
        return "📄"

    def _round_rect_points(self, x1, y1, x2, y2, r):
        return [
            x1 + r, y1, x2 - r, y1, x2, y1, x2, y1 + r,
            x2, y2 - r, x2, y2, x2 - r, y2, x1 + r, y2,
            x1, y2, x1, y2 - r, x1, y1 + r, x1, y1,
        ]

    @staticmethod
    def _lighten(hex_color, amount=0.4):
        """Blend a hex colour towards white by *amount* (0..1).

        Tkinter Canvas does not understand alpha channels, so we have to
        synthesise translucency by mixing colours up-front.
        """
        try:
            r = int(hex_color[1:3], 16)
            g = int(hex_color[3:5], 16)
            b = int(hex_color[5:7], 16)
        except (ValueError, IndexError):
            return "#FFFFFF"
        r = int(r + (255 - r) * amount)
        g = int(g + (255 - g) * amount)
        b = int(b + (255 - b) * amount)
        return f"#{r:02X}{g:02X}{b:02X}"

    @staticmethod
    def _trunc(text, max_chars):
        if len(text) <= max_chars:
            return text
        if max_chars <= 1:
            return "…"
        return text[:max_chars - 1] + "…"

    def _draw_rect(self, node, x, y, w, h, idx):
        if w < 1 or h < 1:
            return
        # Compute "% of parent" — total sum of all visible tiles.
        total = sum(max(n.size, 1) for n in self._data) or 1
        pct = node.size / total * 100
        color = self._tile_color(node, idx)

        # ---- Rounded tile (with a small inset gutter between cards) ----
        inset = 5
        r = max(6, min(18, h * 0.06, w * 0.06))
        x1, y1 = x + inset, y + inset
        x2, y2 = x + w - inset, y + h - inset
        iw = x2 - x1
        ih = y2 - y1
        rect = self.create_polygon(
            self._round_rect_points(x1, y1, x2, y2, r),
            smooth=True, fill=color, outline="")

        # ---- Decorative dot pattern (top-left, very subtle) ----
        if w > 120 and h > 90:
            for i in range(3):
                for j in range(3):
                    self.create_oval(
                        x1 + 10 + i * 10, y1 + 10 + j * 10,
                        x1 + 12 + i * 10, y1 + 12 + j * 10,
                        fill=self._lighten(color, 0.3), outline="")

        # ---- Pick a layout depending on tile size ----
        text_id = None
        icon = self._tile_icon(node)

        if iw >= 300 and ih >= 220:
            # BIG tile — centered icon (large) + stacked text
            cx = (x1 + x2) / 2
            cy_icon = y1 + ih * 0.30
            # Glass circle behind icon
            cr = min(48, ih * 0.12)
            self.create_oval(cx - cr, cy_icon - cr,
                             cx + cr, cy_icon + cr,
                             fill=self._lighten(color, 0.5), outline="")
            self.create_text(cx, cy_icon, text=icon,
                             font=("Segoe UI Emoji", int(cr * 0.95)),
                             fill="white")
            # Name (top of text block) — truncated to fit
            ty = cy_icon + cr + 26
            max_chars = max(10, int((iw - 30) / 9))
            self.create_text(cx, ty,
                             text=self._trunc(node.name, max_chars),
                             fill="white",
                             font=("Segoe UI Semibold", 16))
            # Size — huge
            self.create_text(cx, ty + 34, text=human_size(node.size),
                             fill="white",
                             font=("Segoe UI Semibold", 26))
            # Percent chip
            chip_w, chip_h = 70, 24
            cy_chip = ty + 80
            self.create_polygon(
                self._round_rect_points(cx - chip_w/2, cy_chip - chip_h/2,
                                        cx + chip_w/2, cy_chip + chip_h/2,
                                        chip_h/2),
                smooth=True, fill=self._lighten(color, 0.5), outline="")
            self.create_text(cx, cy_chip, text=f"{pct:.1f}%",
                             fill="white",
                             font=("Segoe UI Semibold", 11))
            # Inline progress bar
            bw_total = min(w * 0.55, 240)
            bx = cx - bw_total / 2
            by = cy_chip + 26
            self.create_polygon(
                self._round_rect_points(bx, by, bx + bw_total, by + 6, 3),
                smooth=True, fill=self._lighten(color, 0.3), outline="")
            fill_w = max(bw_total * pct / 100, 6)
            self.create_polygon(
                self._round_rect_points(bx, by, bx + fill_w, by + 6, 3),
                smooth=True, fill="#FFFFFF", outline="")
            text_id = rect
        elif iw >= 180 and ih >= 100 and iw >= ih * 1.15:
            # WIDE-MEDIUM tile — icon left, stacked text right
            cr = min(28, ih * 0.18)
            cx_icon = x1 + 14 + cr
            cy_icon = (y1 + y2) / 2
            self.create_oval(cx_icon - cr, cy_icon - cr,
                             cx_icon + cr, cy_icon + cr,
                             fill=self._lighten(color, 0.5), outline="")
            self.create_text(cx_icon, cy_icon, text=icon,
                             font=("Segoe UI Emoji", int(cr * 0.95)),
                             fill="white")
            tx = cx_icon + cr + 14
            ty = y1 + 14
            text_avail = max(iw - (tx - x1) - 12, 30)
            max_chars = max(8, int(text_avail / 7.2))
            self.create_text(tx, ty, anchor="nw",
                             text=self._trunc(node.name, max_chars),
                             fill="white",
                             font=("Segoe UI Semibold", 12))
            self.create_text(tx, ty + 24, anchor="nw",
                             text=human_size(node.size),
                             fill="white",
                             font=("Segoe UI Semibold", 17))
            # Percent only when tile is comfortably tall
            if ih >= 150:
                self.create_text(tx, ty + 58, anchor="nw",
                                 text=f"{pct:.1f}%",
                                 fill=self._lighten(color, 0.85),
                                 font=("Segoe UI", 10))
                bw_total = iw - 28
                bx = x1 + 14
                by = y2 - 12
                self.create_polygon(
                    self._round_rect_points(bx, by, bx + bw_total, by + 4, 2),
                    smooth=True, fill=self._lighten(color, 0.3), outline="")
                fill_w = max(bw_total * pct / 100, 4)
                self.create_polygon(
                    self._round_rect_points(bx, by, bx + fill_w, by + 4, 2),
                    smooth=True, fill="#FFFFFF", outline="")
            text_id = rect
        elif iw >= 110 and ih >= 130:
            # TALL-MEDIUM tile — icon on TOP, text below (centred)
            cx = (x1 + x2) / 2
            cr = min(22, ih * 0.13, iw * 0.18)
            cy_icon = y1 + cr + 14
            self.create_oval(cx - cr, cy_icon - cr,
                             cx + cr, cy_icon + cr,
                             fill=self._lighten(color, 0.5), outline="")
            self.create_text(cx, cy_icon, text=icon,
                             font=("Segoe UI Emoji", int(cr * 0.9)),
                             fill="white")
            max_chars = max(6, int((iw - 16) / 7))
            self.create_text(cx, cy_icon + cr + 14,
                             text=self._trunc(node.name, max_chars),
                             fill="white",
                             font=("Segoe UI Semibold", 10))
            self.create_text(cx, cy_icon + cr + 36,
                             text=human_size(node.size),
                             fill="white",
                             font=("Segoe UI Semibold", 12))
            # Percent only when tile is very tall — else it visually
            # collides with the size text above.
            if ih >= 210:
                self.create_text(cx, cy_icon + cr + 72,
                                 text=f"{pct:.1f}%",
                                 fill=self._lighten(color, 0.85),
                                 font=("Segoe UI", 9))
            text_id = rect
        elif iw >= 80 and ih >= 38:
            # SMALL tile — truncated name (no wrap) + size (if room)
            tx = x1 + 8
            ty = y1 + 6
            # Aggressive char cap so size text never runs off the edge
            max_chars = max(3, int((iw - 12) / 7.4))
            self.create_text(tx, ty, anchor="nw",
                             text=self._trunc(node.name, max_chars),
                             fill="white",
                             font=("Segoe UI Semibold", 9))
            # Only draw size when tile is wide AND tall enough to fit it
            # (14pt font ~ 65 px wide for "1024 MB" or "5.89 GB")
            if iw >= 120 and ih >= 46:
                self.create_text(tx, ty + 16, anchor="nw",
                                 text=human_size(node.size),
                                 fill="white",
                                 font=("Segoe UI Semibold", 10))
            text_id = rect
        # else: very small — no text, just the coloured tile

        self._items.append((rect, text_id, node, (x, y, x + w, y + h)))

    def _on_motion(self, event):
        for rect, t, node, bb in self._items:
            x1, y1, x2, y2 = bb
            if x1 <= event.x <= x2 and y1 <= event.y <= y2:
                self._show_tip(event.x_root, event.y_root,
                               f"{node.name}\n{node.path}\n{human_size(node.size)}")
                return
        self._hide_tip()

    def _on_click(self, event):
        for rect, t, node, bb in self._items:
            x1, y1, x2, y2 = bb
            if x1 <= event.x <= x2 and y1 <= event.y <= y2:
                if self.on_click:
                    self.on_click(node)
                return


class PieChart(BaseChart):
    def __init__(self, master, theme, **kw):
        super().__init__(master, theme, **kw)
        self._slices = []

    # Pastel chip tints matching the bar-chart palette
    CHIP_TINTS = [
        ("#3B82F6", "#DBEAFE"), ("#22C55E", "#DCFCE7"),
        ("#F97316", "#FFEDD5"), ("#9333EA", "#EDE9FE"),
        ("#EF4444", "#FEE2E2"), ("#14B8A6", "#CCFBF1"),
        ("#C026D3", "#FAE8FF"), ("#F59E0B", "#FEF3C7"),
        ("#475569", "#E2E8F0"), ("#0EA5E9", "#E0F2FE"),
        ("#65A30D", "#ECFCCB"), ("#DB2777", "#FCE7F3"),
    ]

    def _rrpts(self, x1, y1, x2, y2, r):
        return [
            x1 + r, y1, x2 - r, y1, x2, y1, x2, y1 + r,
            x2, y2 - r, x2, y2, x2 - r, y2, x1 + r, y2,
            x1, y2, x1, y2 - r, x1, y1 + r, x1, y1,
        ]

    def redraw(self):
        self.delete("all")
        self._slices = []
        data = self._data
        w = self.winfo_width()
        h = self.winfo_height()
        if not data:
            self.create_text(w / 2, h / 2, text="No data",
                             fill=self.theme["muted"], font=UI_FONT)
            return
        total = sum(max(v, 0) for _, v, *_ in data)
        if total <= 0:
            return
        # Donut on the left
        size = min(w * 0.42, h * 0.92)
        cx = w * 0.22
        cy = h / 2
        r = size / 2
        bbox = (cx - r, cy - r, cx + r, cy + r)
        start = 90.0
        for i, item in enumerate(data):
            label, value = item[0], item[1]
            base_color, _ = self.CHIP_TINTS[i % len(self.CHIP_TINTS)]
            color = item[2] if len(item) >= 3 else base_color
            extent = -value / total * 360.0
            self.create_arc(*bbox, start=start, extent=extent, fill=color,
                            outline=self.theme["panel"], width=2)
            self._slices.append((start, extent, label, value, color))
            start += extent
        ir = r * 0.58
        self.create_oval(cx - ir, cy - ir, cx + ir, cy + ir,
                         fill=self.theme["panel"], outline="")
        self.create_text(cx, cy - 8, text=human_size(total),
                         font=("Segoe UI Semibold", 18),
                         fill=self.theme["chart_text"])
        self.create_text(cx, cy + 16, text="Total",
                         font=UI_FONT, fill=self.theme["muted"])
        # Breakdown column on the right (with tinted % chips)
        lx = w * 0.50
        ly = h * 0.08
        self.create_text(lx, ly, anchor="nw", text="Breakdown",
                         font=("Segoe UI Semibold", 12),
                         fill=self.theme["chart_text"])
        line_y = ly + 28
        max_lines = int((h - line_y - 16) / 28)
        for i, item in enumerate(data[:max(max_lines, 1)]):
            label, value = item[0], item[1]
            base_color, tint = self.CHIP_TINTS[i % len(self.CHIP_TINTS)]
            pct = value / total * 100
            yy = line_y + i * 28
            # color dot
            self.create_oval(lx, yy + 4, lx + 12, yy + 16,
                             fill=base_color, outline="")
            # name
            short = label if len(label) <= 16 else label[:13] + "…"
            self.create_text(lx + 20, yy + 10, anchor="w",
                             text=short, font=UI_FONT,
                             fill=self.theme["chart_text"])
            # tinted % chip at fixed right offset
            chip_x1 = lx + 290
            chip_x2 = chip_x1 + 56
            # size right-anchored, with guaranteed 10px gap before chip
            self.create_text(chip_x1 - 10, yy + 10, anchor="e",
                             text=human_size(value),
                             font=UI_FONT_BOLD,
                             fill=self.theme["chart_text"])
            self.create_polygon(
                self._rrpts(chip_x1, yy + 2, chip_x2, yy + 18, 9),
                smooth=True, fill=tint, outline="")
            self.create_text((chip_x1 + chip_x2) / 2, yy + 10,
                             text=f"{pct:.1f}%", fill=base_color,
                             font=("Segoe UI Semibold", 9))

    def _on_motion(self, event):
        w = self.winfo_width()
        h = self.winfo_height()
        size = min(w * 0.55, h * 0.85)
        cx = w * 0.3
        cy = h / 2
        r = size / 2
        dx = event.x - cx
        dy = event.y - cy
        dist = math.hypot(dx, dy)
        if dist > r or dist < r * 0.55:
            self._hide_tip()
            return
        ang = math.degrees(math.atan2(-dy, dx))
        if ang < 0:
            ang += 360
        for start, extent, label, value, color in self._slices:
            end = start + extent
            lo, hi = (end, start) if extent < 0 else (start, end)
            lo, hi = lo % 360, hi % 360
            inside = (lo <= ang <= hi) if lo <= hi else (ang >= lo or ang <= hi)
            if inside:
                self._show_tip(event.x_root, event.y_root,
                               f"{label}\n{human_size(value)}")
                return
        self._hide_tip()


class BarChart(BaseChart):
    """Premium horizontal bar list (icon · name · bar · size · chip)."""

    # Pastel tints for percent-chip backgrounds, paired with TILE_COLORS.
    CHIP_TINTS = [
        ("#3B82F6", "#DBEAFE"),  # blue
        ("#22C55E", "#DCFCE7"),  # green
        ("#F97316", "#FFEDD5"),  # orange
        ("#9333EA", "#EDE9FE"),  # purple
        ("#EF4444", "#FEE2E2"),  # red
        ("#14B8A6", "#CCFBF1"),  # teal
        ("#C026D3", "#FAE8FF"),  # magenta
        ("#F59E0B", "#FEF3C7"),  # amber
        ("#475569", "#E2E8F0"),  # slate
        ("#0EA5E9", "#E0F2FE"),  # sky
        ("#65A30D", "#ECFCCB"),  # lime
        ("#DB2777", "#FCE7F3"),  # rose
    ]

    def __init__(self, master, theme, **kw):
        super().__init__(master, theme, **kw)
        self._bars = []

    def _rrpts(self, x1, y1, x2, y2, r):
        return [
            x1 + r, y1, x2 - r, y1, x2, y1, x2, y1 + r,
            x2, y2 - r, x2, y2, x2 - r, y2, x1 + r, y2,
            x1, y2, x1, y2 - r, x1, y1 + r, x1, y1,
        ]

    def redraw(self):
        self.delete("all")
        self._bars = []
        data = self._data
        if not data:
            self.create_text(self.winfo_width() / 2, self.winfo_height() / 2,
                             text="No data", fill=self.theme["muted"],
                             font=UI_FONT)
            return
        w = self.winfo_width()
        h = self.winfo_height()
        # Column widths
        icon_w  = 28
        name_w  = 180
        size_w  = 90
        chip_w  = 70
        gap     = 12
        margin_l = 18
        margin_r = 18
        margin_t = 12
        margin_b = 12
        bar_x1 = margin_l + icon_w + gap + name_w + gap
        bar_x2 = w - margin_r - chip_w - gap - size_w - gap
        bar_total = max(bar_x2 - bar_x1, 60)
        n = len(data)
        avail_h = max(h - margin_t - margin_b, 50)
        row_h = min(34, (avail_h - (n - 1) * 6) / max(n, 1))
        if row_h < 14:
            row_h = max(14, (avail_h - (n - 1) * 3) / n)
        max_val = max((v for _, v, *_ in data), default=1) or 1
        sum_total = sum(v for _, v, *_ in data) or 1
        for i, item in enumerate(data):
            label, value = item[0], item[1]
            base_color, tint = self.CHIP_TINTS[i % len(self.CHIP_TINTS)]
            y = margin_t + i * (row_h + 6)
            yc = y + row_h / 2
            # Icon (a small filled rounded square with a single emoji)
            icon = self._guess_icon(label)
            ix1 = margin_l
            ix2 = margin_l + icon_w
            self.create_polygon(
                self._rrpts(ix1, y + 4, ix2, y + row_h - 4, 6),
                smooth=True, fill=tint, outline="")
            self.create_text((ix1 + ix2) / 2, yc, text=icon,
                             fill=base_color,
                             font=("Segoe UI Emoji", 12))
            # Name
            short = label if len(label) <= 32 else label[:29] + "…"
            self.create_text(margin_l + icon_w + gap, yc, anchor="w",
                             text=short, font=UI_FONT,
                             fill=self.theme["fg"])
            # Track
            self.create_polygon(
                self._rrpts(bar_x1, yc - 5, bar_x1 + bar_total, yc + 5, 5),
                smooth=True, fill=self.theme["panel_alt"], outline="")
            # Fill
            bw = max(bar_total * value / max_val, 6)
            rect = self.create_polygon(
                self._rrpts(bar_x1, yc - 5, bar_x1 + bw, yc + 5, 5),
                smooth=True, fill=base_color, outline="")
            # Size text
            self.create_text(bar_x2 + gap, yc, anchor="w",
                             text=human_size(value), font=UI_FONT_BOLD,
                             fill=self.theme["fg"])
            # Percent chip (rounded pill, tinted bg)
            cx1 = w - margin_r - chip_w
            cx2 = w - margin_r
            cy1 = yc - 10
            cy2 = yc + 10
            self.create_polygon(
                self._rrpts(cx1, cy1, cx2, cy2, 10),
                smooth=True, fill=tint, outline="")
            pct = value / sum_total * 100
            self.create_text((cx1 + cx2) / 2, yc, text=f"{pct:.1f}%",
                             fill=base_color,
                             font=("Segoe UI Semibold", 9))
            self._bars.append(
                (rect, label, value,
                 (margin_l, y, w - margin_r, y + row_h)))

    @staticmethod
    def _guess_icon(name):
        n = name.lower()
        if n == "free space" or n.startswith("[free"):
            return "🖴"
        if "inaccessib" in n:
            return "🔒"
        if n.startswith("user"):
            return "👤"
        if n in ("windows", "winnt", "system32"):
            return "🪟"
        if "program" in n:
            return "📁"
        if "recovery" in n:
            return "💾"
        ext = os.path.splitext(name)[1].lower() if "." in name else ""
        if ext in (".sys", ".dll"):
            return "🛠"
        if ext in (".jpg", ".png", ".gif", ".webp"):
            return "🖼"
        if ext in (".mp4", ".mov", ".avi"):
            return "🎬"
        if ext in (".mp3", ".wav", ".flac"):
            return "🎵"
        if ext in (".zip", ".rar", ".7z"):
            return "📦"
        if ext in (".exe", ".msi"):
            return "⚡"
        if ext in (".log", ".tmp"):
            return "📝"
        if ext:
            return "📄"
        return "📁"

    def _on_motion(self, event):
        for rect, label, value, bb in self._bars:
            x1, y1, x2, y2 = bb
            if x1 <= event.x <= x2 and y1 <= event.y <= y2:
                self._show_tip(event.x_root, event.y_root,
                               f"{label}\n{human_size(value)}")
                return
        self._hide_tip()


class OverviewCanvas(BaseChart):
    """Stat cards + top-folder bars. Pure-canvas dashboard."""
    def set_payload(self, payload):
        self._data = payload
        self.redraw()

    def redraw(self):
        self.delete("all")
        if not self._data:
            self.create_text(self.winfo_width() / 2, self.winfo_height() / 2,
                             text="Select a folder and press Scan to begin.",
                             fill=self.theme["muted"], font=UI_FONT_BIG)
            return
        d = self._data
        t = self.theme
        w = self.winfo_width()
        h = self.winfo_height()
        pad = 20

        # ----- Disk-usage hero bar -----
        disk_total = d.get("disk_total") or 0
        disk_used  = d.get("disk_used") or 0
        disk_free  = d.get("disk_free") or 0
        scanned    = d.get("total_bytes") or 0
        gap_bytes  = max(disk_used - scanned, 0)

        hero_y = pad
        hero_h = 88
        self._draw_shadow_card(pad, hero_y, w - pad, hero_y + hero_h, 14)
        self.create_text(pad + 20, hero_y + 18, anchor="w",
                         text=f"Disk usage  ·  {d.get('root_path', '')}",
                         font=UI_FONT, fill=t["muted"])
        # Numbers
        self.create_text(pad + 20, hero_y + 44, anchor="w",
                         text=human_size(disk_used) if disk_total else
                              human_size(scanned),
                         font=("Segoe UI Semibold", 22), fill=t["fg"])
        if disk_total:
            self.create_text(pad + 20, hero_y + 70, anchor="w",
                             text=f"used of {human_size(disk_total)}  "
                                  f"·  free {human_size(disk_free)}",
                             font=UI_FONT, fill=t["muted"])
        # Segmented bar on the right
        bar_x1 = pad + 280
        bar_x2 = w - pad - 20
        bar_w  = max(bar_x2 - bar_x1, 80)
        bar_y  = hero_y + 36
        bar_h  = 22
        self._round_rect(bar_x1, bar_y, bar_x1 + bar_w, bar_y + bar_h,
                         bar_h / 2, fill=t["panel_alt"], outline="")
        # Layout: scanned (accent), gap (warn-ish), free (transparent)
        total_for_bar = disk_total if disk_total else max(scanned, 1)
        seg_scanned = (scanned / total_for_bar) * bar_w
        seg_gap     = (gap_bytes / total_for_bar) * bar_w
        cur = bar_x1
        if seg_scanned > 0:
            self._round_rect(cur, bar_y, cur + seg_scanned, bar_y + bar_h,
                             bar_h / 2, fill=t["accent"], outline="")
            cur += seg_scanned
        if seg_gap > 1:
            self.create_rectangle(cur, bar_y, cur + seg_gap, bar_y + bar_h,
                                  fill=t["warn"], outline="")
            cur += seg_gap
        # Legend dots
        leg_y = bar_y + bar_h + 6
        if disk_total:
            self._legend_dot(bar_x1,         leg_y, t["accent"],
                             f"Scanned {human_size(scanned)}", t)
            self._legend_dot(bar_x1 + 170,   leg_y, t["warn"],
                             f"Inaccessible {human_size(gap_bytes)}", t)
            self._legend_dot(bar_x1 + 360,   leg_y, t["panel_alt"],
                             f"Free {human_size(disk_free)}", t)

        # ----- Stat cards row -----
        cards_y = hero_y + hero_h + 18
        card_h = 96
        n_cards = 4
        gap = 14
        card_w = (w - pad * 2 - gap * (n_cards - 1)) / n_cards
        cards = [
            ("Files",        f"{d['total_files']:,}",       "#107C10", "📄"),
            ("Folders",      f"{d['folder_count']:,}",      "#D83B01", "📁"),
            ("File types",   f"{len(d['extensions'])}",     "#5C2D91", "🏷"),
            ("Inaccessible", f"{d.get('denied', 0):,}",     "#B7791F", "🔒"),
        ]
        for i, (label, value, color, icon) in enumerate(cards):
            cx = pad + i * (card_w + gap)
            cy = cards_y
            self._draw_shadow_card(cx, cy, cx + card_w, cy + card_h, 12)
            # Icon badge
            self._round_rect(cx + 14, cy + 14, cx + 48, cy + 48, 8,
                             fill=color, outline="")
            self.create_text(cx + 31, cy + 31, text=icon,
                             font=("Segoe UI Emoji", 14), fill="white")
            self.create_text(cx + 60, cy + 22, anchor="w", text=label,
                             font=UI_FONT, fill=t["muted"])
            self.create_text(cx + 60, cy + 50, anchor="w", text=value,
                             font=("Segoe UI Semibold", 18), fill=t["fg"])

        # ----- Top folders bar list -----
        top_y = cards_y + card_h + 22
        self.create_text(pad, top_y, anchor="nw", text="Top folders",
                         font=UI_FONT_BIG, fill=t["fg"])
        bar_top = top_y + 26
        avail_h = h - bar_top - pad
        items = d.get("top_folders", [])[:8]
        if not items:
            return
        max_v = max((v for _, v, _ in items), default=1) or 1
        row_h = min(32, (avail_h - (len(items) - 1) * 8) / max(len(items), 1))
        if row_h < 14:
            row_h = 14
        for i, (name, size, pct) in enumerate(items):
            y = bar_top + i * (row_h + 8)
            short = name if len(name) <= 30 else name[:27] + "…"
            self.create_text(pad, y + row_h / 2, anchor="w",
                             text=short, font=UI_FONT, fill=t["fg"])
            tx = pad + 220
            tw = w - tx - 140 - pad
            if tw < 40:
                tw = 40
            self._round_rect(tx, y, tx + tw, y + row_h, row_h / 2,
                             fill=t["panel_alt"], outline="")
            bw = max(tw * (size / max_v), 6)
            self._round_rect(tx, y, tx + bw, y + row_h, row_h / 2,
                             fill=PALETTE[i % len(PALETTE)], outline="")
            self.create_text(tx + tw + 10, y + row_h / 2, anchor="w",
                             text=f"{human_size(size)}  ({pct:.1f}%)",
                             font=UI_FONT_BOLD, fill=t["fg"])

    def _draw_shadow_card(self, x1, y1, x2, y2, r):
        """Soft drop-shadow + rounded card body."""
        t = self.theme
        # Fake shadow with 2 offset rounded rects of soft outline color
        shadow = "#D8DCE4" if t["panel"] == "#FFFFFF" else "#0A0C12"
        # Slight offset shadow
        self._round_rect(x1 + 2, y1 + 3, x2 + 2, y2 + 4, r,
                         fill=shadow, outline="")
        self._round_rect(x1, y1, x2, y2, r,
                         fill=t["panel"], outline=t["border"], width=1)

    def _legend_dot(self, x, y, color, text, t):
        self.create_oval(x, y + 2, x + 10, y + 12, fill=color, outline="")
        self.create_text(x + 16, y + 7, anchor="w", text=text,
                         font=UI_FONT, fill=t["muted"])

    def _round_rect(self, x1, y1, x2, y2, r, **kw):
        r = min(r, (x2 - x1) / 2, (y2 - y1) / 2)
        points = [
            x1 + r, y1,  x2 - r, y1,  x2, y1,  x2, y1 + r,
            x2, y2 - r,  x2, y2,  x2 - r, y2,  x1 + r, y2,
            x1, y2,  x1, y2 - r,  x1, y1 + r,  x1, y1,
        ]
        return self.create_polygon(points, smooth=True, **kw)


# ---------------------------------------------------------------------------
# Ribbon button (icon over label, hover effect)
# ---------------------------------------------------------------------------

class RibbonButton(tk.Frame):
    def __init__(self, master, theme, icon, label, command,
                 primary=False, **kw):
        super().__init__(master, bg=theme["ribbon_bg"], **kw)
        self.theme = theme
        self.command = command
        self.primary = primary
        self.enabled = True
        self.icon_lbl = tk.Label(self, text=icon, bg=theme["ribbon_bg"],
                                 fg=theme["accent"] if primary else theme["fg"],
                                 font=ICON_FONT, cursor="hand2")
        self.text_lbl = tk.Label(self, text=label, bg=theme["ribbon_bg"],
                                 fg=theme["fg"], font=UI_FONT, cursor="hand2")
        self.icon_lbl.pack(pady=(6, 0), padx=10)
        self.text_lbl.pack(pady=(2, 6), padx=10)
        for w in (self, self.icon_lbl, self.text_lbl):
            w.bind("<Enter>", self._on_enter)
            w.bind("<Leave>", self._on_leave)
            w.bind("<Button-1>", self._on_click)

    def _on_enter(self, _):
        if not self.enabled:
            return
        bg = self.theme["panel_hover"]
        self.configure(bg=bg)
        self.icon_lbl.configure(bg=bg)
        self.text_lbl.configure(bg=bg)

    def _on_leave(self, _):
        bg = self.theme["ribbon_bg"]
        self.configure(bg=bg)
        self.icon_lbl.configure(bg=bg)
        self.text_lbl.configure(bg=bg)

    def _on_click(self, _):
        if self.enabled and self.command:
            self.command()

    def set_enabled(self, enabled):
        self.enabled = enabled
        fg = self.theme["fg"] if enabled else self.theme["muted"]
        icon_fg = (self.theme["accent"] if (self.primary and enabled) else fg)
        self.icon_lbl.configure(fg=icon_fg)
        self.text_lbl.configure(fg=fg)

    def apply_theme(self, theme):
        self.theme = theme
        bg = theme["ribbon_bg"]
        self.configure(bg=bg)
        self.icon_lbl.configure(bg=bg, fg=theme["accent"] if self.primary else theme["fg"])
        self.text_lbl.configure(bg=bg, fg=theme["fg"])



# ---------------------------------------------------------------------------
# Sidebar nav item
# ---------------------------------------------------------------------------

class SidebarItem(tk.Frame):
    def __init__(self, master, theme, icon, label, command, has_submenu=False):
        self.theme = theme
        self.command = command
        self.has_submenu = has_submenu
        self.active = False
        super().__init__(master, bg=theme["sb_bg"])
        self._build(icon, label)

    def _build(self, icon, label):
        t = self.theme
        self.inner = tk.Frame(self, bg=t["sb_bg"])
        self.inner.pack(fill="x", padx=10, pady=2)
        self.icon_lbl = tk.Label(self.inner, text=icon, bg=t["sb_bg"],
                                 fg=t["sb_fg"],
                                 font=("Segoe UI Emoji", 13), cursor="hand2")
        self.icon_lbl.pack(side="left", padx=(12, 12), pady=8)
        self.text_lbl = tk.Label(self.inner, text=label, bg=t["sb_bg"],
                                 fg=t["sb_fg"], font=("Segoe UI", 10),
                                 cursor="hand2")
        self.text_lbl.pack(side="left", pady=10)
        if self.has_submenu:
            self.chev = tk.Label(self.inner, text="›", bg=t["sb_bg"],
                                 fg=t["sb_fg"], font=("Segoe UI", 11),
                                 cursor="hand2")
            self.chev.pack(side="right", padx=(0, 14))
        for w in (self, self.inner, self.icon_lbl, self.text_lbl):
            w.bind("<Enter>", self._on_enter)
            w.bind("<Leave>", self._on_leave)
            w.bind("<Button-1>", self._on_click)

    def _on_enter(self, _):
        if self.active:
            return
        bg = self.theme["sb_bg_hover"]
        self.inner.configure(bg=bg)
        for w in (self.icon_lbl, self.text_lbl):
            w.configure(bg=bg)
        if self.has_submenu:
            self.chev.configure(bg=bg)

    def _on_leave(self, _):
        if self.active:
            return
        bg = self.theme["sb_bg"]
        self.inner.configure(bg=bg)
        for w in (self.icon_lbl, self.text_lbl):
            w.configure(bg=bg)
        if self.has_submenu:
            self.chev.configure(bg=bg)

    def _on_click(self, _):
        if self.command:
            self.command()

    def set_active(self, on):
        self.active = on
        t = self.theme
        bg = t["sb_bg_active"] if on else t["sb_bg"]
        fg = t["sb_fg_active"] if on else t["sb_fg"]
        self.inner.configure(bg=bg)
        self.icon_lbl.configure(bg=bg, fg=fg)
        self.text_lbl.configure(bg=bg, fg=fg)
        if self.has_submenu:
            self.chev.configure(bg=bg, fg=fg)



class IconButton(tk.Frame):
    """Header pill button: icon + text. Variants: primary, ghost."""
    def __init__(self, master, theme, icon, label, command, variant="ghost"):
        self.theme = theme
        self.variant = variant
        self.command = command
        bg = theme["accent"] if variant == "primary" else theme["panel"]
        fg = "white" if variant == "primary" else theme["fg"]
        super().__init__(master, bg=bg)
        self._bg = bg
        self._fg = fg
        pad_x_ic = (18, 8) if variant == "primary" else (14, 6)
        pad_x_tx = (0, 18)  if variant == "primary" else (0, 14)
        pad_y    = 10       if variant == "primary" else 8
        icon_font = ("Segoe UI Emoji", 13) if variant == "primary" \
                    else ("Segoe UI Emoji", 11)
        text_font = ("Segoe UI Semibold", 10) if variant == "primary" \
                    else UI_FONT
        self.icon_lbl = tk.Label(self, text=icon, bg=bg, fg=fg,
                                 font=icon_font, cursor="hand2")
        self.icon_lbl.pack(side="left", padx=pad_x_ic, pady=pad_y)
        self.text_lbl = tk.Label(self, text=label, bg=bg, fg=fg,
                                 font=text_font, cursor="hand2")
        self.text_lbl.pack(side="left", padx=pad_x_tx, pady=pad_y)
        for w in (self, self.icon_lbl, self.text_lbl):
            w.bind("<Enter>", self._on_enter)
            w.bind("<Leave>", self._on_leave)
            w.bind("<Button-1>", self._on_click)

    def _on_enter(self, _):
        bg = self.theme["accent_dark"] if self.variant == "primary" \
             else self.theme["panel_hover"]
        self.configure(bg=bg)
        self.icon_lbl.configure(bg=bg)
        self.text_lbl.configure(bg=bg)

    def _on_leave(self, _):
        self.configure(bg=self._bg)
        self.icon_lbl.configure(bg=self._bg)
        self.text_lbl.configure(bg=self._bg)

    def _on_click(self, _):
        if self.command:
            self.command()


class GhostIconButton(tk.Label):
    """Small icon-only ghost button (theme toggle, kebab)."""
    def __init__(self, master, theme, icon, command):
        self.theme = theme
        self.command = command
        super().__init__(master, text=icon, bg=theme["panel"],
                         fg=theme["fg"], font=("Segoe UI Emoji", 13),
                         cursor="hand2", padx=10, pady=6)
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)
        self.bind("<Button-1>", lambda e: command() if command else None)

    def _on_enter(self, _):
        self.configure(bg=self.theme["panel_hover"])

    def _on_leave(self, _):
        self.configure(bg=self.theme["panel"])


# ---------------------------------------------------------------------------
# Main application — dashboard layout matching the mock
# ---------------------------------------------------------------------------

class CapacitraApp:
    NAV_ITEMS = [
        ("overview",   "🏠", "Overview",    False),
        ("treemap",    "🗺", "Treemap",     False),
        ("charts",     "📊", "Charts",      True),
        ("duplicates", "⎘",  "Duplicates",  False),
        ("large",      "📦", "Large Files", False),
        ("types",      "🏷", "File Types",  False),
        ("export",     "📤", "Export",      False),
        ("settings",   "⚙",  "Settings",    False),
        ("history",    "🕓", "History",     False),
        ("about",      "ⓘ",  "About",       False),
    ]

    def __init__(self, root):
        self.root = root
        self._alive = True   # cleared by _on_close so _poll_queue stops
        self.theme_name = "light"
        self.theme = THEMES[self.theme_name]
        root.title(f"{APP_NAME}  ·  v{APP_VERSION}")
        # Size the window to fit the screen — never larger than the
        # available workspace (leaves room for the Windows taskbar).
        try:
            sw = root.winfo_screenwidth()
            sh = root.winfo_screenheight()
        except Exception:
            sw, sh = 1440, 900
        # Reserve ~60px for the taskbar and a small offset from the top.
        target_w = min(1440, max(1100, sw - 80))
        target_h = min(900, max(680, sh - 100))
        pos_x = max(0, (sw - target_w) // 2)
        pos_y = max(0, (sh - target_h) // 2 - 20)
        root.geometry(f"{target_w}x{target_h}+{pos_x}+{pos_y}")
        root.minsize(1000, 640)
        # On Windows, start maximised so the user gets the whole workspace
        # but the taskbar is respected by the OS.
        if sys.platform == "win32":
            try:
                root.state("zoomed")
            except Exception:
                pass
        root.configure(bg=self.theme["bg"])
        # Intercept the close button so a long-running scan or duplicate
        # search isn't killed mid-flight without confirmation.
        try:
            root.protocol("WM_DELETE_WINDOW", self._on_close)
        except Exception:
            pass

        self.scan_thread = None
        self.dup_thread = None
        self.stop_event = threading.Event()
        self.dup_stop = threading.Event()
        self.result_queue = queue.Queue()
        self.scan_root = None
        self.scan_result = None
        self._dup_result = None
        self._downloads_hint = None
        self._cheat_win = None
        self.current_node = None
        self._overview_payload = None
        self._tree_node_map = {}
        self._node_iid_map = {}
        self._expanded_iids = set()
        self._sort_state = {}
        self._suppress_select = False
        self._nav_items = {}
        self._panels = {}
        self._active_panel = None

        # User-editable exclude list (folder names skipped at scan time)
        self._excludes = ["$Recycle.Bin", "System Volume Information"]
        # Advanced-filter parsed state
        self._adv = {"name": None, "min": 0, "max": float("inf"),
                     "age_max_days": None, "ext": None}

        self._setup_style()
        self._build_layout()
        self._populate_drives()
        self._apply_theme()
        self._select_panel("overview")
        self._bind_shortcuts()
        root.after(80, self._poll_queue)


    # ----- Keyboard cheatsheet overlay (v4.2) -----
    def _show_cheatsheet(self):
        """Full-screen ? overlay showing all keyboard shortcuts."""
        if getattr(self, "_cheat_win", None) is not None:
            try:
                self._cheat_win.destroy()
            except Exception:
                pass
            self._cheat_win = None
            return
        t = self.theme
        win = tk.Toplevel(self.root)
        self._cheat_win = win
        win.title("Keyboard shortcuts")
        win.configure(bg=t["panel"])
        win.transient(self.root)
        # Center over the main window
        self.root.update_idletasks()
        w, h = 640, 520
        x = self.root.winfo_rootx() + (self.root.winfo_width() - w) // 2
        y = self.root.winfo_rooty() + (self.root.winfo_height() - h) // 2
        win.geometry(f"{w}x{h}+{max(0,x)}+{max(0,y)}")
        win.minsize(520, 440)
        win.resizable(True, True)

        outer = tk.Frame(win, bg=t["panel"])
        outer.pack(fill="both", expand=True, padx=28, pady=24)
        tk.Label(outer, text="Keyboard shortcuts", bg=t["panel"],
                 fg=t["fg"], font=("Segoe UI Semibold", 18)).pack(anchor="w")
        tk.Label(outer, text="Press ? or Esc to close",
                 bg=t["panel"], fg=t["muted"],
                 font=UI_FONT).pack(anchor="w", pady=(2, 18))

        groups = [
            ("Scan", [
                ("Ctrl+O",       "Browse and scan folder"),
                ("F5",           "Rescan current root"),
                ("Esc",          "Cancel running scan"),
            ]),
            ("Search", [
                ("Ctrl+F",       "Focus filter box"),
                ("Ctrl+Shift+F", "Global find dialog"),
            ]),
            ("Export", [
                ("Ctrl+E",       "Open export menu"),
            ]),
            ("Files", [
                ("Del",          "Move selected to Recycle Bin"),
                ("Enter",        "Open selected file"),
            ]),
            ("Help", [
                ("?",            "Toggle this cheatsheet"),
            ]),
        ]

        for title, rows in groups:
            tk.Label(outer, text=title, bg=t["panel"], fg=t["accent"],
                     font=("Segoe UI Semibold", 11)).pack(
                         anchor="w", pady=(10, 4))
            grid = tk.Frame(outer, bg=t["panel"])
            grid.pack(fill="x", padx=6)
            for i, (keys, desc) in enumerate(rows):
                key_lbl = tk.Label(grid, text=keys, bg=t["panel_alt"],
                                   fg=t["fg"], font=("Consolas", 10, "bold"),
                                   padx=10, pady=4)
                key_lbl.grid(row=i, column=0, sticky="w", pady=2)
                tk.Label(grid, text=desc, bg=t["panel"], fg=t["fg_subtle"],
                         font=UI_FONT).grid(row=i, column=1, sticky="w",
                                            padx=(14, 0), pady=2)

        def _close(_=None):
            try:
                win.destroy()
            except Exception:
                pass
            self._cheat_win = None

        win.bind("<Escape>", _close)
        win.bind("<Key-question>", _close)
        win.protocol("WM_DELETE_WINDOW", _close)
        win.focus_set()


    # =========================================================
    # v4.2 Sprint A: Empty folders / Suspicious exes / Downloads
    # =========================================================
    _SUSPICIOUS_EXTENSIONS = {
        ".exe", ".dll", ".bat", ".ps1", ".vbs", ".scr", ".cmd", ".msi",
        ".jar", ".hta", ".wsf",
    }
    _SUSPICIOUS_SAFE_PREFIXES = None  # populated lazily

    def _safe_path_prefixes(self):
        """Windows path prefixes that we treat as 'known-good' (i.e. not
        surfaced by the suspicious executable finder)."""
        if self._SUSPICIOUS_SAFE_PREFIXES is None:
            prefixes = []
            for env in ("ProgramFiles", "ProgramFiles(x86)", "ProgramW6432",
                        "windir", "SystemRoot"):
                v = os.environ.get(env)
                if v:
                    prefixes.append(v.rstrip("\\").lower())
            # Common Microsoft-owned data trees
            for extra in (r"C:\ProgramData\Microsoft",
                          r"C:\ProgramData\Package Cache",
                          r"C:\Windows",
                          r"C:\Program Files",
                          r"C:\Program Files (x86)"):
                prefixes.append(extra.lower())
            self.__class__._SUSPICIOUS_SAFE_PREFIXES = list(set(prefixes))
        return self._SUSPICIOUS_SAFE_PREFIXES

    def _is_safe_path(self, p):
        p = (p or "").lower()
        for pref in self._safe_path_prefixes():
            if p.startswith(pref):
                return True
        return False

    # ----- Empty folder finder ---------------------------------------
    def _find_empty_folders(self):
        if not self.scan_result:
            self._warn("Run a scan first",
                       "Pick a folder or drive and click New Scan, "
                       "then try again.")
            return
        empties = []
        def _walk(node):
            if not node.is_dir:
                return
            # A folder counts as empty when it has no children at all
            # (recursively descending into subdirs that themselves are
            # empty). We only surface the outermost empty folder so the
            # user can Recycle just it, not every leaf below.
            has_content = any(
                (not c.is_dir) or (c.size or 0) > 0 or c.children
                for c in node.children
            )
            if not has_content and node.parent is not None:
                empties.append(node.path)
                return
            for c in node.children:
                _walk(c)
        try:
            _walk(self.scan_result["root"])
        except RecursionError:
            self._error("Recursion limit",
                        "Folder tree is too deep to walk fully.")
            return

        if not empties:
            self._info("No empty folders",
                       "Nothing to clean up. Every folder in the scan "
                       "contains at least one non-empty child.")
            return

        # Show a simple dialog with a listbox and "Move to Recycle Bin"
        t = self.theme
        win = tk.Toplevel(self.root)
        win.title(f"Empty folders — {len(empties)} found")
        win.configure(bg=t["panel"])
        win.transient(self.root)
        win.minsize(560, 420)
        win.geometry("720x520")

        tk.Label(win, text=f"Empty folders — {len(empties):,} found",
                 bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 15)).pack(
                     anchor="w", padx=24, pady=(20, 4))
        tk.Label(win,
                 text="Select rows to move to the Recycle Bin. "
                      "Nothing is hard-deleted.",
                 bg=t["panel"], fg=t["muted"],
                 font=UI_FONT, wraplength=640, justify="left").pack(
                     anchor="w", padx=24, pady=(0, 12))

        list_frame = tk.Frame(win, bg=t["panel"])
        list_frame.pack(fill="both", expand=True, padx=24, pady=(0, 12))
        lb = tk.Listbox(list_frame, selectmode="extended",
                        bg=t["panel_alt"], fg=t["fg"],
                        selectbackground=t["accent"],
                        selectforeground="#FFFFFF",
                        font=UI_FONT, activestyle="none",
                        highlightthickness=0, borderwidth=0)
        vs = ttk.Scrollbar(list_frame, orient="vertical", command=lb.yview)
        lb.configure(yscrollcommand=vs.set)
        lb.pack(side="left", fill="both", expand=True)
        vs.pack(side="right", fill="y")
        for p in empties:
            lb.insert("end", p)

        btn_bar = tk.Frame(win, bg=t["panel"])
        btn_bar.pack(fill="x", padx=24, pady=(0, 20))

        def _select_all():
            lb.selection_set(0, "end")

        def _recycle_selected():
            idxs = lb.curselection()
            if not idxs:
                if self._ask(
                        "Recycle all?",
                        f"No rows selected. Move all {len(empties)} "
                        "empty folders to the Recycle Bin?"):
                    idxs = tuple(range(lb.size()))
                else:
                    return
            paths = [lb.get(i) for i in idxs]
            if not self._ask(
                    "Move to Recycle Bin",
                    f"Move {len(paths)} folder(s) to the Recycle Bin?\n"
                    "Nothing is hard-deleted."):
                return
            ok = 0
            for p in paths:
                try:
                    if send_to_recycle_bin(p):
                        ok += 1
                except Exception:
                    pass
            self._info("Done",
                       f"Moved {ok}/{len(paths)} to Recycle Bin.")
            try:
                self._refresh_after_recycle(paths)
            except Exception:
                pass
            try:
                win.destroy()
            except Exception:
                pass

        tk.Button(btn_bar, text="Select all", command=_select_all,
                  bg=t["panel_alt"], fg=t["fg"], relief="flat",
                  padx=14, pady=6).pack(side="left")
        tk.Button(btn_bar, text="Move selected to Recycle Bin",
                  command=_recycle_selected,
                  bg=t["accent"], fg="#FFFFFF", relief="flat",
                  padx=16, pady=6, font=UI_FONT_BOLD).pack(side="right")
        tk.Button(btn_bar, text="Close", command=win.destroy,
                  bg=t["panel_alt"], fg=t["fg"], relief="flat",
                  padx=14, pady=6).pack(side="right", padx=(0, 8))

    # ----- Suspicious executable finder -------------------------------
    def _show_suspicious_exes(self):
        if not self.scan_result:
            self._warn("Run a scan first",
                       "Pick a folder or drive and click New Scan.")
            return
        hits = []
        exts = self._SUSPICIOUS_EXTENSIONS

        def _walk(node):
            if node.is_dir:
                for c in node.children:
                    _walk(c)
                return
            p = node.path
            if not p or p.startswith("("):
                return
            ext = os.path.splitext(p)[1].lower()
            if ext not in exts:
                return
            if self._is_safe_path(p):
                return
            hits.append((node.mtime or 0, node.size or 0, p, ext))

        try:
            _walk(self.scan_result["root"])
        except RecursionError:
            self._error("Recursion limit",
                        "Folder tree is too deep to walk fully.")
            return

        hits.sort(reverse=True)  # newest first

        if not hits:
            self._info(
                "No suspicious executables",
                "Every .exe / .dll / .bat / .ps1 / .vbs found is inside "
                "Program Files, Windows or ProgramData. That's the "
                "normal state on a clean system.")
            return

        t = self.theme
        win = tk.Toplevel(self.root)
        win.title(f"Suspicious executables — {len(hits)} found")
        win.configure(bg=t["panel"])
        win.transient(self.root)
        win.minsize(720, 480)
        win.geometry("960x600")

        tk.Label(win,
                 text=f"Executables outside standard locations — "
                      f"{len(hits):,} found",
                 bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 15)).pack(
                     anchor="w", padx=24, pady=(20, 4))
        tk.Label(win,
                 text="Sorted by modification time, newest first. "
                      "A file appearing here is not necessarily malicious, "
                      "but every one deserves a second look on unfamiliar "
                      "systems.",
                 bg=t["panel"], fg=t["muted"], font=UI_FONT,
                 justify="left", wraplength=880).pack(
                     anchor="w", padx=24, pady=(0, 12))

        tree_frame = tk.Frame(win, bg=t["panel"])
        tree_frame.pack(fill="both", expand=True, padx=24, pady=(0, 12))
        tv = ttk.Treeview(tree_frame,
                          columns=("modified", "size", "type"),
                          show="tree headings")
        tv.heading("#0", text="Path")
        tv.heading("modified", text="Modified")
        tv.heading("size", text="Size")
        tv.heading("type", text="Type")
        tv.column("#0", width=560, anchor="w")
        tv.column("modified", width=160, anchor="w")
        tv.column("size", width=100, anchor="e")
        tv.column("type", width=80, anchor="w")
        vs = ttk.Scrollbar(tree_frame, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=vs.set)
        tv.pack(side="left", fill="both", expand=True)
        vs.pack(side="right", fill="y")

        for mt, sz, p, ext in hits[:5000]:  # cap for UI perf
            when = (time.strftime("%Y-%m-%d %H:%M", time.localtime(mt))
                    if mt else "")
            tv.insert("", "end", text=p,
                      values=(when, human_size(sz), ext))

        btn_bar = tk.Frame(win, bg=t["panel"])
        btn_bar.pack(fill="x", padx=24, pady=(0, 20))

        def _reveal():
            sel = tv.selection()
            if not sel:
                return
            p = tv.item(sel[0], "text")
            try:
                subprocess.Popen(["explorer", "/select,", p])
            except Exception:
                pass

        def _copy():
            sel = tv.selection()
            if not sel:
                return
            p = tv.item(sel[0], "text")
            try:
                self.root.clipboard_clear()
                self.root.clipboard_append(p)
            except Exception:
                pass

        tk.Button(btn_bar, text="Show in Explorer", command=_reveal,
                  bg=t["panel_alt"], fg=t["fg"], relief="flat",
                  padx=14, pady=6).pack(side="left")
        tk.Button(btn_bar, text="Copy path", command=_copy,
                  bg=t["panel_alt"], fg=t["fg"], relief="flat",
                  padx=14, pady=6).pack(side="left", padx=(8, 0))
        tk.Button(btn_bar, text="Close", command=win.destroy,
                  bg=t["accent"], fg="#FFFFFF", relief="flat",
                  padx=16, pady=6, font=UI_FONT_BOLD).pack(side="right")

    # ----- Downloads folder aging (post-scan analysis) ---------------
    _DL_AGE_DAYS = 30
    _DL_AGE_MIN_BYTES = 100 * 1024 * 1024  # 100 MB threshold

    def _analyze_downloads_aging(self):
        """Look for a Downloads folder in the scan result and compute
        aged (>30 days) bytes. Returns dict or None."""
        if not self.scan_result:
            return None
        root = self.scan_result.get("root")
        if root is None:
            return None
        target = None
        # Prefer %USERPROFILE%\Downloads if it lives inside the scan
        userprofile = os.environ.get("USERPROFILE", "")
        expected = (os.path.join(userprofile, "Downloads").lower()
            if userprofile else None)

        def _find(node):
            nonlocal target
            if target is not None:
                return
            if node.is_dir:
                p = (node.path or "").lower()
                nm = (node.name or "").lower()
                if expected and p == expected:
                    target = node
                    return
                if nm == "downloads" and target is None:
                    target = node  # tentative, keep looking for exact match
                for c in node.children:
                    _find(c)
        try:
            _find(root)
        except RecursionError:
            return None

        if target is None:
            return None

        now = time.time()
        thr = self._DL_AGE_DAYS * 86400
        aged_bytes = 0
        aged_count = 0

        def _walk(n):
            nonlocal aged_bytes, aged_count
            if n.is_dir:
                for c in n.children:
                    _walk(c)
                return
            if n.mtime and (now - n.mtime) > thr:
                aged_bytes += (n.size or 0)
                aged_count += 1

        try:
            _walk(target)
        except RecursionError:
            return None

        if aged_bytes < self._DL_AGE_MIN_BYTES:
            return None
        return {"path": target.path, "bytes": aged_bytes,
                "count": aged_count, "days": self._DL_AGE_DAYS}


    def _show_downloads_hint(self):
        info = getattr(self, "_downloads_hint", None)
        if not info:
            self._info(
                "No aged Downloads content",
                "Your Downloads folder either wasn't in the scan, "
                "or has less than 100 MB of files older than "
                f"{self._DL_AGE_DAYS} days.")
            return
        t = self.theme
        win = tk.Toplevel(self.root)
        win.title("Downloads folder aging")
        win.configure(bg=t["panel"])
        win.transient(self.root)
        win.minsize(560, 320)
        win.geometry("640x360")
        tk.Label(win, text="Downloads folder aging",
                 bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 15)).pack(
                     anchor="w", padx=24, pady=(20, 4))
        tk.Label(win,
                 text=f"{info['path']}",
                 bg=t["panel"], fg=t["muted"], font=UI_FONT,
                 wraplength=500).pack(anchor="w", padx=24, pady=(0, 14))
        tk.Label(win,
                 text=human_size(info["bytes"]),
                 bg=t["panel"], fg=t["accent"],
                 font=("Segoe UI Semibold", 28)).pack(anchor="w", padx=24)
        tk.Label(win,
                 text=f"in {info['count']:,} files older than "
                      f"{info['days']} days.",
                 bg=t["panel"], fg=t["fg_subtle"], font=UI_FONT).pack(
                     anchor="w", padx=24, pady=(0, 18))
        btn = tk.Frame(win, bg=t["panel"])
        btn.pack(fill="x", padx=24, pady=(0, 20))
        def _open_folder():
            try:
                open_path(info["path"])
            except Exception:
                pass
        tk.Button(btn, text="Open in Explorer", command=_open_folder,
                  bg=t["accent"], fg="#FFFFFF", relief="flat",
                  padx=16, pady=6, font=UI_FONT_BOLD).pack(side="left")
        tk.Button(btn, text="Close", command=win.destroy,
                  bg=t["panel_alt"], fg=t["fg"], relief="flat",
                  padx=14, pady=6).pack(side="right")

    # ----- Duplicate cluster viewer ----------------------------------
    def _show_duplicate_clusters(self):
        """Group duplicate SHA-1 groups by parent folder and list total
        wasted bytes per folder. Actionable view."""
        dups = getattr(self, "_dup_result", None) or []
        if not dups:
            self._info(
                "Run duplicate scan first",
                "Open the Duplicates panel and click Find duplicates, "
                "then come back here.")
            return
        by_folder = {}
        # Each entry from _on_dup_done is (size, hash, paths)
        for entry in dups:
            try:
                file_sz, _sha, paths = entry[0], entry[1], entry[2]
            except (IndexError, TypeError):
                continue
            if not paths or len(paths) < 2:
                continue
            for p in paths:
                fold = os.path.dirname(p)
                if not fold:
                    continue
                b = by_folder.setdefault(fold, {"wasted": 0, "groups": 0})
                # Attribute the waste to this folder proportional to how
                # many copies of the group live here.
                b["wasted"] += file_sz
                b["groups"] += 1
        # Sort by wasted desc
        rows = sorted(by_folder.items(),
                      key=lambda kv: kv[1]["wasted"], reverse=True)
        if not rows:
            self._info("No duplicate clusters",
                       "No folder groups host duplicate content.")
            return

        t = self.theme
        win = tk.Toplevel(self.root)
        win.title(f"Duplicate clusters — {len(rows)} folders with duplicates")
        win.configure(bg=t["panel"])
        win.transient(self.root)
        win.minsize(720, 480)
        win.geometry("880x560")

        tk.Label(win, text="Duplicate clusters by folder",
                 bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 15)).pack(
                     anchor="w", padx=24, pady=(20, 4))
        tk.Label(win,
                 text="Which folders host the most wasted space? "
                      "The top rows are usually Downloads / backup dirs.",
                 bg=t["panel"], fg=t["muted"], font=UI_FONT,
                 wraplength=760, justify="left").pack(
                     anchor="w", padx=24, pady=(0, 12))

        tree_frame = tk.Frame(win, bg=t["panel"])
        tree_frame.pack(fill="both", expand=True, padx=24, pady=(0, 12))
        tv = ttk.Treeview(tree_frame,
                          columns=("wasted", "groups"),
                          show="tree headings")
        tv.heading("#0", text="Folder")
        tv.heading("wasted", text="Wasted bytes")
        tv.heading("groups", text="Groups")
        tv.column("#0", width=520, anchor="w")
        tv.column("wasted", width=140, anchor="e")
        tv.column("groups", width=90, anchor="e")
        vs = ttk.Scrollbar(tree_frame, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=vs.set)
        tv.pack(side="left", fill="both", expand=True)
        vs.pack(side="right", fill="y")
        for path, info in rows[:1000]:
            tv.insert("", "end", text=path,
                      values=(human_size(info["wasted"]),
                              f"{info['groups']:,}"))

        btn_bar = tk.Frame(win, bg=t["panel"])
        btn_bar.pack(fill="x", padx=24, pady=(0, 20))
        tk.Button(btn_bar, text="Close", command=win.destroy,
                  bg=t["accent"], fg="#FFFFFF", relief="flat",
                  padx=16, pady=6, font=UI_FONT_BOLD).pack(side="right")

    # ----- Keyboard shortcuts -----
    def _bind_shortcuts(self):
        b = self.root.bind_all
        b("<Control-o>", lambda e: self._browse_and_scan())
        b("<Control-O>", lambda e: self._browse_and_scan())
        b("<Control-e>", lambda e: self._show_export_menu())
        b("<Control-E>", lambda e: self._show_export_menu())
        b("<F5>",        lambda e: self._rescan())
        b("<Control-f>", lambda e: self.filter_entry.focus_set())
        b("<Control-F>", lambda e: self.filter_entry.focus_set())
        # Global filename search across the entire scan tree
        b("<Control-Shift-F>", lambda e: self._open_find_dialog())
        b("<Control-Shift-f>", lambda e: self._open_find_dialog())
        b("<Escape>",
          lambda e: self._cancel_scan()
          if self.scan_thread and self.scan_thread.is_alive() else None)
        b("<Delete>",    lambda e: self._batch_recycle())
        b("<Key-question>", lambda e: self._show_cheatsheet())
        b("<Shift-slash>",  lambda e: self._show_cheatsheet())

    # ----- Batch recycle (called by Delete shortcut) -----
    def _batch_recycle(self):
        focused = self.root.focus_get()
        target = None
        for tv in (getattr(self, "tree", None),
                   getattr(self, "top_tree", None),
                   getattr(self, "dup_tree", None)):
            if tv is focused:
                target = tv
                break
        if target is None:
            return
        sel = target.selection()
        if not sel:
            return
        paths = []
        for iid in sel:
            if target is getattr(self, "tree", None):
                n = self._tree_node_map.get(iid)
                if n and n.path not in ("(free)", "(gap)"):
                    paths.append(n.path)
            elif target is getattr(self, "top_tree", None):
                v = target.item(iid, "values")
                if len(v) >= 5:
                    paths.append(v[4])
            elif target is getattr(self, "dup_tree", None):
                tags = target.item(iid, "tags")
                if "file" in tags:
                    paths.append(target.item(iid, "text"))
        if not paths:
            return
        if not self._ask(
                "Confirm",
                f"Move {len(paths)} item(s) to the Recycle Bin?"):
            return
        self.status_var.set(
            f"Moving {len(paths)} item(s) to Recycle Bin …")
        self.status_icon.configure(fg=self.theme["warn"])

        def _worker(plist=list(paths)):
            ok = 0
            for p in plist:
                try:
                    if send_to_recycle_bin(p):
                        ok += 1
                except Exception:
                    pass
            self.root.after(0,
                            lambda: self._batch_recycle_done(plist, ok))
        threading.Thread(target=_worker, daemon=True).start()

    def _refresh_after_recycle(self, paths):
        """Called after Recycle Bin operations. Removes matching rows
        from ALL Treeviews (Overview, Duplicates, Large Files) so the
        UI does not show ghost rows the user just deleted."""
        removed = set(paths)
        for tv_name in ("tree", "top_tree", "dup_tree"):
            tv = getattr(self, tv_name, None)
            if tv is None:
                continue
            def _walk(parent):
                for iid in tv.get_children(parent):
                    node = self._tree_node_map.get(iid) \
                        if tv_name == "tree" else None
                    txt = tv.item(iid, "text") or ""
                    values = tv.item(iid, "values") or ()
                    row_path = None
                    if node is not None:
                        row_path = node.path
                    elif values and len(values) > 1:
                        # Last value in top_tree/dup_tree is the path
                        row_path = str(values[-1])
                    if row_path and row_path in removed:
                        try:
                            tv.delete(iid)
                        except tk.TclError:
                            pass
                        continue
                    _walk(iid)
            try:
                _walk("")
            except tk.TclError:
                pass

    def _batch_recycle_done(self, paths, ok):
        n = len(paths)
        if ok == n:
            self.status_icon.configure(fg=self.theme["success"])
            self.status_var.set(
                f"Moved {ok} item(s) to Recycle Bin.")
        else:
            self.status_icon.configure(fg=self.theme["warn"])
            self.status_var.set(
                f"Moved {ok}/{n} item(s) to Recycle Bin.")
        kill = set(paths)
        for tv in (getattr(self, "top_tree", None),
                   getattr(self, "dup_tree", None)):
            if tv is None:
                continue
            for iid in list(tv.get_children()):
                v = tv.item(iid, "values")
                if (v and any(c in kill for c in v)) or \
                        tv.item(iid, "text") in kill:
                    tv.delete(iid)
                    continue
                for cid in list(tv.get_children(iid)):
                    if tv.item(cid, "text") in kill:
                        tv.delete(cid)

    # ----- ttk style baseline -----
        # Refresh all Treeviews so deleted paths disappear immediately
        try:
            self._refresh_after_recycle(ok)
        except Exception:
            pass

    def _setup_style(self):
        self.style = ttk.Style()
        try:
            self.style.theme_use("clam")
        except tk.TclError:
            pass


    def _debounce(self, key, delay_ms, fn):
        """Coalesce rapid <Configure> events. Cancels a pending
        after() for `key` and re-schedules `fn` after `delay_ms`."""
        if not hasattr(self, "_debounce_map"):
            self._debounce_map = {}
        tok = self._debounce_map.get(key)
        if tok is not None:
            try:
                self.root.after_cancel(tok)
            except Exception:
                pass
        try:
            self._debounce_map[key] = self.root.after(delay_ms, fn)
        except Exception:
            try:
                fn()
            except Exception:
                pass

    def _apply_theme(self):
        t = self.theme
        self.root.configure(bg=t["bg"])
        s = self.style
        s.configure("Treeview", background=t["panel"],
                    fieldbackground=t["panel"], foreground=t["fg"],
                    rowheight=30, font=UI_FONT, borderwidth=0)
        s.configure("Treeview.Heading", background=t["panel_alt"],
                    foreground=t["muted"], font=("Segoe UI", 9, "bold"),
                    relief="flat", padding=(10, 6), borderwidth=0)
        s.map("Treeview",
              background=[("selected", t["select"])],
              foreground=[("selected", t["fg"])])
        s.map("Treeview.Heading",
              background=[("active", t["panel_hover"])])
        s.configure("TProgressbar", troughcolor=t["panel_alt"],
                    background=t["accent"], borderwidth=0)
        # Thin, modern scrollbar — no arrow buttons, accent thumb
        s.layout("Vertical.TScrollbar",
                 [("Vertical.Scrollbar.trough",
                   {"children": [("Vertical.Scrollbar.thumb",
                                  {"expand": "1", "sticky": "nswe"})],
                    "sticky": "ns"})])
        s.layout("Horizontal.TScrollbar",
                 [("Horizontal.Scrollbar.trough",
                   {"children": [("Horizontal.Scrollbar.thumb",
                                  {"expand": "1", "sticky": "nswe"})],
                    "sticky": "ew"})])
        s.configure("Vertical.TScrollbar",
                    background=t["border_strong"],
                    troughcolor=t["panel_alt"],
                    bordercolor=t["panel_alt"],
                    lightcolor=t["panel_alt"],
                    darkcolor=t["panel_alt"],
                    arrowcolor=t["muted"],
                    gripcount=0,
                    relief="flat",
                    borderwidth=0,
                    width=10)
        s.configure("Horizontal.TScrollbar",
                    background=t["border_strong"],
                    troughcolor=t["panel_alt"],
                    bordercolor=t["panel_alt"],
                    lightcolor=t["panel_alt"],
                    darkcolor=t["panel_alt"],
                    arrowcolor=t["muted"],
                    gripcount=0,
                    relief="flat",
                    borderwidth=0,
                    width=10)
        s.map("Vertical.TScrollbar",
              background=[("active", t["accent"]),
                          ("pressed", t["accent_dark"])])
        s.map("Horizontal.TScrollbar",
              background=[("active", t["accent"]),
                          ("pressed", t["accent_dark"])])
        s.configure("Card.TFrame", background=t["panel"])
        if hasattr(self, "tree"):
            # Folder-tree row colours: the unicode bars in the "bar" column
            # inherit the row foreground, so we tint rows with the brand
            # accent. Names stay readable; bars become Capacitra-blue.
            self.tree.tag_configure("even", background=t["panel"],
                                    foreground=t["accent_dark"])
            self.tree.tag_configure("odd",  background=t["row_alt"],
                                    foreground=t["accent_dark"])
            self.tree.tag_configure("synth", foreground=t["muted"])

    def _toggle_theme(self):
        """Instant, crash-safe theme toggle.

        Does NOT destroy or rebuild the widget tree. Instead, walks every
        widget recursively and remaps its colour attributes from the old
        theme to the new theme. Chart canvases are then asked to redraw.

        This replaces the older destroy+rebuild approach that could take
        many seconds and occasionally crash on large scan results
        (>300k tree items).
        """
        try:
            self.root.configure(cursor="watch")
            self.root.update_idletasks()
        except Exception:
            pass
        try:
            old_theme = dict(self.theme)
            self.theme_name = "dark" if self.theme_name == "light" else "light"
            self.theme = THEMES[self.theme_name]

            # Build old-colour -> new-colour mapping so any widget currently
            # painted in an old-theme colour will get its new-theme
            # equivalent.
            colour_map = {}
            for k, old_c in old_theme.items():
                new_c = self.theme.get(k)
                if old_c and new_c and old_c != new_c:
                    colour_map[old_c] = new_c

            # ttk styles (Treeview, Scrollbar, Progressbar) get updated
            # here — this covers dozens of widgets in a single call.
            try:
                self._apply_theme()
            except Exception:
                pass

            # Walk every tk widget and remap colour attributes
            try:
                self._retint_widget(self.root, colour_map)
            except Exception:
                pass

            # Also walk any open Toplevel dialogs
            try:
                for w in self.root.winfo_children():
                    if isinstance(w, tk.Toplevel):
                        self._retint_widget(w, colour_map)
            except Exception:
                pass

            # Redraw known chart canvases so their painted items match
            for attr in ("hero_canvas", "bar_chart", "pie",
                         "sb_progress", "top_folders_canvas"):
                w = getattr(self, attr, None)
                if w is None:
                    continue
                try:
                    if hasattr(w, "cget"):
                        cur = w.cget("bg")
                        if cur in colour_map:
                            w.configure(bg=colour_map[cur])
                    if hasattr(w, "redraw"):
                        w.redraw()
                except Exception:
                    pass

            # Refresh the treemap if it holds data
            try:
                if hasattr(self, "treemap") and hasattr(self.treemap, "redraw"):
                    self.treemap.redraw()
            except Exception:
                pass

            # Update the moon/sun button icon
            try:
                if hasattr(self, "theme_btn"):
                    self.theme_btn.configure(
                        text="🌙" if self.theme_name == "light" else "☀"
                    )
            except Exception:
                pass

            # Force a final layout pass
            try:
                self.root.configure(bg=self.theme["bg"])
                self.root.update_idletasks()
            except Exception:
                pass
        finally:
            try:
                self.root.configure(cursor="")
            except Exception:
                pass

    def _retint_widget(self, w, colour_map):
        """Recursively walk widgets, remapping colour attributes.

        Handles bg, fg, highlightbackground, highlightcolor,
        selectbackground, selectforeground, insertbackground,
        activebackground, activeforeground, disabledforeground,
        readonlybackground. Silently skips widgets that don't support
        a given attribute.
        """
        attrs = ("bg", "fg", "highlightbackground", "highlightcolor",
                 "selectbackground", "selectforeground",
                 "insertbackground", "activebackground",
                 "activeforeground", "disabledforeground",
                 "readonlybackground", "troughcolor")
        for attr in attrs:
            try:
                cur = w.cget(attr)
                if cur in colour_map:
                    w.configure(**{attr: colour_map[cur]})
            except Exception:
                pass
        for c in w.winfo_children():
            self._retint_widget(c, colour_map)

    # ----- top-level layout -----
    def _build_layout(self):
        t = self.theme
        # Outer container
        outer = tk.Frame(self.root, bg=t["bg"])
        outer.pack(fill="both", expand=True)
        # Sidebar fixed width
        self.sidebar = tk.Frame(outer, bg=t["sb_bg"], width=240)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self._build_sidebar()
        # Right column: header + address + content + status
        right = tk.Frame(outer, bg=t["bg"])
        right.pack(side="left", fill="both", expand=True)
        self.right_col = right
        self._build_header(right)
        self._build_address(right)
        self._build_content_stack(right)
        self._build_status(right)

    # ----- Sidebar -----
    def _build_sidebar(self):
        t = self.theme
        # Logo box
        logo_row = tk.Frame(self.sidebar, bg=t["sb_bg"])
        logo_row.pack(fill="x", padx=16, pady=(20, 16))
        self._sb_logo_row = logo_row  # exposed for toggle
        canvas = tk.Canvas(logo_row, width=44, height=44, bg=t["sb_bg"],
                           highlightthickness=0)
        canvas.pack(side="left")
        self._draw_capacitra_mark(canvas, 22, 22, 20)
        text_box = tk.Frame(logo_row, bg=t["sb_bg"])
        text_box.pack(side="left", padx=10)
        self._sb_logo_text_box = text_box  # exposed for toggle
        tk.Label(text_box, text=APP_NAME, bg=t["sb_bg"], fg="white",
                 font=("Segoe UI Semibold", 13)).pack(anchor="w")
        tk.Label(text_box, text=APP_TAGLINE, bg=t["sb_bg"],
                 fg=t["sb_label"], font=("Segoe UI", 8)).pack(anchor="w")

        # Nav items
        nav_box = tk.Frame(self.sidebar, bg=t["sb_bg"])
        nav_box.pack(fill="x", pady=(10, 0))
        for key, icon, label, sub in self.NAV_ITEMS:
            it = SidebarItem(nav_box, t, icon, label,
                             command=lambda k=key: self._select_panel(k),
                             has_submenu=sub)
            it.pack(fill="x")
            self._nav_items[key] = it
        # Bottom mini disk card
        spacer = tk.Frame(self.sidebar, bg=t["sb_bg"])
        spacer.pack(fill="both", expand=True)
        self.sb_disk_card = tk.Frame(self.sidebar, bg=t["sb_bg_active"],
                                     highlightthickness=0)
        self.sb_disk_card.pack(side="bottom", fill="x",
                               padx=14, pady=14, ipady=2)
        inner = tk.Frame(self.sb_disk_card, bg=t["sb_bg_active"])
        inner.pack(fill="x", padx=10, pady=10)
        tk.Label(inner, text="🖴", bg=t["sb_bg_active"], fg=t["sb_fg_active"],
                 font=("Segoe UI Emoji", 12)).pack(side="left", padx=(0, 6))
        col = tk.Frame(inner, bg=t["sb_bg_active"])
        col.pack(side="left", fill="x", expand=True)
        self.sb_disk_title = tk.Label(col, text="No drive scanned yet",
                                      bg=t["sb_bg_active"],
                                      fg=t["sb_fg_active"],
                                      font=("Segoe UI", 9, "bold"),
                                      anchor="w", wraplength=170,
                                      justify="left")
        self.sb_disk_title.pack(anchor="w")
        self.sb_disk_sub = tk.Label(col, text="—",
                                    bg=t["sb_bg_active"],
                                    fg=t["sb_label"], font=("Segoe UI", 8),
                                    anchor="w", wraplength=170,
                                    justify="left")
        self.sb_disk_sub.pack(anchor="w")
        self.sb_progress = tk.Canvas(self.sb_disk_card,
                                     height=6,
                                     bg=t["sb_bg_active"],
                                     highlightthickness=0)
        self.sb_progress.pack(fill="x", padx=12, pady=(2, 4))
        self.sb_progress_pct = tk.Label(self.sb_disk_card, text="",
                                        bg=t["sb_bg_active"],
                                        fg=t["sb_label"],
                                        font=("Segoe UI", 8))
        self.sb_progress_pct.pack(anchor="w", padx=12, pady=(0, 8))

        # Theme toggle switch at the very bottom
        self._build_theme_switch()

    def _draw_capacitra_mark(self, canvas, cx, cy, r):
            """Hexagonal Capacitra brand mark with a folder + bars inside."""
            # Pointy-top hex points
            import math as _m
            pts = []
            for i in range(6):
                ang = _m.radians(-90 + 60 * i)
                pts.extend([cx + r * _m.cos(ang), cy + r * _m.sin(ang)])
            # Hex body — deep blue with brighter edge
            canvas.create_polygon(pts, fill="#1E3A8A", outline="#3B82F6", width=2)
            # Inner soft-blue hex (lighter middle)
            pts2 = []
            for i in range(6):
                ang = _m.radians(-90 + 60 * i)
                pts2.extend([cx + (r - 2) * _m.cos(ang),
                             cy + (r - 2) * _m.sin(ang)])
            canvas.create_polygon(pts2, fill="#1D4ED8", outline="")
            # Folder body
            fx, fy = cx - r * 0.55, cy - r * 0.35
            fw, fh = r * 1.1, r * 0.8
            # Tab
            canvas.create_polygon(
                fx, fy,                  fx + fw * 0.45, fy,
                fx + fw * 0.55, fy + 3,  fx + fw, fy + 3,
                fx + fw, fy + fh,        fx, fy + fh,
                fill="#0EA5E9", outline="#22D3EE", width=1
            )
            # Bars inside folder (3 ascending)
            bx = fx + 3
            by = fy + fh - 2
            bw = (fw - 6) / 5
            heights = [fh * 0.35, fh * 0.55, fh * 0.8]
            for i, h in enumerate(heights):
                x = bx + (i * 1.6 + 0.5) * bw
                canvas.create_rectangle(x, by - h, x + bw, by,
                                        fill="#FFFFFF", outline="")

    def _build_theme_switch(self):
        t = self.theme
        switch_row = tk.Frame(self.sidebar, bg=t["sb_bg"])
        switch_row.pack(side="bottom", fill="x", padx=24, pady=(0, 18))
        self._sb_theme_row = switch_row  # exposed for toggle
        sun_lbl = tk.Label(switch_row, text="☀", bg=t["sb_bg"],
                           fg=t["sb_fg_active"] if self.theme_name == "light"
                              else t["sb_label"],
                           font=("Segoe UI Emoji", 13), cursor="hand2")
        sun_lbl.pack(side="left")
        pill = tk.Canvas(switch_row, width=46, height=24, bg=t["sb_bg"],
                         highlightthickness=0, cursor="hand2")
        pill.pack(side="left", padx=8)
        # Pill background
        pill.create_oval(0, 0, 24, 24, fill=t["accent"], outline="")
        pill.create_oval(22, 0, 46, 24, fill=t["accent"], outline="")
        pill.create_rectangle(12, 0, 34, 24, fill=t["accent"], outline="")
        # Knob — right side when in dark mode, left when light
        if self.theme_name == "dark":
            pill.create_oval(24, 2, 44, 22, fill="white", outline="")
        else:
            pill.create_oval(2, 2, 22, 22, fill="white", outline="")
        moon_lbl = tk.Label(switch_row, text="☾", bg=t["sb_bg"],
                            fg=t["sb_fg_active"] if self.theme_name == "dark"
                               else t["sb_label"],
                            font=("Segoe UI Emoji", 13), cursor="hand2")
        moon_lbl.pack(side="left")
        for w in (sun_lbl, pill, moon_lbl, switch_row):
            w.bind("<Button-1>", lambda e: self._toggle_theme())

    def _update_sb_disk(self, used, total, label):
        if total <= 0:
            self.sb_disk_title.configure(text="No drive scanned yet")
            self.sb_disk_sub.configure(text="—")
            self.sb_progress_pct.configure(text="")
            self.sb_progress.delete("all")
            return
        pct = used / total * 100 if total else 0
        self.sb_disk_title.configure(text=label or "Local Disk")
        self.sb_disk_sub.configure(
            text=f"{human_size(used)} / {human_size(total)} used")
        self.sb_progress.delete("all")
        w = self.sb_progress.winfo_width() or 200
        h = 6
        self.sb_progress.create_rectangle(0, 0, w, h,
                                          fill=self.theme["sb_bg"],
                                          outline="")
        self.sb_progress.create_rectangle(0, 0, max(w * pct / 100, 4), h,
                                          fill=self.theme["accent"],
                                          outline="")
        self.sb_progress_pct.configure(text=f"{pct:.0f}% Used")

    # ----- Header -----
    def _build_header(self, parent):
        t = self.theme
        bar = tk.Frame(parent, bg=t["bg"])
        bar.pack(side="top", fill="x", padx=24, pady=(16, 8))
        # Left: hamburger placeholder + title block
        left = tk.Frame(bar, bg=t["bg"])
        left.pack(side="left")
        burger = tk.Label(left, text="≡", bg=t["bg"], fg=t["muted"],
                          font=("Segoe UI", 16), cursor="hand2",
                          padx=8, pady=2)
        burger.pack(side="left", padx=(0, 14))
        burger.bind("<Enter>", lambda e: burger.configure(fg=t["fg"]))
        burger.bind("<Leave>", lambda e: burger.configure(fg=t["muted"]))
        burger.bind("<Button-1>", lambda e: self._toggle_sidebar())
        tcol = tk.Frame(left, bg=t["bg"])
        tcol.pack(side="left")
        self.header_title_lbl = tk.Label(
            tcol, text="Overview", bg=t["bg"], fg=t["fg"],
            font=("Segoe UI Semibold", 20))
        self.header_title_lbl.pack(anchor="w")
        self.header_sub_lbl = tk.Label(
            tcol, text="Get a quick overview of your disk usage",
            bg=t["bg"], fg=t["muted"], font=UI_FONT)
        self.header_sub_lbl.pack(anchor="w", pady=(2, 0))
        # Right: action buttons
        right = tk.Frame(bar, bg=t["bg"])
        right.pack(side="right")
        self.new_scan_btn = IconButton(right, t, "🔍", "New Scan",
                                       self._on_primary_action,
                                       variant="primary")
        self.new_scan_btn.pack(side="left", padx=4)
        self.rescan_btn = IconButton(right, t, "↻", "Rescan",
                                     self._rescan, variant="ghost")
        self.rescan_btn.pack(side="left", padx=4)
        self.export_btn = IconButton(right, t, "📤", "Export",
                                     self._show_export_menu, variant="ghost")
        self.export_btn.pack(side="left", padx=4)
        self.theme_btn = GhostIconButton(right, t,
                                         "🌙" if self.theme_name == "light"
                                         else "☀",
                                         self._toggle_theme)
        self.theme_btn.pack(side="left", padx=4)
        self.kebab_btn = GhostIconButton(right, t, "⋮",
                                         self._show_kebab_menu)
        self.kebab_btn.pack(side="left", padx=4)

    def _toggle_sidebar(self):
        """Collapse sidebar to a slim icon-rail / expand back."""
        if not hasattr(self, "_sidebar_collapsed"):
            self._sidebar_collapsed = False
        if self._sidebar_collapsed:
            # Expand back
            self.sidebar.configure(width=240)
            for it in self._nav_items.values():
                try:
                    it.text_lbl.pack(side="left", pady=10)
                except Exception:
                    pass
                # Restore chevron if this item has a submenu
                if getattr(it, "chev", None) is not None:
                    try:
                        it.chev.pack(side="right", padx=(0, 14))
                    except Exception:
                        pass
                # Icon back to left-aligned
                try:
                    it.icon_lbl.pack_forget()
                    it.icon_lbl.pack(side="left", padx=(12, 12))
                except Exception:
                    pass
            # Restore logo text
            try:
                self._sb_logo_text_box.pack(side="left", padx=10)
            except Exception:
                pass
            # Restore theme switch row
            try:
                self._sb_theme_row.pack(side="bottom", fill="x",
                                        padx=24, pady=(0, 18))
            except Exception:
                pass
            # Restore disk card
            try:
                self.sb_disk_card.pack(side="bottom", fill="x",
                                       padx=14, pady=14, ipady=2)
            except Exception:
                pass
            # Restore logo row with original 16px padding
            try:
                self._sb_logo_row.pack_forget()
                self._sb_logo_row.pack(fill="x", padx=16, pady=(20, 16))
            except Exception:
                pass
            self._sidebar_collapsed = False
        else:
            # Collapse to 64px icon-rail
            self.sidebar.configure(width=64)
            for it in self._nav_items.values():
                try:
                    it.text_lbl.pack_forget()
                except Exception:
                    pass
                if getattr(it, "chev", None) is not None:
                    try:
                        it.chev.pack_forget()
                    except Exception:
                        pass
                # Center the icon in the collapsed 64px rail
                try:
                    it.icon_lbl.pack_forget()
                    it.icon_lbl.pack(pady=10)
                except Exception:
                    pass
            # Hide logo text — only the mark remains visible
            try:
                self._sb_logo_text_box.pack_forget()
            except Exception:
                pass
            # Hide theme switch row (pill is too wide for 64px)
            try:
                self._sb_theme_row.pack_forget()
            except Exception:
                pass
            # Hide disk card entirely
            try:
                self.sb_disk_card.pack_forget()
            except Exception:
                pass
            # Recenter the 44px logo inside 64px rail (padx 10 vs 16)
            try:
                self._sb_logo_row.pack_forget()
                self._sb_logo_row.pack(fill="x", padx=10, pady=(20, 16))
            except Exception:
                pass
            self._sidebar_collapsed = True
        # Force immediate layout redraw
        try:
            self.root.update_idletasks()
        except Exception:
            pass

    def _show_kebab_menu(self):
        m = tk.Menu(self.root, tearoff=0)
        m.add_command(label="Export CSV…",
                      command=self._export_csv)
        m.add_command(label="Export HTML…",
                      command=self._export_html)
        if HAS_REPORTLAB:
            m.add_command(label="Export PDF…",
                          command=self._export_pdf)
        if HAS_OPENPYXL:
            m.add_command(label="Export Excel…",
                          command=self._export_excel)
        m.add_separator()
        m.add_command(label="Save snapshot…",
                      command=self._save_snapshot)
        m.add_command(label="Load snapshot…",
                      command=self._load_snapshot)
        m.add_command(label="Compare with snapshot…",
                      command=self._compare_snapshot)
        m.add_separator()
        m.add_command(label="Schedule daily scan…",
                      command=self._install_schedule)
        m.add_separator()
        m.add_command(label="Find duplicates",
                      command=self._find_duplicates)
        m.add_command(label="Duplicate clusters by folder…",
                      command=self._show_duplicate_clusters)
        m.add_separator()
        m.add_command(label="Find empty folders…",
                      command=self._find_empty_folders)
        m.add_command(label="Suspicious executables…",
                      command=self._show_suspicious_exes)
        m.add_command(label="Downloads folder aging…",
                      command=self._show_downloads_hint)
        m.add_separator()
        m.add_command(label="Keyboard shortcuts  (?)",
                      command=self._show_cheatsheet)
        m.add_separator()
        m.add_command(label="About Capacitra",
                      command=lambda: self._select_panel("about"))
        try:
            x = self.kebab_btn.winfo_rootx()
            y = self.kebab_btn.winfo_rooty() + self.kebab_btn.winfo_height()
            m.tk_popup(x, y)
        finally:
            m.grab_release()

    # ----- Address bar -----
    def _build_address(self, parent):
        t = self.theme
        bar = tk.Frame(parent, bg=t["bg"])
        bar.pack(side="top", fill="x", padx=24, pady=(0, 12))
        # Drive picker pill
        pill = tk.Frame(bar, bg=t["panel"], highlightthickness=1,
                        highlightbackground=t["border"])
        pill.pack(side="left")
        tk.Label(pill, text="🖴", bg=t["panel"], fg=t["accent"],
                 font=("Segoe UI Emoji", 11)).pack(side="left",
                                                   padx=(10, 6), pady=6)
        self.path_var = tk.StringVar()
        self.path_combo = ttk.Combobox(pill, textvariable=self.path_var,
                                       width=22, state="readonly")
        self.path_combo.pack(side="left", padx=(0, 4), pady=4)
        self.path_combo.bind("<<ComboboxSelected>>",
                             lambda e: self._on_drive_change())
        # Browse… button inside the same pill
        browse_btn = tk.Label(pill, text="📂  Browse…", bg=t["panel"],
                              fg=t["accent"], cursor="hand2",
                              font=("Segoe UI", 9, "bold"),
                              padx=10, pady=6)
        browse_btn.pack(side="left", padx=(0, 4), pady=4)
        browse_btn.bind("<Enter>",
                        lambda e: browse_btn.configure(bg=t["panel_hover"]))
        browse_btn.bind("<Leave>",
                        lambda e: browse_btn.configure(bg=t["panel"]))
        browse_btn.bind("<Button-1>", lambda e: self._browse_and_scan())
        # Breadcrumb area
        self.crumb_box = tk.Frame(bar, bg=t["bg"])
        self.crumb_box.pack(side="left", padx=14, fill="x", expand=True)
        # Search box on right
        search_frame = tk.Frame(bar, bg=t["panel"], highlightthickness=1,
                                highlightbackground=t["border"])
        search_frame.pack(side="right", padx=(0, 6))
        tk.Label(search_frame, text="🔎", bg=t["panel"], fg=t["muted"],
                 font=("Segoe UI Emoji", 10)).pack(side="left",
                                                   padx=(10, 4), pady=6)
        self.filter_var = tk.StringVar()
        self.filter_entry = tk.Entry(search_frame, textvariable=self.filter_var,
                                     bd=0, bg=t["panel"], fg=t["fg"],
                                     width=24, font=UI_FONT)
        self.filter_entry.pack(side="left", padx=(0, 8), pady=4)
        self.filter_entry.bind("<KeyRelease>", lambda e: self._apply_filter())
        # Filter icon button
        fbtn = tk.Label(bar, text="⛃", bg=t["panel"], fg=t["fg"],
                        font=("Segoe UI Emoji", 12), padx=10, pady=6,
                        highlightthickness=1,
                        highlightbackground=t["border"],
                        cursor="hand2")
        fbtn.pack(side="right")

    def _refresh_breadcrumb(self):
        for w in self.crumb_box.winfo_children():
            w.destroy()
        t = self.theme
        node = self.current_node
        if not node:
            return
        # Build path top-down
        parts = []
        cur = node
        while cur:
            parts.append(cur)
            cur = cur.parent
        parts.reverse()

        # Render helpers — keeps the breadcrumb compact even on deep paths
        def sep():
            tk.Label(self.crumb_box, text="›", bg=t["bg"],
                     fg=t["muted"], font=UI_FONT
                     ).pack(side="left", padx=4)

        def crumb(label, node_ref=None, last=False):
            short = label if len(label) <= 22 else label[:19] + "…"
            lbl = tk.Label(self.crumb_box, text=short, bg=t["bg"],
                           fg=t["fg"] if last else t["accent"],
                           font=UI_FONT_BOLD if last else UI_FONT,
                           cursor="" if last else "hand2")
            lbl.pack(side="left")
            if node_ref is not None and not last:
                lbl.bind("<Button-1>",
                         lambda e, nn=node_ref: self._focus_node(nn))

        # "This PC" anchor
        tk.Label(self.crumb_box, text="This PC", bg=t["bg"],
                 fg=t["accent"], font=UI_FONT
                 ).pack(side="left")

        # Pretty-name the drive (first part)
        def pretty(name):
            if len(name) <= 4:
                return f"Local Disk ({name.rstrip(chr(92)).rstrip('/')})"
            return name

        n_parts = len(parts)
        # Collapse middle segments if path is deep
        if n_parts <= 4:
            for i, n in enumerate(parts):
                sep()
                crumb(pretty(n.name) if i == 0 else n.name,
                      node_ref=n, last=(i == n_parts - 1))
        else:
            # Show: drive › first child › … › parent › current
            sep()
            crumb(pretty(parts[0].name), node_ref=parts[0])
            sep()
            crumb(parts[1].name, node_ref=parts[1])
            sep()
            # Ellipsis dropdown (popup with hidden segments)
            ell = tk.Label(self.crumb_box, text="…", bg=t["bg"],
                           fg=t["muted"], font=UI_FONT_BOLD,
                           cursor="hand2", padx=4)
            ell.pack(side="left")
            hidden = parts[2:-2]

            def show_hidden(event, items=hidden):
                m = tk.Menu(self.root, tearoff=0)
                for nn in items:
                    nm = nn.name if len(nn.name) <= 40 else nn.name[:37] + "…"
                    m.add_command(label=nm,
                                  command=lambda x=nn: self._focus_node(x))
                m.tk_popup(event.x_root, event.y_root)
            ell.bind("<Button-1>", show_hidden)
            sep()
            crumb(parts[-2].name, node_ref=parts[-2])
            sep()
            crumb(parts[-1].name, node_ref=parts[-1], last=True)

    def _on_drive_change(self):
        # User picked a drive from address bar; remember and offer scan.
        pass

    # ----- Content stack -----
    def _build_content_stack(self, parent):
        t = self.theme
        stack = tk.Frame(parent, bg=t["bg"])
        stack.pack(side="top", fill="both", expand=True, padx=24, pady=0)
        self.content_stack = stack
        self._build_panel_overview(stack)
        self._build_panel_treemap(stack)
        self._build_panel_charts(stack)
        self._build_panel_duplicates(stack)
        self._build_panel_large(stack)
        self._build_panel_types(stack)
        self._build_panel_settings(stack)
        self._build_panel_history(stack)
        self._build_panel_export(stack)
        self._build_panel_about(stack)


    def _select_panel(self, key):
        # Hide current
        if self._active_panel and self._active_panel in self._panels:
            self._panels[self._active_panel].pack_forget()
        # Show selected
        panel = self._panels.get(key)
        if panel:
            panel.pack(fill="both", expand=True)
        self._active_panel = key
        # Sync the Duplicates threshold display with the current value
        # so the two threshold editors (Settings vs Duplicates panel)
        # never drift apart.
        if key == "duplicates" and hasattr(self, "dup_thr_var"):
            try:
                self.dup_thr_var.set(
                    str(max(1, ScanWorker.DUP_THRESHOLD // 1024)))
            except Exception:
                pass
        # Toggle nav active state
        for k, it in self._nav_items.items():
            it.set_active(k == key)
        # Update header subtitle/title
        title_map = {
            "overview":   ("Overview", "Get a quick overview of your disk usage"),
            "treemap":    ("Treemap", "Hierarchical view of disk usage"),
            "charts":     ("Charts", "Pie and bar views of the current folder"),
            "duplicates": ("Duplicates", "Find identical files wasting space"),
            "large":      ("Large Files", "Top files by size"),
            "types":      ("File Types", "Disk usage grouped by extension"),
            "export":     ("Export", "Save reports as CSV or HTML"),
            "settings":   ("Settings", "Customize how Capacitra behaves"),
            "history":    ("History", "Recent scans in this session"),
            "about":      ("About", "Capacitra — Storage capacity intelligence"),
        }
        ttl, sub = title_map.get(key, ("Overview", ""))
        self.header_title_lbl.configure(text=ttl)
        self.header_sub_lbl.configure(text=sub)
        # Side actions: certain nav items trigger a flow instead of just
        # showing a panel.

        # Export now has its own panel (no popup menu)

    # ----- Status bar -----
    def _build_status(self, parent):
        t = self.theme
        bar = tk.Frame(parent, bg=t["panel"], highlightthickness=1,
                       highlightbackground=t["border"])
        bar.pack(side="bottom", fill="x", padx=24, pady=(0, 16))
        inner = tk.Frame(bar, bg=t["panel"])
        inner.pack(fill="x", padx=18, pady=10)
        # Check icon
        self.status_icon = tk.Label(inner, text="●", bg=t["panel"],
                                    fg=t["success"], font=("Segoe UI", 12))
        self.status_icon.pack(side="left", padx=(0, 8))
        self.status_var = tk.StringVar(
            value=f"Ready  ·  {APP_NAME} {APP_VERSION}  ·  Pick a drive from the sidebar and click New Scan to begin.")
        tk.Label(inner, textvariable=self.status_var, bg=t["panel"],
                 fg=t["muted"], font=UI_FONT, anchor="w"
                 ).pack(side="left", fill="x", expand=True)
        self.progress = ttk.Progressbar(inner, mode="indeterminate",
                                        length=180, style="TProgressbar")
        self.progress.pack(side="right")

    # ----- Panels -----
    def _card(self, parent, padx=22, pady=16):
        """Return a white rounded-ish card Frame inside parent."""
        t = self.theme
        wrapper = tk.Frame(parent, bg=t["bg"])
        # Shadow underneath
        shadow = tk.Frame(wrapper, bg=t["shadow"], height=1)
        shadow.pack(fill="x", side="bottom")
        card = tk.Frame(wrapper, bg=t["panel"], highlightthickness=1,
                        highlightbackground=t["border"])
        card.pack(fill="both", expand=True, ipadx=padx, ipady=pady)
        return wrapper, card

    def _build_panel_overview(self, parent):
        t = self.theme
        panel = tk.Frame(parent, bg=t["bg"])
        # Default: tree takes the full width. Dashboard (right column) is
        # initially hidden and is shown side-by-side when the user clicks
        # the dashboard toggle in the tree header.
        left = tk.Frame(panel, bg=t["bg"])
        left.pack(side="left", fill="both", expand=True)
        right = tk.Frame(panel, bg=t["bg"], width=480)
        right.pack_propagate(False)
        # NOTE: right pane is not packed here — it is shown by
        # _toggle_dashboard. Saved on self for later access.
        self._overview_left = left
        self._overview_right = right
        self._dashboard_visible = False
        # Folder tree card (no fixed width — let it breathe)
        tree_wrap = tk.Frame(left, bg=t["panel"], highlightthickness=1,
                             highlightbackground=t["border"])
        tree_wrap.pack(fill="both", expand=True)
        head = tk.Frame(tree_wrap, bg=t["panel"])
        head.pack(fill="x", padx=20, pady=(16, 10))
        tk.Label(head, text="Folder Tree", bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 13)).pack(side="left")
        # Tree header action icons (right side)
        icons_box = tk.Frame(head, bg=t["panel"])
        icons_box.pack(side="right")

        def _hdr_btn(parent, icon, tip, cmd):
            b = tk.Label(parent, text=icon, bg=t["panel"], fg=t["muted"],
                         font=("Segoe UI Emoji", 11), padx=8, pady=4,
                         cursor="hand2")
            b.pack(side="right", padx=2)
            b.bind("<Enter>", lambda e: b.configure(bg=t["panel_hover"],
                                                    fg=t["fg"]))
            b.bind("<Leave>", lambda e: b.configure(bg=t["panel"],
                                                    fg=t["muted"]))
            b.bind("<Button-1>", lambda e: cmd())
            # Simple tooltip
            tip_win = {"tw": None}
            def show_tip(_):
                if tip_win["tw"]:
                    return
                tip_win["tw"] = tk.Toplevel(b)
                tip_win["tw"].wm_overrideredirect(True)
                tk.Label(tip_win["tw"], text=tip, bg="#111", fg="white",
                         font=UI_FONT, padx=8, pady=4).pack()
                x = b.winfo_rootx() + b.winfo_width() // 2 - 60
                y = b.winfo_rooty() + b.winfo_height() + 6
                tip_win["tw"].geometry(f"+{x}+{y}")
            def hide_tip(_):
                if tip_win["tw"]:
                    tip_win["tw"].destroy()
                    tip_win["tw"] = None
            b.bind("<Enter>", show_tip, add="+")
            b.bind("<Leave>", hide_tip, add="+")
            return b

        _hdr_btn(icons_box, "▦", "Toggle dashboard",
                 self._toggle_dashboard)
        _hdr_btn(icons_box, "⛶", "Expand folder tree",
                 self._toggle_tree_fullscreen)

        self.tree_stat_lbl = tk.Label(head, text="—",
                                      bg=t["panel"], fg=t["muted"],
                                      font=UI_FONT)
        self.tree_stat_lbl.pack(side="right", padx=(0, 14))
        tree_box = tk.Frame(tree_wrap, bg=t["panel"])
        tree_box.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        cols = ("size", "allocated", "files", "folders", "bar",
                "percent", "modified", "accessed", "owner")
        self.tree = ttk.Treeview(tree_box, columns=cols,
                                 show="tree headings", selectmode="extended")
        self.tree.heading("#0",        text="Name", anchor="w",
                          command=lambda: self._sort_tree("#0"))
        self.tree.heading("size",      text="Size",
                          command=lambda: self._sort_tree("size"))
        self.tree.heading("allocated", text="Alloc",
                          command=lambda: self._sort_tree("allocated"))
        self.tree.heading("files",     text="Files",
                          command=lambda: self._sort_tree("files"))
        self.tree.heading("folders",   text="Folders",
                          command=lambda: self._sort_tree("folders"))
        self.tree.heading("bar",       text="")
        self.tree.heading("percent",   text="%",
                          command=lambda: self._sort_tree("percent"))
        self.tree.heading("modified",  text="Modified",
                          command=lambda: self._sort_tree("modified"))
        self.tree.heading("accessed",  text="Accessed",
                          command=lambda: self._sort_tree("accessed"))
        self.tree.heading("owner",     text="Owner")
        self.tree.column("#0",        width=300, minwidth=200, anchor="w", stretch=True)
        self.tree.column("size",      width=96,  minwidth=76,  anchor="e", stretch=False)
        self.tree.column("allocated", width=90,  minwidth=72,  anchor="e", stretch=False)
        self.tree.column("files",     width=92,  minwidth=72,  anchor="e", stretch=False)
        self.tree.column("folders",   width=96,  minwidth=76,  anchor="e", stretch=False)
        self.tree.column("bar",       width=100, minwidth=64,  anchor="w", stretch=False)
        self.tree.column("percent",   width=72,  minwidth=52,  anchor="e", stretch=False)
        self.tree.column("modified",  width=124, minwidth=100, anchor="w", stretch=False)
        self.tree.column("accessed",  width=124, minwidth=100, anchor="w", stretch=False)
        self.tree.column("owner",     width=280, minwidth=160, anchor="w", stretch=False)
        # Owner column cache: maps path -> resolved owner string
        if not hasattr(self, "_owner_cache"):
            self._owner_cache = OrderedDict()
        vsb = ttk.Scrollbar(tree_box, orient="vertical",
                            command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self.tree.bind("<<TreeviewOpen>>", self._on_tree_open)
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        self._make_context_menu(self.tree, kind="tree")

        # Right column: hero + cards + top folders
        # Hero card (disk usage)
        hero_wrap = tk.Frame(right, bg=t["panel"], highlightthickness=1,
                             highlightbackground=t["border"])
        hero_wrap.pack(fill="x", pady=(0, 14))
        hi = tk.Frame(hero_wrap, bg=t["panel"])
        hi.pack(fill="x", padx=26, pady=22)
        self.hero_label = tk.Label(hi, text="DISK USAGE", bg=t["panel"],
                                   fg=t["muted"], font=("Segoe UI Semibold", 9))
        self.hero_label.pack(anchor="w")
        # Two columns: big text left, bar right
        body = tk.Frame(hi, bg=t["panel"])
        body.pack(fill="x", pady=(6, 0))
        bigcol = tk.Frame(body, bg=t["panel"])
        bigcol.pack(side="left", anchor="n")
        self.hero_size_lbl = tk.Label(bigcol, text="—", bg=t["panel"],
                                      fg=t["fg"],
                                      font=("Segoe UI Semibold", 28))
        self.hero_size_lbl.pack(anchor="w")
        self.hero_sub_lbl = tk.Label(bigcol, text="",
                                     bg=t["panel"], fg=t["muted"],
                                     font=UI_FONT)
        self.hero_sub_lbl.pack(anchor="w", pady=(2, 0))
        # Bar UNDER the big number (so legend has room to breathe)
        self.hero_canvas = tk.Canvas(hi, height=20, bg=t["panel"],
                                     highlightthickness=0)
        self.hero_canvas.pack(fill="x", pady=(14, 6))
        self.hero_legend = tk.Frame(hi, bg=t["panel"])
        self.hero_legend.pack(fill="x")
        # Redraw whenever the canvas resizes
        self.hero_canvas.bind("<Configure>", lambda e: self._debounce("hero", 80, self._redraw_hero))

        # 4 stat cards as a 2x2 grid (fits narrow dashboard)
        cards = tk.Frame(right, bg=t["bg"])
        cards.pack(fill="x", pady=(0, 14))
        cards.grid_columnconfigure(0, weight=1, uniform="c")
        cards.grid_columnconfigure(1, weight=1, uniform="c")
        self._stat_cards = []
        defs = [
            ("Files",        "📄", "#16A34A", "Total files"),
            ("Folders",      "📁", "#EA580C", "Total folders"),
            ("File types",   "🏷", "#7C3AED", "Total types"),
            ("Inaccessible", "🔒", "#B45309", "Files"),
        ]
        for i, (lbl, icon, color, sub) in enumerate(defs):
            row, col = i // 2, i % 2
            wrap = tk.Frame(cards, bg=t["panel"], highlightthickness=1,
                            highlightbackground=t["border"])
            wrap.grid(row=row, column=col, sticky="nsew",
                      padx=(0 if col == 0 else 6, 6 if col == 0 else 0),
                      pady=(0 if row == 0 else 6, 0))
            inner = tk.Frame(wrap, bg=t["panel"])
            inner.pack(fill="x", padx=18, pady=14)
            # icon badge (Frame with rounded look via background + emoji)
            badge_bg_map = {
                "#16A34A": "#DCFCE7",
                "#EA580C": "#FFEDD5",
                "#7C3AED": "#EDE9FE",
                "#B45309": "#FEF3C7",
            }
            badge_bg = badge_bg_map.get(color, "#E5E7EB")
            badge = tk.Label(inner, text=icon, bg=badge_bg, fg=color,
                             font=("Segoe UI Emoji", 14), padx=10, pady=4)
            badge.pack(anchor="w")
            tk.Label(inner, text=lbl, bg=t["panel"], fg=t["muted"],
                     font=UI_FONT).pack(anchor="w", pady=(10, 0))
            val_lbl = tk.Label(inner, text="—", bg=t["panel"], fg=t["fg"],
                               font=("Segoe UI Semibold", 18))
            val_lbl.pack(anchor="w", pady=(2, 0))
            tk.Label(inner, text=sub, bg=t["panel"], fg=t["muted"],
                     font=("Segoe UI", 8)).pack(anchor="w", pady=(4, 0))
            self._stat_cards.append(val_lbl)

        # Top folders card
        topwrap = tk.Frame(right, bg=t["panel"], highlightthickness=1,
                           highlightbackground=t["border"])
        topwrap.pack(fill="both", expand=True)
        ti = tk.Frame(topwrap, bg=t["panel"])
        ti.pack(fill="x", padx=22, pady=(16, 8))
        tk.Label(ti, text="Top folders by size", bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 12)).pack(side="left")
        self.top_folders_canvas = tk.Canvas(topwrap, bg=t["panel"],
                                            highlightthickness=0)
        self.top_folders_canvas.pack(fill="both", expand=True,
                                     padx=22, pady=(0, 18))
        self.top_folders_canvas.bind(
            "<Configure>", lambda e: self._redraw_top_folders())

        self._panels["overview"] = panel

    def _toggle_tree_fullscreen(self):
        """Same as toggle_dashboard from the tree's perspective —
        the only way to change the layout is to show/hide the dashboard."""
        self._toggle_dashboard()

    def _toggle_dashboard(self):
        """Show the dashboard column on the right (or hide it again).
        Scan results (tree) always stay visible — they're the primary view."""
        if not hasattr(self, "_overview_right"):
            return
        if self._dashboard_visible:
            self._overview_right.pack_forget()
            self._overview_left.pack_configure(padx=0)
            self._dashboard_visible = False
        else:
            # Tree shrinks (but keeps majority); dashboard slides in on right.
            self._overview_left.pack_configure(padx=(0, 14))
            self._overview_right.pack(side="right", fill="y", expand=False)
            self._dashboard_visible = True
            # Refresh charts after the canvases get their geometry
            if self.current_node:
                self.root.after_idle(
                    lambda n=self.current_node: self._focus_node(n))
                self.root.after(120, self._redraw_hero)
                self.root.after(120, self._redraw_top_folders)

    def _redraw_top_folders(self):
        cnv = getattr(self, "top_folders_canvas", None)
        if cnv is None or not self.scan_result:
            return
        cnv.delete("all")
        t = self.theme
        root = self.scan_result["root"]
        # Percent relative to root.size (disk_total when available)
        denom = root.size or self.scan_result["total_bytes"] or 1
        items = [(c.name, c.size, c.size / denom * 100, c.path)
                 for c in root.children[:8]]
        if not items:
            cnv.create_text(cnv.winfo_width() / 2, cnv.winfo_height() / 2,
                            text="No data", fill=t["muted"], font=UI_FONT)
            return
        w = cnv.winfo_width() or 600
        h = cnv.winfo_height() or 360
        max_v = max(it[1] for it in items) or 1
        n = len(items)
        gap = 10
        row_h = max(min(36, (h - gap * (n - 1)) / n), 16)
        for i, (name, size, pct, path) in enumerate(items):
            y = i * (row_h + gap)
            short = name if len(name) <= 30 else name[:27] + "…"
            cnv.create_text(0, y + row_h / 2, anchor="w",
                            text=short, fill=t["fg"], font=UI_FONT)
            tx = 170
            tw = max(w - tx - 130, 60)
            cnv.create_rectangle(tx, y + row_h / 2 - 4,
                                 tx + tw, y + row_h / 2 + 4,
                                 fill=t["panel_alt"], outline="")
            bw = max(tw * (size / max_v), 4)
            # Special colors for synthetic rows
            if path == "(free)":
                color = "#9CA3AF"
            elif path == "(gap)":
                color = t["warn"]
            else:
                color = PALETTE[i % len(PALETTE)]
            cnv.create_rectangle(tx, y + row_h / 2 - 4,
                                 tx + bw, y + row_h / 2 + 4,
                                 fill=color, outline="")
            cnv.create_text(tx + tw + 10, y + row_h / 2, anchor="w",
                            text=f"{human_size(size)}  ({pct:.1f}%)",
                            fill=t["fg"], font=UI_FONT_BOLD)

    def _build_panel_treemap(self, parent):
        t = self.theme
        panel = tk.Frame(parent, bg=t["bg"])
        wrap = tk.Frame(panel, bg=t["panel"], highlightthickness=1,
                        highlightbackground=t["border"])
        wrap.pack(fill="both", expand=True)
        head = tk.Frame(wrap, bg=t["panel"])
        head.pack(fill="x", padx=18, pady=(14, 6))
        tk.Label(head, text="Treemap", bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 12)).pack(side="left")
        tk.Label(head,
                 text="Click any tile to drill down  ·  Hover for details",
                 bg=t["panel"], fg=t["muted"],
                 font=UI_FONT).pack(side="left", padx=10)
        IconButton(head, t, "↑", "Up",
                   self._tree_up, variant="ghost").pack(side="right", padx=2)
        IconButton(head, t, "⌂", "Root",
                   self._tree_home, variant="ghost").pack(side="right", padx=2)
        # Color-mode toggle (depth-based or extension-based)
        self._treemap_color_btn = IconButton(
            head, t, "🎨", "Color: depth",
            self._toggle_treemap_color, variant="ghost")
        self._treemap_color_btn.pack(side="right", padx=2)
        self.treemap = TreemapChart(wrap, t, on_click=self._on_treemap_click)
        self.treemap.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        self._panels["treemap"] = panel

    def _toggle_treemap_color(self):
        """Flip between depth-based and extension-based tile colours."""
        cur = getattr(TreemapChart, "color_mode", "by_depth")
        new_mode = "by_extension" if cur == "by_depth" else "by_depth"
        TreemapChart.set_color_mode(new_mode)
        # Update the button label so the user knows the current state
        try:
            label_text = "Color: file type" if new_mode == "by_extension" else "Color: depth"
            self._treemap_color_btn.text_lbl.configure(text=label_text)
        except Exception:
            pass
        # Force a redraw of the treemap with the new colour scheme
        try:
            self.treemap.redraw()
        except Exception:
            pass

    def _build_panel_charts(self, parent):
        t = self.theme
        panel = tk.Frame(parent, bg=t["bg"])
        # Two cards stacked: Bar (top), Pie (bottom)
        bar_wrap = tk.Frame(panel, bg=t["panel"], highlightthickness=1,
                            highlightbackground=t["border"])
        bar_wrap.pack(fill="both", expand=True, pady=(0, 14))
        tk.Label(bar_wrap, text="Bar Chart — largest items",
                 bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 12)
                 ).pack(anchor="w", padx=18, pady=(14, 4))
        self.bar_chart = BarChart(bar_wrap, t)
        self.bar_chart.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        pie_wrap = tk.Frame(panel, bg=t["panel"], highlightthickness=1,
                            highlightbackground=t["border"])
        pie_wrap.pack(fill="both", expand=True)
        tk.Label(pie_wrap, text="Pie Chart — distribution",
                 bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 12)
                 ).pack(anchor="w", padx=18, pady=(14, 4))
        self.pie = PieChart(pie_wrap, t)
        self.pie.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        self._panels["charts"] = panel

    def _build_panel_duplicates(self, parent):
        t = self.theme
        panel = tk.Frame(parent, bg=t["bg"])
        wrap = tk.Frame(panel, bg=t["panel"], highlightthickness=1,
                        highlightbackground=t["border"])
        wrap.pack(fill="both", expand=True)
        head = tk.Frame(wrap, bg=t["panel"])
        head.pack(fill="x", padx=18, pady=(14, 6))
        tk.Label(head, text="Duplicate Files",
                 bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 12)).pack(side="left")
        IconButton(head, t, "🔍", "Find Duplicates",
                   self._find_duplicates, variant="primary"
                   ).pack(side="right", padx=2)
        # Threshold quick-control row
        thr_row = tk.Frame(wrap, bg=t["panel"])
        thr_row.pack(fill="x", padx=18, pady=(0, 4))
        tk.Label(thr_row, text="Min file size:",
                 bg=t["panel"], fg=t["muted"], font=UI_FONT
                 ).pack(side="left")
        cur_kb = max(1, ScanWorker.DUP_THRESHOLD // 1024)
        self.dup_thr_var = tk.StringVar(value=str(cur_kb))
        thr_entry = tk.Entry(thr_row, textvariable=self.dup_thr_var,
                             bg=t["panel_alt"], fg=t["fg"],
                             relief="flat", width=8,
                             highlightthickness=1,
                             highlightbackground=t["border_strong"],
                             highlightcolor=t["accent"],
                             font=UI_FONT)
        thr_entry.pack(side="left", padx=(6, 4))
        tk.Label(thr_row, text="KB",
                 bg=t["panel"], fg=t["muted"], font=UI_FONT
                 ).pack(side="left")

        def _apply_thr():
            try:
                kb = max(1, int(self.dup_thr_var.get().strip() or "256"))
            except ValueError:
                kb = 256
                self.dup_thr_var.set("256")
            ScanWorker.DUP_THRESHOLD = kb * 1024
            # (v4.3: results now populated by _on_dup_done directly)
            self.dup_status_var.set(
                f"Threshold set to {kb} KB. Click Find Duplicates to "
                f"re-scan for matching candidates.")
        IconButton(thr_row, t, "↻", "Apply",
                   _apply_thr, variant="ghost"
                   ).pack(side="left", padx=6)
        tk.Label(thr_row,
                 text="Lower = more candidates · slower hash step",
                 bg=t["panel"], fg=t["muted"], font=("Segoe UI", 9)
                 ).pack(side="left", padx=(12, 0))

        self.dup_status_var = tk.StringVar(
            value=f"Run a scan first, then click Find Duplicates "
                  f"(default ≥{cur_kb} KB).")
        tk.Label(wrap, textvariable=self.dup_status_var,
                 bg=t["panel"], fg=t["muted"], font=UI_FONT
                 ).pack(anchor="w", padx=18, pady=(0, 4))

        # Live progress indicator (visible only while a search runs).
        # Big horizontal blue bar + percentage label so the user never
        # wonders whether the search is alive.
        self.dup_progress_frame = tk.Frame(wrap, bg=t["panel"])
        # NOTE: not packed by default; _find_duplicates packs it.
        prog_row = tk.Frame(self.dup_progress_frame, bg=t["panel"])
        prog_row.pack(fill="x", padx=18, pady=(2, 0))
        self.dup_progress_bar = ttk.Progressbar(
            prog_row, mode="determinate", length=420,
            style="TProgressbar", maximum=100)
        self.dup_progress_bar.pack(side="left", fill="x", expand=True)
        self.dup_progress_pct = tk.Label(
            prog_row, text="0%", bg=t["panel"], fg=t["accent"],
            font=("Segoe UI Semibold", 11), width=6)
        self.dup_progress_pct.pack(side="left", padx=(10, 0))
        IconButton(prog_row, t, "✕", "Cancel",
                   self._cancel_duplicates, variant="ghost"
                   ).pack(side="left", padx=(8, 0))
        self.dup_progress_detail = tk.Label(
            self.dup_progress_frame,
            text="Hashing candidates…",
            bg=t["panel"], fg=t["muted"], font=UI_FONT)
        self.dup_progress_detail.pack(anchor="w", padx=18, pady=(4, 6))
        cont = tk.Frame(wrap, bg=t["panel"])
        cont.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        cols = ("size", "count", "waste")
        self.dup_tree = ttk.Treeview(cont, columns=cols,
                                     show="tree headings",
                                     selectmode="extended")
        self.dup_tree.heading("#0", text="Group / File")
        self.dup_tree.heading("size", text="File size")
        self.dup_tree.heading("count", text="Copies")
        self.dup_tree.heading("waste", text="Wasted")
        self.dup_tree.column("#0", width=560, anchor="w")
        self.dup_tree.column("size", width=110, anchor="e")
        self.dup_tree.column("count", width=80, anchor="e")
        self.dup_tree.column("waste", width=110, anchor="e")
        vsb = ttk.Scrollbar(cont, orient="vertical",
                            command=self.dup_tree.yview)
        self.dup_tree.configure(yscrollcommand=vsb.set)
        self.dup_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self._make_context_menu(self.dup_tree, kind="dup")
        self._panels["duplicates"] = panel

    def _build_panel_large(self, parent):
        t = self.theme
        panel = tk.Frame(parent, bg=t["bg"])

        # ----- KPI row -----
        kpi_row = tk.Frame(panel, bg=t["bg"])
        kpi_row.pack(fill="x", pady=(0, 14))
        self._large_kpi = {}
        kpi_defs = [
            ("Total Size",        "📄"),
            ("Files",             "📄"),
            ("Folders",           "📁"),
            ("Average File Size", "🥧"),
        ]
        for i, (lbl, icon) in enumerate(kpi_defs):
            card = tk.Frame(kpi_row, bg=t["panel"], highlightthickness=1,
                            highlightbackground=t["border"])
            card.pack(side="left", fill="both", expand=True,
                      padx=(0 if i == 0 else 7, 7 if i < 3 else 0))
            ci = tk.Frame(card, bg=t["panel"])
            ci.pack(fill="x", padx=20, pady=18)
            badge = tk.Label(ci, text=icon, bg=t["accent_soft"],
                             fg=t["accent"],
                             font=("Segoe UI Emoji", 16),
                             padx=12, pady=8)
            badge.pack(side="left", padx=(0, 16))
            tc = tk.Frame(ci, bg=t["panel"])
            tc.pack(side="left", fill="x", expand=True)
            tk.Label(tc, text=lbl, bg=t["panel"], fg=t["muted"],
                     font=UI_FONT).pack(anchor="w")
            val_lbl = tk.Label(tc, text="—", bg=t["panel"], fg=t["fg"],
                               font=("Segoe UI Semibold", 20))
            val_lbl.pack(anchor="w", pady=(2, 0))
            self._large_kpi[lbl] = val_lbl

        # ----- Table card -----
        wrap = tk.Frame(panel, bg=t["panel"], highlightthickness=1,
                        highlightbackground=t["border"])
        wrap.pack(fill="both", expand=True)
        cont = tk.Frame(wrap, bg=t["panel"])
        cont.pack(fill="both", expand=True, padx=10, pady=10)
        cols = ("idx", "name", "size", "age", "path")
        self.top_tree = ttk.Treeview(cont, columns=cols, show="headings", selectmode="extended")
        self.top_tree.heading("idx", text="#")
        self.top_tree.heading("name", text="File Name  ↕",
                              command=lambda: self._sort_treeview(
                                  self.top_tree, "name"))
        self.top_tree.heading("size", text="Size  ↕",
                              command=lambda: self._sort_treeview(
                                  self.top_tree, "size"))
        self.top_tree.heading("age", text="Age  ↕",
                              command=lambda: self._sort_treeview(
                                  self.top_tree, "age"))
        self.top_tree.heading("path", text="Path  ↕",
                              command=lambda: self._sort_treeview(
                                  self.top_tree, "path"))
        self.top_tree.column("idx", width=46, anchor="center", stretch=False)
        self.top_tree.column("name", width=280, anchor="w")
        self.top_tree.column("size", width=100, anchor="center",
                             stretch=False)
        self.top_tree.column("age", width=130, anchor="w", stretch=False)
        self.top_tree.column("path", width=600, anchor="w")
        # Subtle styling: pill-like size column via tag
        self.top_tree.tag_configure("largerow", foreground=t["fg"])
        self.top_tree.tag_configure("zebra",
                                    background=t["row_alt"],
                                    foreground=t["fg"])
        vsb = ttk.Scrollbar(cont, orient="vertical",
                            command=self.top_tree.yview)
        hsb = ttk.Scrollbar(cont, orient="horizontal",
                            command=self.top_tree.xview)
        self.top_tree.configure(yscrollcommand=vsb.set,
                                xscrollcommand=hsb.set)
        self.top_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        cont.rowconfigure(0, weight=1)
        cont.columnconfigure(0, weight=1)
        self.top_tree.bind("<Double-Button-1>",
                           lambda e: self._open_top_selected())
        self._make_context_menu(self.top_tree, kind="top")
        self._panels["large"] = panel

    @staticmethod
    def _file_emoji(path):
        """Return an emoji that represents the file's type."""
        ext = os.path.splitext(path)[1].lower()
        if ext in {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp",
                   ".svg", ".tiff", ".tif", ".ico", ".heic", ".raw"}:
            return "🖼"
        if ext in {".mp4", ".mov", ".avi", ".mkv", ".wmv", ".flv",
                   ".webm", ".m4v", ".mpg", ".mpeg", ".3gp"}:
            return "🎬"
        if ext in {".mp3", ".wav", ".flac", ".aac", ".ogg", ".m4a",
                   ".wma", ".opus", ".aiff"}:
            return "🎵"
        if ext in {".pdf"}:
            return "📕"
        if ext in {".doc", ".docx", ".odt", ".rtf"}:
            return "📘"
        if ext in {".xls", ".xlsx", ".ods", ".csv"}:
            return "📗"
        if ext in {".ppt", ".pptx", ".odp"}:
            return "📙"
        if ext in {".zip", ".rar", ".7z", ".tar", ".gz", ".bz2",
                   ".xz", ".iso"}:
            return "📦"
        if ext in {".exe", ".msi", ".dmg", ".apk", ".deb", ".rpm"}:
            return "⚡"
        if ext in {".dll", ".sys", ".so"}:
            return "🛠"
        if ext in {".vhd", ".vhdx", ".img", ".dmg"}:
            return "💿"
        if ext in {".db", ".sqlite", ".ost", ".pst"}:
            return "💾"
        if ext in {".py", ".js", ".ts", ".html", ".css", ".java",
                   ".c", ".cpp", ".h", ".cs", ".go", ".rs", ".rb",
                   ".php", ".swift", ".kt", ".sh"}:
            return "⚙"
        if ext in {".txt", ".md", ".log"}:
            return "📝"
        return "📄"

    def _build_panel_types(self, parent):
        t = self.theme
        panel = tk.Frame(parent, bg=t["bg"])
        # Top: category cards row
        self.cat_row = tk.Frame(panel, bg=t["bg"])
        self.cat_row.pack(fill="x", pady=(0, 14))
        # Bottom: table card
        wrap = tk.Frame(panel, bg=t["panel"], highlightthickness=1,
                        highlightbackground=t["border"])
        wrap.pack(fill="both", expand=True)
        head = tk.Frame(wrap, bg=t["panel"])
        head.pack(fill="x", padx=18, pady=(14, 6))
        tk.Label(head, text="All extensions",
                 bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 12)).pack(side="left")
        tk.Label(head, text="Sorted by total disk usage",
                 bg=t["panel"], fg=t["muted"],
                 font=UI_FONT).pack(side="left", padx=10)
        cont = tk.Frame(wrap, bg=t["panel"])
        cont.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        cols = ("ext", "category", "size", "percent", "count")
        self.ext_tree = ttk.Treeview(cont, columns=cols, show="headings", selectmode="extended")
        for c, txt, w, anc in [("ext", "Extension", 160, "w"),
                               ("category", "Category", 140, "w"),
                               ("size", "Total size", 130, "e"),
                               ("percent", "%", 80, "e"),
                               ("count", "Files", 100, "e")]:
            self.ext_tree.heading(c, text=txt)
            self.ext_tree.column(c, width=w, anchor=anc)
        vsb = ttk.Scrollbar(cont, orient="vertical",
                            command=self.ext_tree.yview)
        self.ext_tree.configure(yscrollcommand=vsb.set)
        self.ext_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self._panels["types"] = panel


    def _build_panel_settings(self, parent):
        t = self.theme
        panel = tk.Frame(parent, bg=t["bg"])
        wrap = tk.Frame(panel, bg=t["panel"], highlightthickness=1,
                        highlightbackground=t["border"])
        wrap.pack(fill="both", expand=True)
        tk.Label(wrap, text="Settings",
                 bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 14)
                 ).pack(anchor="w", padx=22, pady=(20, 4))
        tk.Label(wrap, text="Customize how Capacitra behaves.",
                 bg=t["panel"], fg=t["muted"], font=UI_FONT
                 ).pack(anchor="w", padx=22, pady=(0, 16))
        body = tk.Frame(wrap, bg=t["panel"])
        body.pack(fill="both", expand=True, padx=22, pady=(0, 22))

        def row(label, sub, action_label, action):
            r = tk.Frame(body, bg=t["panel_alt"],
                         highlightthickness=1,
                         highlightbackground=t["border"])
            r.pack(fill="x", pady=6, ipady=6)
            tcol = tk.Frame(r, bg=t["panel_alt"])
            tcol.pack(side="left", padx=16, pady=8, fill="x", expand=True)
            tk.Label(tcol, text=label, bg=t["panel_alt"], fg=t["fg"],
                     font=UI_FONT_BOLD).pack(anchor="w")
            tk.Label(tcol, text=sub, bg=t["panel_alt"], fg=t["muted"],
                     font=UI_FONT).pack(anchor="w")
            IconButton(r, t, "→", action_label, action,
                       variant="ghost").pack(side="right", padx=12)

        row("Theme",
            f"Currently using the {self.theme_name} theme.",
            "Toggle", self._toggle_theme)
        # Exclude paths editor
        ex_card = tk.Frame(body, bg=t["panel_alt"],
                           highlightthickness=1,
                           highlightbackground=t["border"])
        ex_card.pack(fill="x", pady=6, ipady=6)
        tk.Label(ex_card, text="Excluded folder names", bg=t["panel_alt"],
                 fg=t["fg"], font=UI_FONT_BOLD
                 ).pack(anchor="w", padx=16, pady=(10, 0))
        tk.Label(ex_card,
                 text="Folders with these names are skipped during scans. "
                      "One per line.",
                 bg=t["panel_alt"], fg=t["muted"], font=UI_FONT
                 ).pack(anchor="w", padx=16, pady=(2, 6))
        self.ex_text = tk.Text(ex_card, height=4, bg=t["panel"],
                               fg=t["fg"], relief="flat",
                               highlightthickness=1,
                               highlightbackground=t["border_strong"],
                               highlightcolor=t["accent"],
                               font=("Consolas", 9))
        self.ex_text.pack(fill="x", padx=16, pady=(0, 6))
        self.ex_text.insert("1.0", "\n".join(self._excludes))

        def save_excludes():
            txt = self.ex_text.get("1.0", "end").strip()
            self._excludes = [l.strip() for l in txt.splitlines()
                              if l.strip()]
            self.status_var.set(
                f"Excludes saved ({len(self._excludes)} item(s)).")
        save_btn = tk.Frame(ex_card, bg=t["panel_alt"])
        save_btn.pack(fill="x", padx=16, pady=(0, 8))
        IconButton(save_btn, t, "💾", "Save excludes",
                   save_excludes, variant="primary"
                   ).pack(side="right")
        # Duplicate detection threshold — inline editor
        thr_card = tk.Frame(body, bg=t["panel_alt"],
                            highlightthickness=1,
                            highlightbackground=t["border"])
        thr_card.pack(fill="x", pady=6, ipady=6)
        thr_in = tk.Frame(thr_card, bg=t["panel_alt"])
        thr_in.pack(fill="x", padx=16, pady=10)
        tk.Label(thr_in, text="Duplicate-detection threshold",
                 bg=t["panel_alt"], fg=t["fg"], font=UI_FONT_BOLD
                 ).pack(anchor="w")
        tk.Label(thr_in, text="Files smaller than this are skipped when "
                              "searching for duplicates. Bigger value = "
                              "faster scan, fewer candidates.",
                 bg=t["panel_alt"], fg=t["muted"],
                 font=UI_FONT, wraplength=720, justify="left"
                 ).pack(anchor="w", pady=(2, 8))
        thr_row = tk.Frame(thr_in, bg=t["panel_alt"])
        thr_row.pack(anchor="w")
        # Default value in MB
        current_mb = ScanWorker.DUP_THRESHOLD // (1024 * 1024)
        self._dup_threshold_var = tk.StringVar(value=str(current_mb))
        tk.Entry(thr_row, textvariable=self._dup_threshold_var,
                 width=8, bg=t["panel"], fg=t["fg"], relief="flat",
                 highlightthickness=1,
                 highlightbackground=t["border_strong"],
                 highlightcolor=t["accent"],
                 font=("Segoe UI", 10)).pack(side="left", ipady=4)
        tk.Label(thr_row, text="  MB", bg=t["panel_alt"], fg=t["muted"],
                 font=UI_FONT).pack(side="left")

        def apply_threshold():
            try:
                mb = float(self._dup_threshold_var.get())
                if mb < 0.001 or mb > 100000:
                    raise ValueError
                ScanWorker.DUP_THRESHOLD = int(mb * 1024 * 1024)
                self.status_var.set(
                    f"Duplicate threshold set to {mb:g} MB. "
                    "Re-run the scan for the new threshold to take effect.")
                self.status_icon.configure(fg=self.theme["success"])
            except ValueError:
                self.status_var.set(
                    "Invalid threshold value — enter a number in MB.")
                self.status_icon.configure(fg=self.theme["danger"])
        IconButton(thr_in, t, "✓", "Apply threshold",
                   apply_threshold, variant="primary"
                   ).pack(anchor="e", pady=(8, 0))

        # Run as Administrator — explanation only, no popup
        admin_card = tk.Frame(body, bg=t["panel_alt"],
                              highlightthickness=1,
                              highlightbackground=t["border"])
        admin_card.pack(fill="x", pady=6, ipady=6)
        ai = tk.Frame(admin_card, bg=t["panel_alt"])
        ai.pack(fill="x", padx=16, pady=12)
        tk.Label(ai, text="Run as Administrator",
                 bg=t["panel_alt"], fg=t["fg"],
                 font=UI_FONT_BOLD).pack(anchor="w")
        tk.Label(ai,
                 text="Some Windows folders (System Volume Information, "
                      "$Recycle.Bin, other users' profiles, locked system "
                      "files like pagefile.sys / hiberfil.sys) are only "
                      "readable by an elevated process.\n\n"
                      "To include them: close Capacitra, right-click the "
                      "executable (or Capacitra.pyw), and choose "
                      "“Run as administrator”.",
                 bg=t["panel_alt"], fg=t["muted"],
                 font=UI_FONT, wraplength=720, justify="left"
                 ).pack(anchor="w", pady=(2, 0))

        self._panels["settings"] = panel

    def _build_panel_history(self, parent):
        t = self.theme
        panel = tk.Frame(parent, bg=t["bg"])
        wrap = tk.Frame(panel, bg=t["panel"], highlightthickness=1,
                        highlightbackground=t["border"])
        wrap.pack(fill="both", expand=True)
        tk.Label(wrap, text="Scan History",
                 bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 14)
                 ).pack(anchor="w", padx=22, pady=(20, 4))
        tk.Label(wrap, text="Recent scans completed in this session.",
                 bg=t["panel"], fg=t["muted"], font=UI_FONT
                 ).pack(anchor="w", padx=22, pady=(0, 16))
        cont = tk.Frame(wrap, bg=t["panel"])
        cont.pack(fill="both", expand=True, padx=12, pady=(0, 16))
        cols = ("when", "scanned", "files", "folders", "denied")
        self.history_tree = ttk.Treeview(cont, columns=cols, show="tree headings")
        for c, txt, w, anc in [
            ("when", "When", 160, "w"),
            ("scanned", "Scanned size", 140, "e"),
            ("files", "Files", 100, "e"),
            ("folders", "Folders", 100, "e"),
            ("denied", "Skipped", 100, "e"),
        ]:
            self.history_tree.heading(c, text=txt)
            self.history_tree.column(c, width=w, anchor=anc)
        self.history_tree.heading("#0", text="Path")
        self.history_tree.column("#0", width=400, minwidth=200, anchor="w", stretch=True)
        vsb = ttk.Scrollbar(cont, orient="vertical",
                            command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=vsb.set)
        self.history_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self._panels["history"] = panel
        if not hasattr(self, "_scan_history"):
            self._scan_history = []

    def _record_history(self, result):
        if not hasattr(self, "_scan_history"):
            self._scan_history = []
        self._scan_history.insert(0, {
            "when": time.strftime("%Y-%m-%d %H:%M"),
            "path": self.scan_root,
            "scanned": result["total_bytes"],
            "files": result["total_files"],
            "folders": result.get("total_dirs", 0),
            "denied": result.get("denied", 0),
        })
        if hasattr(self, "history_tree"):
            self.history_tree.delete(*self.history_tree.get_children())
            for h in self._scan_history[:50]:
                self.history_tree.insert(
                    "", "end",
                    text=h["path"],
                    values=(h["when"], human_size(h["scanned"]),
                            f"{h['files']:,}", f"{h['folders']:,}",
                            f"{h['denied']:,}"))

    def _build_panel_export(self, parent):
        t = self.theme
        panel = tk.Frame(parent, bg=t["bg"])
        # Header card
        hero = tk.Frame(panel, bg=t["panel"], highlightthickness=1,
                        highlightbackground=t["border"])
        hero.pack(fill="x", pady=(0, 14))
        h_in = tk.Frame(hero, bg=t["panel"])
        h_in.pack(fill="x", padx=22, pady=20)
        tk.Label(h_in, text="Export reports", bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 16)).pack(anchor="w")
        tk.Label(h_in,
                 text="Pick a format below to save a snapshot of the "
                      "current scan. All exports are written to a local "
                      "file you choose, nothing is uploaded anywhere.",
                 bg=t["panel"], fg=t["muted"], font=UI_FONT,
                 wraplength=820, justify="left"
                 ).pack(anchor="w", pady=(4, 0))

        # Status line for export state
        self.export_status_var = tk.StringVar(
            value="Run a scan first to enable exports.")
        tk.Label(h_in, textvariable=self.export_status_var,
                 bg=t["panel"], fg=t["accent"], font=UI_FONT
                 ).pack(anchor="w", pady=(10, 0))

        # Grid of format cards
        grid = tk.Frame(panel, bg=t["bg"])
        grid.pack(fill="both", expand=True)
        grid.grid_columnconfigure(0, weight=1, uniform="ex")
        grid.grid_columnconfigure(1, weight=1, uniform="ex")

        def card(row, col, icon, title, desc, badge, badge_color,
                 action, enabled=True):
            c = tk.Frame(grid, bg=t["panel"], highlightthickness=1,
                         highlightbackground=t["border"])
            c.grid(row=row, column=col, sticky="nsew",
                   padx=(0 if col == 0 else 7, 7 if col == 0 else 0),
                   pady=7, ipady=4)
            inner = tk.Frame(c, bg=t["panel"])
            inner.pack(fill="both", expand=True, padx=22, pady=18)
            top = tk.Frame(inner, bg=t["panel"])
            top.pack(fill="x")
            ico = tk.Label(top, text=icon, bg=t["accent_soft"],
                           fg=t["accent"],
                           font=("Segoe UI Emoji", 18),
                           padx=10, pady=4)
            ico.pack(side="left")
            if badge:
                tk.Label(top, text=badge, bg=badge_color, fg="#FFF",
                         font=("Segoe UI Semibold", 8),
                         padx=8, pady=3
                         ).pack(side="right", anchor="ne")
            tk.Label(inner, text=title, bg=t["panel"], fg=t["fg"],
                     font=("Segoe UI Semibold", 13)
                     ).pack(anchor="w", pady=(12, 4))
            tk.Label(inner, text=desc, bg=t["panel"], fg=t["muted"],
                     font=UI_FONT, wraplength=360, justify="left"
                     ).pack(anchor="w", pady=(0, 14))
            btn = IconButton(
                inner, t, "💾",
                "Save…" if enabled else "Unavailable",
                action if enabled else (lambda: None),
                variant="primary" if enabled else "ghost",
            )
            btn.pack(anchor="w")
            return c

        card(0, 0, "📄", "CSV",
             "Comma-separated table of every folder and file. Opens "
             "in Excel, Numbers or any data tool. Best for further "
             "analysis or scripting.",
             None, None,
             self._export_csv, enabled=True)
        card(0, 1, "🌐", "HTML",
             "Self-contained HTML report with bars and the folder "
             "tree. Share by email or open in any browser. No "
             "external assets.",
             None, None,
             self._export_html, enabled=True)
        card(1, 0, "📕", "PDF",
             "Printable, paginated report with the dashboard summary "
             "and top folders. Ideal for archiving or sending to a "
             "manager.",
             None, None,
             self._export_pdf, enabled=True)
        card(1, 1, "📊", "Excel",
             "Native .xlsx workbook with formatted sheets for folders, "
             "duplicates and large files. Drop straight into your "
             "reporting flow.",
             None, None,
             self._export_excel, enabled=True)

        # Tips footer
        tips = tk.Frame(panel, bg=t["panel_alt"], highlightthickness=1,
                        highlightbackground=t["border"])
        tips.pack(fill="x", pady=(14, 0))
        ti = tk.Frame(tips, bg=t["panel_alt"])
        ti.pack(fill="x", padx=22, pady=14)
        tk.Label(ti, text="💡  Tip",
                 bg=t["panel_alt"], fg=t["accent"],
                 font=("Segoe UI Semibold", 11)).pack(anchor="w")
        tk.Label(ti,
                 text=("Press Ctrl+E from any panel to open this Export "
                       "screen quickly. Exports always reflect the most "
                       "recent scan plus any filters you have applied."),
                 bg=t["panel_alt"], fg=t["fg_subtle"],
                 font=UI_FONT, wraplength=820, justify="left"
                 ).pack(anchor="w", pady=(4, 0))

        self._panels["export"] = panel

    def _refresh_export_status(self):
        """Called after a scan to update the Export panel status line."""
        if not hasattr(self, "export_status_var"):
            return
        if not self.scan_result:
            self.export_status_var.set("Run a scan first to enable exports.")
        else:
            root = self.scan_result.get("root")
            total = self.scan_result.get("total_bytes", 0)
            self.export_status_var.set(
                f"Ready · {root.name if root else 'scan'} · "
                f"{human_size(total)} of data ready to export.")

    def _build_panel_about(self, parent):
        t = self.theme
        panel = tk.Frame(parent, bg=t["bg"])
        # Outer scrollable canvas so the About content is never cut off
        # on shorter screens. The inner shell is centered with a max
        # width of 820 px for readable line lengths.
        outer = tk.Canvas(panel, bg=t["bg"], highlightthickness=0,
                          borderwidth=0)
        outer.pack(side="left", fill="both", expand=True)
        ysb = ttk.Scrollbar(panel, orient="vertical", command=outer.yview)
        ysb.pack(side="right", fill="y")
        outer.configure(yscrollcommand=ysb.set)
        inner = tk.Frame(outer, bg=t["bg"])
        outer_win = outer.create_window((0, 0), window=inner, anchor="nw")

        def _resize_inner(event=None):
            outer.configure(scrollregion=outer.bbox("all"))
            # Make inner frame match canvas width so content centers properly
            try:
                outer.itemconfigure(outer_win, width=outer.winfo_width())
            except tk.TclError:
                pass
        inner.bind("<Configure>", _resize_inner)
        outer.bind("<Configure>", _resize_inner)

        # Mouse wheel scrolling — bound only while pointer is over the
        # About panel, so other panels keep their normal wheel behaviour.
        def _on_wheel(e):
            outer.yview_scroll(int(-1 * (e.delta / 120)), "units")
        def _bind_wheel(_e):
            outer.bind_all("<MouseWheel>", _on_wheel)
        def _unbind_wheel(_e):
            outer.unbind_all("<MouseWheel>")
        outer.bind("<Enter>", _bind_wheel)
        outer.bind("<Leave>", _unbind_wheel)
        inner.bind("<Enter>", _bind_wheel)
        inner.bind("<Leave>", _unbind_wheel)

        # Center column inside the scrollable inner frame
        inner.grid_columnconfigure(0, weight=1)
        inner.grid_columnconfigure(1, weight=0, minsize=1080)
        inner.grid_columnconfigure(2, weight=1)
        shell = tk.Frame(inner, bg=t["bg"])
        shell.grid(row=0, column=1, sticky="nsew", pady=(0, 28))
        # Hero band with logo + name
        hero = tk.Frame(shell, bg=t["panel"], highlightthickness=1,
                        highlightbackground=t["border"])
        hero.pack(fill="x", pady=(0, 14))
        h_in = tk.Frame(hero, bg=t["panel"])
        h_in.pack(fill="x", padx=28, pady=24)
        logo_canvas = tk.Canvas(h_in, width=72, height=72, bg=t["panel"],
                                highlightthickness=0)
        logo_canvas.pack(side="left", padx=(0, 20))
        self._draw_capacitra_mark(logo_canvas, 36, 36, 32)
        col = tk.Frame(h_in, bg=t["panel"])
        col.pack(side="left", anchor="w")
        tk.Label(col, text=APP_NAME, bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 22)).pack(anchor="w")
        tk.Label(col, text=APP_TAGLINE, bg=t["panel"], fg=t["accent"],
                 font=("Segoe UI", 12)).pack(anchor="w", pady=(2, 0))
        tk.Label(col, text=f"Version {APP_VERSION}  ·  Build 2026.07",
                 bg=t["panel"], fg=t["muted"],
                 font=UI_FONT).pack(anchor="w", pady=(8, 0))

        # Body card with corporate description
        body = tk.Frame(shell, bg=t["panel"], highlightthickness=1,
                        highlightbackground=t["border"])
        body.pack(fill="both", expand=True)
        b_in = tk.Frame(body, bg=t["panel"])
        b_in.pack(fill="both", expand=True, padx=28, pady=22)

        def section(title, paras):
            tk.Label(b_in, text=title, bg=t["panel"], fg=t["fg"],
                     font=("Segoe UI Semibold", 13)
                     ).pack(anchor="w", pady=(10, 6))
            for p in paras:
                tk.Label(b_in, text=p, bg=t["panel"], fg=t["fg_subtle"],
                         font=UI_FONT, wraplength=980, justify="left"
                         ).pack(anchor="w", pady=(0, 6))

        section("What is Capacitra?", [
            "Capacitra is a storage capacity intelligence platform "
            "designed for individuals, IT administrators and "
            "enterprises who need to understand, and reclaim, every "
            "byte on their disks.",
            "From a single workstation to fleets of servers, Capacitra "
            "delivers fast, transparent insight into where storage is "
            "being consumed and what can safely be cleaned up.",
        ])
        section("Privacy by design", [
            "Capacitra runs entirely on your machine. Your scan "
            "results never leave your device. No telemetry, no cloud "
            "uploads, no third-party tracking.",
            "We believe storage analytics is a local concern, not a "
            "service to subscribe to. Every byte we read stays in "
            "your hands.",
        ])
        section("Built for performance", [
            "A multi-threaded scanning engine processes terabyte-scale "
            "volumes in minutes, with a live progress indicator and a "
            "responsive interface throughout the scan.",
            "Free space, accessible data and inaccessible system "
            "regions are all accounted for, so the numbers add up to "
            "the full capacity of the disk.",
        ])
        section("Designed for Windows", [
            "Capacitra is purpose-built for Windows 10 and 11, with a "
            "lean installer footprint and a single-executable distribution.",
            "Network shares, removable drives and mounted images are "
            "all fully supported inside the Windows shell.",
        ])
        section("Capabilities at a glance", [
            "Treemap, bar, pie and folder-tree visualizations · "
            "duplicate detection with SHA-1 verification · file-age "
            "cohorts · file-type categorization · large-file finder · "
            "inline size bars · CSV, HTML, PDF and Excel report export · "
            "dark and light themes · keyboard shortcuts · exclude paths · "
            "advanced filtering syntax (size, age, regex).",
        ])

        # Resources card with clickable links
        rs = tk.Frame(shell, bg=t["panel"], highlightthickness=1,
                      highlightbackground=t["border"])
        rs.pack(fill="x", pady=(14, 0))
        ri = tk.Frame(rs, bg=t["panel"])
        ri.pack(fill="x", padx=28, pady=18)
        tk.Label(ri, text="Resources", bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 13)
                 ).pack(anchor="w", pady=(0, 10))
        link_rows = [
            ("🌐  Website",        "https://capacitra.com"),
            ("📚  Documentation",  "https://capacitra.com/docs.html"),
            ("✉  Support",         "mailto:info@capacitra.com"),
            ("🔒  Privacy policy", "https://capacitra.com/privacy.html"),
        ]
        for lbl, url in link_rows:
            row = tk.Frame(ri, bg=t["panel"])
            row.pack(fill="x", pady=2)
            tk.Label(row, text=lbl, bg=t["panel"], fg=t["fg"],
                     font=UI_FONT, width=20, anchor="w"
                     ).pack(side="left")
            link = tk.Label(row, text=url, bg=t["panel"],
                            fg=t["accent"], font=UI_FONT,
                            cursor="hand2")
            link.pack(side="left")
            link.bind("<Button-1>",
                      lambda e, u=url: self._open_link(u))

        # Footer
        foot = tk.Frame(shell, bg=t["bg"])
        foot.pack(fill="x", pady=(14, 6))
        tk.Label(foot,
                 text="© 2026 Capacitra. All rights reserved.",
                 bg=t["bg"], fg=t["muted"],
                 font=("Segoe UI", 9)).pack(side="left")
        tk.Label(foot,
                 text="Made with care · No trackers, no cookies, no nonsense",
                 bg=t["bg"], fg=t["muted"],
                 font=("Segoe UI", 9)).pack(side="right")

        self._panels["about"] = panel

    def _open_find_dialog(self):
        """Modal global Find: type a substring, see all matching files
        and folders across the whole scan, double-click to jump to that
        node in the main tree."""
        if not self.scan_result:
            self._warn("Run a scan first",
                       "Pick a drive or folder and click New Scan, "
                       "then come back to search.")
            return
        t = self.theme
        win = tk.Toplevel(self.root)
        win.title("Find in scan results")
        win.configure(bg=t["panel"])
        win.transient(self.root)
        try:
            if getattr(self.root, "_capacitra_icon", None) is not None:
                win.iconphoto(False, self.root._capacitra_icon)
        except Exception:
            pass

        outer = tk.Frame(win, bg=t["panel"])
        outer.pack(fill="both", expand=True, padx=20, pady=18)

        # Header + input
        tk.Label(outer, text="🔍  Find file or folder by name",
                 bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 13)).pack(anchor="w")
        tk.Label(outer,
                 text="Case-insensitive substring match across all "
                      "scanned paths. Double-click a result to jump "
                      "to it in the Folder Tree.",
                 bg=t["panel"], fg=t["muted"], font=UI_FONT,
                 wraplength=560, justify="left"
                 ).pack(anchor="w", pady=(2, 10))

        q_var = tk.StringVar()
        entry = tk.Entry(outer, textvariable=q_var,
                         bg=t["panel_alt"], fg=t["fg"],
                         relief="flat",
                         highlightthickness=1,
                         highlightbackground=t["border_strong"],
                         highlightcolor=t["accent"],
                         font=("Segoe UI", 11))
        entry.pack(fill="x", ipady=6)
        entry.focus_set()

        # Status / counter
        status_var = tk.StringVar(value="Type to search…")
        tk.Label(outer, textvariable=status_var,
                 bg=t["panel"], fg=t["muted"], font=UI_FONT
                 ).pack(anchor="w", pady=(6, 4))

        # Results table
        cols = ("size", "path")
        table = ttk.Treeview(outer, columns=cols, show="tree headings",
                             selectmode="browse", height=14)
        table.heading("#0",   text="Name", anchor="w")
        table.heading("size", text="Size",  anchor="e")
        table.heading("path", text="Path",  anchor="w")
        table.column("#0",   width=240, anchor="w")
        table.column("size", width=90,  anchor="e", stretch=False)
        table.column("path", width=520, anchor="w")
        vsb = ttk.Scrollbar(outer, orient="vertical", command=table.yview)
        table.configure(yscrollcommand=vsb.set)
        table.pack(side="left", fill="both", expand=True, pady=(4, 0))
        vsb.pack(side="right", fill="y", pady=(4, 0))

        # In-memory results cache for the current query
        results = []  # list of Node

        def _collect(node, q, found, limit=500):
            """Depth-first walk of the scan tree, collect matches up to limit."""
            if len(found) >= limit:
                return
            if q in node.name.lower():
                found.append(node)
                if len(found) >= limit:
                    return
            for c in node.children:
                if c.path in ("(free)", "(gap)"):
                    continue
                _collect(c, q, found, limit)

        def _on_query(*_a):
            q = q_var.get().strip().lower()
            table.delete(*table.get_children())
            results.clear()
            if not q:
                status_var.set("Type to search…")
                return
            root = self.scan_result["root"]
            _collect(root, q, results, limit=500)
            for n in results:
                icon = "📁 " if n.is_dir else "📄 "
                table.insert("", "end",
                             text=icon + n.name,
                             values=(human_size(n.size), n.path),
                             tags=("res",))
            count = len(results)
            if count >= 500:
                status_var.set("Showing first 500 matches. Narrow the query for fewer results.")
            elif count == 0:
                status_var.set(f"No matches for '{q_var.get()}'.")
            else:
                status_var.set(f"{count} match{'es' if count != 1 else ''}.")

        # Debounce keystrokes ~250 ms so a 5M-node scan does not
        # re-walk the whole tree on every character typed.
        _debounce_id = {"id": None}
        def _schedule(*_a):
            if _debounce_id["id"] is not None:
                try:
                    win.after_cancel(_debounce_id["id"])
                except Exception:
                    pass
            _debounce_id["id"] = win.after(250, _on_query)
        q_var.trace_add("write", _schedule)

        def _jump_to_selected(_evt=None):
            sel = table.selection()
            if not sel:
                return
            idx = table.index(sel[0])
            if idx < 0 or idx >= len(results):
                return
            target_node = results[idx]
            # Switch to Overview panel so the tree is visible
            self._select_panel("overview")
            # Walk up from the target and expand each ancestor so the
            # target row is visible in the main tree.
            ancestors = []
            n = target_node
            while n is not None:
                ancestors.append(n)
                n = n.parent
            ancestors.reverse()
            for anc in ancestors:
                iid = self._node_iid_map.get(anc)
                if iid:
                    try:
                        self.tree.item(iid, open=True)
                        if anc.is_dir:
                            self._expand_iid(iid, anc)
                    except tk.TclError:
                        pass
            final_iid = self._node_iid_map.get(target_node)
            if final_iid:
                try:
                    self.tree.see(final_iid)
                    self.tree.selection_set(final_iid)
                    self.tree.focus(final_iid)
                except tk.TclError:
                    pass
            win.destroy()

        table.bind("<Double-Button-1>", _jump_to_selected)
        table.bind("<Return>",          _jump_to_selected)

        # Center the dialog over the root window
        win.update_idletasks()
        rw, rh = self.root.winfo_width(), self.root.winfo_height()
        rx, ry = self.root.winfo_rootx(), self.root.winfo_rooty()
        ww, wh = 720, 560
        x = rx + max(0, (rw - ww) // 2)
        y = ry + max(0, (rh - wh) // 3)
        win.geometry(f"{ww}x{wh}+{x}+{y}")

        win.bind("<Escape>", lambda e: win.destroy())

    def _on_close(self):
        """Window close handler. If a scan or duplicate search is
        running, ask the user to confirm. Otherwise quit immediately."""
        scan_alive = (self.scan_thread is not None
                      and self.scan_thread.is_alive())
        dup_alive = (self.dup_thread is not None
                     and self.dup_thread.is_alive())
        if scan_alive or dup_alive:
            what = []
            if scan_alive: what.append("a disk scan")
            if dup_alive: what.append("a duplicate search")
            label = " and ".join(what)
            if not self._ask(
                "Operation still running",
                f"Capacitra is still running {label}. "
                "Closing now will cancel it and discard partial results.\n\n"
                "Quit anyway?",
                yes="Quit", no="Keep running"):
                return
            # User confirmed: signal both workers to stop cleanly
            try:
                if self.stop_event: self.stop_event.set()
            except Exception:
                pass
            try:
                if self.dup_stop: self.dup_stop.set()
            except Exception:
                pass
        # Tear down
        self._alive = False
        try:
            self.root.destroy()
        except Exception:
            pass

    def _open_link(self, url):
        """Open a URL in the user's default browser or mail client."""
        import webbrowser
        try:
            webbrowser.open(url)
        except Exception:
            self.root.clipboard_clear()
            self.root.clipboard_append(url)
            self.status_var.set(f"Copied to clipboard: {url}")

    # ----- Branded dialogs (replace tkinter messagebox popups) -----
    def _branded_dialog(self, kind, title, message, yes_text="OK",
                        no_text=None):
        """Show a Capacitra-branded modal dialog with our own theme,
        hexagon logo, and clean buttons. Returns True if the user
        clicked the primary button, False otherwise.

        kind: "info", "warn", "error", "ask"
        """
        t = self.theme
        win = tk.Toplevel(self.root)
        win.title(title)
        win.configure(bg=t["panel"])
        win.transient(self.root)
        win.resizable(False, False)
        # Inherit the root window's branded icon
        try:
            if hasattr(self.root, "_capacitra_icon") and \
                    self.root._capacitra_icon is not None:
                win.iconphoto(False, self.root._capacitra_icon)
        except Exception:
            pass

        accent = {
            "info":  t["accent"],
            "warn":  t.get("warn", "#B45309"),
            "error": t.get("danger", "#DC2626"),
            "ask":   t["accent"],
        }.get(kind, t["accent"])

        glyph = {
            "info": "ⓘ", "warn": "⚠", "error": "✕", "ask": "?",
        }.get(kind, "ⓘ")

        outer = tk.Frame(win, bg=t["panel"])
        outer.pack(fill="both", expand=True, padx=28, pady=24)

        head = tk.Frame(outer, bg=t["panel"])
        head.pack(fill="x")
        badge = tk.Label(head, text=glyph, bg=accent, fg="#FFFFFF",
                         font=("Segoe UI", 18, "bold"),
                         padx=12, pady=4)
        badge.pack(side="left", padx=(0, 14))
        tk.Label(head, text=title, bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 14)
                 ).pack(side="left", anchor="w")

        body = tk.Frame(outer, bg=t["panel"])
        body.pack(fill="x", pady=(16, 20))
        tk.Label(body, text=message, bg=t["panel"], fg=t["fg_subtle"],
                 font=UI_FONT, wraplength=380, justify="left"
                 ).pack(anchor="w")

        # Buttons row
        btns = tk.Frame(outer, bg=t["panel"])
        btns.pack(fill="x")
        result = {"value": False}

        def _close(val):
            result["value"] = val
            try:
                win.grab_release()
            except tk.TclError:
                pass
            win.destroy()

        if no_text:
            IconButton(btns, t, "✕", no_text,
                       lambda: _close(False), variant="ghost"
                       ).pack(side="right", padx=(8, 0))
        primary_glyph = "✓" if kind != "error" else "OK"
        IconButton(btns, t, primary_glyph, yes_text,
                   lambda: _close(True), variant="primary"
                   ).pack(side="right")

        # Center over root window
        win.update_idletasks()
        rw = self.root.winfo_width()
        rh = self.root.winfo_height()
        rx = self.root.winfo_rootx()
        ry = self.root.winfo_rooty()
        ww = max(440, win.winfo_reqwidth())
        wh = max(180, win.winfo_reqheight())
        x = rx + max(0, (rw - ww) // 2)
        y = ry + max(0, (rh - wh) // 3)
        win.geometry(f"{ww}x{wh}+{x}+{y}")

        win.bind("<Escape>", lambda e: _close(False))
        win.bind("<Return>", lambda e: _close(True))
        try:
            win.grab_set()
        except tk.TclError:
            pass
        win.focus_set()
        self.root.wait_window(win)
        return result["value"]

    def _info(self, title, msg):
        self._branded_dialog("info", title, msg)

    def _warn(self, title, msg):
        self._branded_dialog("warn", title, msg)

    def _error(self, title, msg):
        self._branded_dialog("error", title, msg)

    def _ask(self, title, msg, yes="Yes", no="No"):
        return self._branded_dialog("ask", title, msg,
                                    yes_text=yes, no_text=no)

    # ----- File / folder properties dialog -----
    def _show_properties_dialog(self, path):
        try:
            st = os.stat(path)
        except OSError as e:
            self._error("Properties", f"Cannot read:\n{e}")
            return
        t = self.theme
        is_dir = os.path.isdir(path)
        # Win-specific attributes (read-only, hidden, system, compressed, …)
        attrs = []
        if sys.platform == "win32":
            try:
                from ctypes import windll
                GetFileAttributesW = windll.kernel32.GetFileAttributesW
                a = GetFileAttributesW(path)
                if a not in (0, 0xFFFFFFFF):
                    if a & 0x01:    attrs.append("Read-only")
                    if a & 0x02:    attrs.append("Hidden")
                    if a & 0x04:    attrs.append("System")
                    if a & 0x20:    attrs.append("Archive")
                    if a & 0x100:   attrs.append("Temporary")
                    if a & 0x400:   attrs.append("Reparse point")
                    if a & 0x800:   attrs.append("NTFS compressed")
                    if a & 0x4000:  attrs.append("Encrypted")
            except Exception:
                pass

        win = tk.Toplevel(self.root)
        win.title("Properties — " + (os.path.basename(path) or path))
        win.minsize(520, 420)
        win.geometry("600x520")
        win.configure(bg=t["bg"])
        # Header
        head = tk.Frame(win, bg=t["panel"], highlightthickness=1,
                        highlightbackground=t["border"])
        head.pack(fill="x", padx=14, pady=(14, 8))
        hi = tk.Frame(head, bg=t["panel"])
        hi.pack(fill="x", padx=18, pady=14)
        emoji = "📁" if is_dir else self._file_emoji(path)
        tk.Label(hi, text=emoji, bg=t["panel"], fg=t["accent"],
                 font=("Segoe UI Emoji", 26)).pack(side="left", padx=(0, 14))
        tc = tk.Frame(hi, bg=t["panel"])
        tc.pack(side="left", fill="x", expand=True)
        tk.Label(tc, text=os.path.basename(path) or path,
                 bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 14)
                 ).pack(anchor="w")
        tk.Label(tc, text=path, bg=t["panel"], fg=t["muted"],
                 font=("Consolas", 9)).pack(anchor="w", pady=(2, 0))
        # Body
        body = tk.Frame(win, bg=t["panel"], highlightthickness=1,
                        highlightbackground=t["border"])
        body.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        bi = tk.Frame(body, bg=t["panel"])
        bi.pack(fill="both", expand=True, padx=22, pady=18)

        def row(label, value):
            r = tk.Frame(bi, bg=t["panel"])
            r.pack(fill="x", pady=4)
            tk.Label(r, text=label, bg=t["panel"], fg=t["muted"],
                     font=UI_FONT, width=20, anchor="w"
                     ).pack(side="left")
            tk.Label(r, text=value, bg=t["panel"], fg=t["fg"],
                     font=UI_FONT, anchor="w", justify="left",
                     wraplength=380).pack(side="left", fill="x", expand=True)

        row("Type", "Folder" if is_dir else "File")
        row("Size", f"{human_size(st.st_size)}  ({st.st_size:,} bytes)")
        row("Last modified",
            time.strftime("%Y-%m-%d %H:%M:%S",
                          time.localtime(st.st_mtime)))
        row("Last accessed",
            time.strftime("%Y-%m-%d %H:%M:%S",
                          time.localtime(st.st_atime)))
        try:
            row("Created",
                time.strftime("%Y-%m-%d %H:%M:%S",
                              time.localtime(st.st_ctime)))
        except Exception:
            pass
        if hasattr(st, "st_nlink") and st.st_nlink > 1:
            row("Hard link count", str(st.st_nlink))
        if attrs:
            row("Attributes", "  ·  ".join(attrs))
        if hasattr(st, "st_uid"):
            try:
                import pwd
                owner = pwd.getpwuid(st.st_uid).pw_name
                row("Owner", f"{owner} ({st.st_uid})")
            except Exception:
                row("Owner UID", str(st.st_uid))
        # Footer with Close button
        foot = tk.Frame(win, bg=t["bg"])
        foot.pack(fill="x", padx=14, pady=(0, 14))
        IconButton(foot, t, "✕", "Close",
                   win.destroy, variant="primary"
                   ).pack(side="right")

    # ----- Snapshot save / load / compare -----
    # Magic header so we can reject random files before pickle touches them.
    _SNAPSHOT_MAGIC = b"CAPSNAP1"

    @staticmethod
    def _safe_unpickler(fp):
        """Return a pickle.Unpickler that refuses to instantiate any
        class except a small allow-list. Blocks the classic
        arbitrary-code-execution attack on pickle.load()."""
        import pickle
        allowed = {
            ("builtins", "dict"),
            ("builtins", "list"),
            ("builtins", "tuple"),
            ("builtins", "set"),
            ("builtins", "frozenset"),
            ("builtins", "str"),
            ("builtins", "int"),
            ("builtins", "float"),
            ("builtins", "bool"),
            ("builtins", "bytes"),
            ("builtins", "NoneType"),
            ("collections", "defaultdict"),
            ("collections", "OrderedDict"),
        }
        class _Restricted(pickle.Unpickler):
            def find_class(self, module, name):
                if name == "Node":
                    return Node
                if (module, name) in allowed:
                    return super().find_class(module, name)
                raise pickle.UnpicklingError(
                    f"Refusing to load unsafe pickle class: "
                    f"{module}.{name}")
        return _Restricted(fp)

    def _save_snapshot(self):
        if not self.scan_result:
            self._warn("First, run a scan",
                                   "Run a scan before saving a snapshot.")
            return
        path = filedialog.asksaveasfilename(
            title="Save scan snapshot",
            defaultextension=".capsnap",
            filetypes=[("Capacitra snapshot", "*.capsnap"),
                       ("All files", "*.*")],
            initialfile="capacitra.capsnap")
        if not path:
            return
        try:
            import pickle
            data = {
                "version": APP_VERSION,
                "saved_at": time.time(),
                "scan_root": self.scan_root,
                "result": self.scan_result,
            }
            with open(path, "wb") as f:
                f.write(self._SNAPSHOT_MAGIC)
                pickle.dump(data, f, pickle.HIGHEST_PROTOCOL)
            self.status_var.set(f"Snapshot saved: {path}")
            self.status_icon.configure(fg=self.theme["success"])
        except Exception as e:
            self._error("Save failed", str(e))

    def _load_snapshot(self):
        path = filedialog.askopenfilename(
            title="Open scan snapshot",
            defaultextension=".capsnap",
            filetypes=[("Capacitra snapshot", "*.capsnap"),
                       ("All files", "*.*")])
        if not path:
            return
        try:
            with open(path, "rb") as f:
                head = f.read(len(self._SNAPSHOT_MAGIC))
                if head != self._SNAPSHOT_MAGIC:
                    self._error(
                        "Not a Capacitra snapshot",
                        "This file does not look like a Capacitra snapshot. "
                        "For your safety the file was not loaded.")
                    return
                data = self._safe_unpickler(f).load()
            self.scan_root = data["scan_root"]
            self._overview_payload = None
            # Reuse the normal post-scan flow with _restore=True so that
            # the synthetic free / inaccessible nodes are not re-appended.
            self._on_scan_done(data["result"], _restore=True)
            when = time.strftime("%Y-%m-%d %H:%M",
                                 time.localtime(data.get("saved_at", 0)))
            self.status_var.set(
                f"Loaded snapshot ({when}) — {self.scan_root}")
            self.status_icon.configure(fg=self.theme["accent"])
        except Exception as e:
            self._error("Load failed", str(e))

    def _compare_snapshot(self):
        if not self.scan_result:
            self._warn("First, run a scan",
                                   "Run a current scan before comparing.")
            return
        path = filedialog.askopenfilename(
            title="Compare with snapshot",
            defaultextension=".capsnap",
            filetypes=[("Capacitra snapshot", "*.capsnap"),
                       ("All files", "*.*")])
        if not path:
            return
        try:
            with open(path, "rb") as f:
                head = f.read(len(self._SNAPSHOT_MAGIC))
                if head != self._SNAPSHOT_MAGIC:
                    self._error(
                        "Not a Capacitra snapshot",
                        "This file does not look like a Capacitra snapshot. "
                        "For your safety the file was not loaded.")
                    return
                data = self._safe_unpickler(f).load()
            old_root = data["result"]["root"]
        except Exception as e:
            self._error("Load failed", str(e))
            return
        # Build flat path → size maps.
        def walk(node, out):
            out[node.path] = node.size
            for c in node.children:
                walk(c, out)
        old_map, new_map = {}, {}
        walk(old_root, old_map)
        walk(self.scan_result["root"], new_map)
        all_paths = set(old_map) | set(new_map)
        diffs = []
        for p in all_paths:
            o = old_map.get(p, 0)
            n = new_map.get(p, 0)
            d = n - o
            if d != 0:
                diffs.append((d, p, o, n))
        diffs.sort(key=lambda x: abs(x[0]), reverse=True)
        self._show_snapshot_diff(diffs, data)


    def _install_schedule(self):
        """Register a Windows Task Scheduler job that runs
        Capacitra.exe headlessly once a day and writes a CSV report."""
        if sys.platform != "win32":
            self._info(
                "Windows only",
                "Scheduled scans use Windows Task Scheduler and are "
                "available on Windows.")
            return
        if not self.scan_root:
            self._warn(
                "Pick a folder first",
                "Choose the drive or folder you want the scheduled "
                "scan to cover, then open this dialog again.")
            return

        # Default output next to the exe: capacitra_daily.csv
        exe_path = sys.executable
        if not exe_path or not exe_path.lower().endswith(".exe"):
            self._info(
                "Only works from the built .exe",
                "Scheduled scans use the packaged Capacitra.exe. "
                "Run it from the download build, not from source.")
            return

        out_dir = os.path.dirname(exe_path)
        out_file = os.path.join(out_dir, "capacitra_daily.csv")

        cmd = (
            f'"{exe_path}" --scan "{self.scan_root}" '
            f'--export "{out_file}" --quiet'
        )

        try:
            # /SC DAILY /ST 03:00 → run once a day at 3 AM
            proc = subprocess.run(
                ["schtasks", "/Create", "/TN", "CapacitraDailyScan",
                 "/TR", cmd, "/SC", "DAILY", "/ST", "03:00", "/F"],
                capture_output=True, text=True, timeout=15,
            )
            if proc.returncode == 0:
                self._info(
                    "Scheduled",
                    "A daily scan of\n\n"
                    f"  {self.scan_root}\n\n"
                    "has been scheduled for 03:00. The CSV report "
                    "will be written to:\n\n"
                    f"  {out_file}\n\n"
                    "Edit or remove it any time from Windows Task "
                    "Scheduler (search 'Task Scheduler' in Start).")
            else:
                self._warn(
                    "Could not install scheduled task",
                    (proc.stderr or proc.stdout or
                     "schtasks returned a non-zero exit code.").strip())
        except Exception as e:
            self._error("Error", str(e))

    def _show_snapshot_diff(self, diffs, snap_data):
        t = self.theme
        win = tk.Toplevel(self.root)
        win.title("Compare with snapshot")
        win.geometry("980x600")
        win.minsize(720, 480)
        win.configure(bg=t["bg"])
        # Header
        head = tk.Frame(win, bg=t["panel"], highlightthickness=1,
                        highlightbackground=t["border"])
        head.pack(fill="x", padx=14, pady=(14, 8))
        hi = tk.Frame(head, bg=t["panel"])
        hi.pack(fill="x", padx=18, pady=14)
        when = time.strftime("%Y-%m-%d %H:%M",
                             time.localtime(snap_data.get("saved_at", 0)))
        old_total = sum(c.size for c in snap_data["result"]["root"].children)
        new_total = sum(c.size for c in self.scan_result["root"].children)
        delta = new_total - old_total
        sign = "+" if delta >= 0 else "−"
        tk.Label(hi, text="Compare with snapshot",
                 bg=t["panel"], fg=t["fg"],
                 font=("Segoe UI Semibold", 14)
                 ).pack(anchor="w")
        tk.Label(hi,
                 text=f"Snapshot saved on {when}  ·  "
                      f"net change {sign}{human_size(abs(delta))}",
                 bg=t["panel"], fg=t["muted"],
                 font=UI_FONT).pack(anchor="w", pady=(2, 0))
        # Body table
        body = tk.Frame(win, bg=t["panel"], highlightthickness=1,
                        highlightbackground=t["border"])
        body.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        cols = ("delta", "old", "new", "path")
        tv = ttk.Treeview(body, columns=cols, show="headings")
        tv.heading("delta", text="Δ change")
        tv.heading("old", text="Old size")
        tv.heading("new", text="New size")
        tv.heading("path", text="Path")
        tv.column("delta", width=140, anchor="e")
        tv.column("old", width=110, anchor="e")
        tv.column("new", width=110, anchor="e")
        tv.column("path", width=580, anchor="w")
        tv.tag_configure("grew",
                         foreground=t["danger"])
        tv.tag_configure("shrank",
                         foreground=t["success"])
        vsb = ttk.Scrollbar(body, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=vsb.set)
        tv.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        for delta_val, p, o, n in diffs[:500]:
            sign = "+" if delta_val > 0 else "−"
            tag = "grew" if delta_val > 0 else "shrank"
            tv.insert("", "end",
                      values=(f"{sign}{human_size(abs(delta_val))}",
                              human_size(o), human_size(n), p),
                      tags=(tag,))

    # ----- Context menu -----
    def _make_context_menu(self, widget, kind):
        m = tk.Menu(widget, tearoff=0)
        m.add_command(label="Open",
                      command=lambda: self._ctx("open", kind))
        m.add_command(label="Show in Explorer",
                      command=lambda: self._ctx("explorer", kind))
        m.add_command(label="Copy path",
                      command=lambda: self._ctx("copy", kind))
        if kind == "tree":
            m.add_separator()
            m.add_command(label="Rescan this folder",
                          command=lambda: self._ctx("rescan", kind))
        m.add_separator()
        m.add_command(label="Move to Recycle Bin",
                      command=lambda: self._ctx("recycle", kind))
        m.add_command(label="Move selected to Recycle Bin  (Del)",
                      command=self._batch_recycle)
        m.add_separator()
        m.add_command(label="Properties…",
                      command=lambda: self._ctx("properties", kind))

        def popup(event):
            iid = widget.identify_row(event.y)
            if iid:
                # Only reset selection if user right-clicked something
                # they haven't already multi-selected.
                if iid not in widget.selection():
                    widget.selection_set(iid)
                widget.focus(iid)
                m.tk_popup(event.x_root, event.y_root)
        widget.bind("<Button-3>", popup)

    def _ctx(self, action, kind):
        path = self._selected_path(kind)
        if not path or path in ("(free)", "(gap)"):
            return
        if action == "open":
            open_path(path)
        elif action == "explorer":
            open_in_explorer(path)
        elif action == "copy":
            self.root.clipboard_clear()
            self.root.clipboard_append(path)
            self.status_var.set(f"Copied: {path}")
        elif action == "rescan":
            # Rescan this folder as a fresh root scan.
            target = path if os.path.isdir(path) else os.path.dirname(path)
            if target:
                self.path_var.set(target)
                self.scan_root = target
                self._start_scan()
        elif action == "properties":
            self._show_properties_dialog(path)
        elif action == "recycle":
            if not self._ask(
                    "Confirm", f"Move this item to the Recycle Bin?\n\n{path}"):
                return
            # Run the (potentially slow) shell op on a worker thread so
            # the UI doesn't freeze on large files.
            self.status_var.set(f"Moving to Recycle Bin: {path} …")
            self.status_icon.configure(fg=self.theme["warn"])

            def _worker(p=path):
                ok = send_to_recycle_bin(p)
                self.root.after(0, lambda: self._recycle_done(p, ok))
            threading.Thread(target=_worker, daemon=True).start()

    def _recycle_done(self, path, ok):
        if ok:
            self.status_icon.configure(fg=self.theme["success"])
            self.status_var.set(f"Moved to Recycle Bin: {path}")
            # Drop row(s) referring to this path from any list view
            for tv in (getattr(self, "top_tree", None),
                       getattr(self, "dup_tree", None)):
                if tv is None:
                    continue
                for iid in list(tv.get_children()):
                    vals = tv.item(iid, "values")
                    text = tv.item(iid, "text")
                    if (vals and path in vals) or (text == path):
                        tv.delete(iid)
                        continue
                    # check children (e.g. duplicates groups)
                    for cid in list(tv.get_children(iid)):
                        if tv.item(cid, "text") == path:
                            tv.delete(cid)
        else:
            self.status_icon.configure(fg=self.theme["danger"])
            self.status_var.set(f"Could not move to Recycle Bin: {path}")
            hint = ""
            if sys.platform != "win32":
                hint = ("\n\nOn macOS / Linux, install the 'send2trash' "
                        "package for safe trashing:\n"
                        "    pip install send2trash")
            self._error("Error",
                                 "Could not move to Recycle Bin." + hint)
        # Refresh all Treeviews
        try:
            self._refresh_after_recycle(ok if isinstance(ok, list) else [path])
        except Exception:
            pass

    def _selected_path(self, kind):
        if kind == "tree":
            sel = self.tree.focus()
            n = self._tree_node_map.get(sel)
            return n.path if n else None
        if kind == "top":
            sel = self.top_tree.focus()
            if not sel:
                return None
            v = self.top_tree.item(sel, "values")
            # New schema: (idx, name, size, age, path) — fall back to last
            if len(v) >= 5:
                return v[4]
            return v[-1] if v else None
        if kind == "dup":
            sel = self.dup_tree.focus()
            if not sel:
                return None
            tags = self.dup_tree.item(sel, "tags")
            if "file" in tags:
                return self.dup_tree.item(sel, "text")
        return None

    def _open_top_selected(self):
        p = self._selected_path("top")
        if p:
            open_in_explorer(p)

    # ----- Drives & scanning -----
    def _populate_drives(self):
        drives = []
        if sys.platform == "win32":
            import string
            for letter in string.ascii_uppercase:
                d = f"{letter}:\\"
                if os.path.exists(d):
                    drives.append(f"{d} (Local Disk)")
        elif sys.platform == "darwin":
            # macOS: list /Volumes/* (each mounted volume)
            drives.append("/ (Macintosh HD)")
            try:
                for name in sorted(os.listdir("/Volumes")):
                    path = "/Volumes/" + name
                    if os.path.isdir(path) and not name.startswith("."):
                        drives.append(f"{path} ({name})")
            except (PermissionError, FileNotFoundError):
                pass
            home = os.path.expanduser("~")
            if home not in [d.split()[0] for d in drives]:
                drives.append(f"{home} (Home)")
        else:
            # Linux / other Unix
            drives.append("/ (Root)")
            home = os.path.expanduser("~")
            drives.append(f"{home} (Home)")
            for base in ("/mnt", "/media", "/run/media"):
                if not os.path.isdir(base):
                    continue
                try:
                    for name in sorted(os.listdir(base)):
                        path = os.path.join(base, name)
                        if os.path.isdir(path):
                            # /run/media has a user subdir level
                            if base == "/run/media":
                                try:
                                    for sub in sorted(os.listdir(path)):
                                        full = os.path.join(path, sub)
                                        if os.path.isdir(full):
                                            drives.append(f"{full} ({sub})")
                                except (PermissionError, FileNotFoundError):
                                    pass
                            else:
                                drives.append(f"{path} ({name})")
                except (PermissionError, FileNotFoundError):
                    pass
        self.path_combo["values"] = drives
        if drives:
            self.path_var.set(drives[0])

    def _selected_drive_path(self):
        v = self.path_var.get().strip()
        # Strip trailing label
        if " (" in v:
            v = v.split(" (")[0]
        return v

    def _on_primary_action(self):
        """Header primary button: starts a scan, or cancels the running one."""
        if self.scan_thread and self.scan_thread.is_alive():
            self._cancel_scan()
        else:
            self._start_scan()

    def _set_primary_button(self, scanning):
        """Morph New Scan ↔ Cancel."""
        if not hasattr(self, "new_scan_btn"):
            return
        if scanning:
            self.new_scan_btn.icon_lbl.configure(text="■")
            self.new_scan_btn.text_lbl.configure(text="Cancel")
        else:
            self.new_scan_btn.icon_lbl.configure(text="🔍")
            self.new_scan_btn.text_lbl.configure(text="New Scan")

    def _start_scan(self):
        path = self._selected_drive_path()
        if not path:
            self._warn("Pick a folder",
                                   "Choose a drive from the dropdown.")
            return
        if not os.path.isdir(path):
            self._error("Not found", f"Folder not found:\n{path}")
            return
        if self.scan_thread and self.scan_thread.is_alive():
            self._info("Busy", "A scan is already running.")
            return
        self._clear_results()
        self.scan_root = path
        self.stop_event = threading.Event()
        self.progress.start(10)
        self.status_icon.configure(fg=self.theme["warn"])
        self.status_var.set(f"Scanning: {path} …")
        self.scan_thread = ScanWorker(path, self.result_queue, self.stop_event,
                                      excludes=self._excludes)
        self.scan_thread.start()
        self._set_primary_button(scanning=True)

    def _rescan(self):
        if self.scan_root:
            for v in self.path_combo["values"]:
                if v.startswith(self.scan_root):
                    self.path_var.set(v)
                    break
        self._start_scan()

    def _cancel_scan(self):
        if self.scan_thread and self.scan_thread.is_alive():
            self.stop_event.set()
        if self.dup_thread and self.dup_thread.is_alive():
            self.dup_stop.set()

    def _browse_and_scan(self):
        """Folder picker → set the path → kick off a scan."""
        f = filedialog.askdirectory(title="Select folder to analyze")
        if not f:
            return
        # Match it against drive list display strings; otherwise show raw path
        display = f
        for v in self.path_combo["values"]:
            if v.startswith(f):
                display = v
                break
        else:
            # Append the chosen folder to the dropdown so it's selectable later
            values = list(self.path_combo["values"]) + [f]
            self.path_combo["values"] = values
        self.path_var.set(display)
        self._start_scan()

    def _show_export_menu(self):
        """Open a small popup with all export options."""
        m = tk.Menu(self.root, tearoff=0)
        m.add_command(label="Export as CSV…",   command=self._export_csv)
        m.add_command(label="Export as HTML…",  command=self._export_html)
        if HAS_REPORTLAB:
            m.add_command(label="Export as PDF…",   command=self._export_pdf)
        if HAS_OPENPYXL:
            m.add_command(label="Export as Excel…", command=self._export_excel)
        m.add_separator()
        m.add_command(label="Save snapshot…",
                      command=self._save_snapshot)
        m.add_command(label="Compare with snapshot…",
                      command=self._compare_snapshot)
        try:
            x = self.export_btn.winfo_rootx()
            y = self.export_btn.winfo_rooty() + self.export_btn.winfo_height()
            m.tk_popup(x, y)
        except Exception:
            m.tk_popup(self.root.winfo_pointerx(),
                       self.root.winfo_pointery())

    def _clear_results(self):
        if hasattr(self, "tree"):
            self.tree.delete(*self.tree.get_children())
        self._tree_node_map.clear()
        self._node_iid_map.clear()
        self._expanded_iids.clear()
        self._overview_payload = None
        if hasattr(self, "treemap"):
            self.treemap.set_data([])
        if hasattr(self, "pie"):
            self.pie.set_data([])
        if hasattr(self, "bar_chart"):
            self.bar_chart.set_data([])
        if hasattr(self, "ext_tree"):
            self.ext_tree.delete(*self.ext_tree.get_children())
        if hasattr(self, "top_tree"):
            self.top_tree.delete(*self.top_tree.get_children())
        if hasattr(self, "dup_tree"):
            self.dup_tree.delete(*self.dup_tree.get_children())
        if hasattr(self, "tree_stat_lbl"):
            self.tree_stat_lbl.configure(text="—")
        for lbl in getattr(self, "_stat_cards", []):
            lbl.configure(text="—")
        self.hero_size_lbl.configure(text="—")
        self.hero_sub_lbl.configure(text="")
        self.hero_canvas.delete("all")
        for w in self.hero_legend.winfo_children():
            w.destroy()
        self.top_folders_canvas.delete("all")

    def _poll_queue(self):
        # Skip work if the window is being torn down. Prevents a race
        # where root.after() fires after root.destroy(), raising
        # TclError from every widget touch.
        if not getattr(self, "_alive", True):
            return
        try:
            while True:
                kind, payload = self.result_queue.get_nowait()
                if kind == "progress":
                    self.status_var.set(
                        f"Scanning… {payload['files']:,} files  ·  "
                        f"{human_size(payload['bytes'])}  ·  "
                        f"{payload['current']}")
                elif kind == "done":
                    self._on_scan_done(payload)
                elif kind == "cancelled":
                    self.progress.stop()
                    self.status_icon.configure(fg=self.theme["warn"])
                    self.status_var.set("Scan cancelled.")
                    self._set_primary_button(scanning=False)
                elif kind == "error":
                    self.progress.stop()
                    self.status_icon.configure(fg=self.theme["danger"])
                    self.status_var.set(f"Error: {payload}")
                    self._set_primary_button(scanning=False)
                    self._error("Scan error", str(payload))
                elif kind == "dup_progress":
                    cur = payload['current']
                    tot = max(1, payload['total'])
                    pct = int(cur * 100 / tot)
                    self.dup_status_var.set(
                        f"Hashing candidate groups · {cur:,} of {tot:,} "
                        f"· {pct}% complete")
                    try:
                        self.dup_progress_bar["value"] = pct
                        self.dup_progress_pct.configure(text=f"{pct}%")
                        self.dup_progress_detail.configure(
                            text=f"Group {cur:,} / {tot:,} processed. "
                                 f"This can take a while on large drives.")
                    except (AttributeError, tk.TclError):
                        pass
                elif kind == "dup_done":
                    try:
                        self.dup_progress_frame.pack_forget()
                        self.dup_progress_bar["value"] = 0
                    except (AttributeError, tk.TclError):
                        pass
                    self._on_dup_done(payload)
        except queue.Empty:
            pass
        self.root.after(80, self._poll_queue)

    def _on_scan_done(self, result, _restore=False):
        self.progress.stop()
        self._set_primary_button(scanning=False)
        # Disk usage info
        if "disk" not in result:
            try:
                import shutil
                usage = shutil.disk_usage(self.scan_root)
                result["disk"] = {"total": usage.total, "used": usage.used,
                                  "free": usage.free}
            except Exception:
                result["disk"] = {"total": 0, "used": 0, "free": 0}
        # Augment root with synthetic Free / Inaccessible entries so the tree,
        # treemap and top-folder bars all reflect the full disk picture.
        if not _restore:
            disk = result["disk"]
            total_scanned = result["total_bytes"]
            gap = max(disk["used"] - total_scanned, 0)
            root = result["root"]
            # Drop any prior synthetic entries (in case of rescan on same dict)
            root.children = [c for c in root.children
                             if c.path not in ("(free)", "(gap)")]
            if disk["total"]:
                if disk["free"] > 0:
                    free_node = Node("[Free space]", "(free)", is_dir=False,
                                     parent=root)
                    free_node.size = disk["free"]
                    root.children.append(free_node)
                if gap > 0:
                    gap_node = Node("[Inaccessible]", "(gap)", is_dir=False,
                                    parent=root)
                    gap_node.size = gap
                    root.children.append(gap_node)
                # Make root.size reflect the disk total so % adds up to 100
                root.size = disk["total"]
            root.children.sort(key=lambda n: n.size, reverse=True)
        self.scan_result = result
        self._overview_payload = None
        try:
            self._downloads_hint = self._analyze_downloads_aging()
        except Exception:
            self._downloads_hint = None
        # Surface a subtle status hint about aged Downloads content
        try:
            if self._downloads_hint:
                self.status_var.set(
                    "Tip: Downloads folder has "
                    f"{human_size(self._downloads_hint['bytes'])} "
                    "in files >30 days old. Open Downloads folder "
                    "aging in the More menu for details.")
        except Exception:
            pass
        # Update export panel state (enable/disable status line)
        try:
            self._refresh_export_status()
        except Exception:
            pass
        total = result["total_bytes"]
        files = result["total_files"]
        dirs = result.get("total_dirs", 0)
        denied = result.get("denied", 0)
        disk = result["disk"]
        # Status bar — percentage is now relative to disk total
        self.status_icon.configure(fg=self.theme["success"])
        if disk["total"]:
            pct = total / disk["total"] * 100
            msg = (
                f"Ready  ·  {files:,} files  ·  {dirs:,} folders  ·  "
                f"Scanned {human_size(total)} ({pct:.1f}% of disk)  ·  "
                f"Used {human_size(disk['used'])}  ·  "
                f"Free {human_size(disk['free'])}  ·  "
                f"Total {human_size(disk['total'])}"
            )
            if denied:
                msg += f"  ·  {denied:,} skipped"
            self.status_var.set(msg)
        else:
            self.status_var.set(
                f"Ready  ·  {files:,} files  ·  {human_size(total)}")
        # Sidebar disk card
        # Sidebar card is narrow — use short label (drive letter only)
        # so text doesn't get clipped.
        _dr = self.scan_root.rstrip(chr(92)).rstrip("/")
        nice_label = _dr if _dr else "Local Disk"
        self._update_sb_disk(disk["used"], disk["total"], nice_label)
        # Folder tree card stats
        self.tree_stat_lbl.configure(
            text=f"{files:,} files  ·  {human_size(total)}")
        # Stat cards
        self._stat_cards[0].configure(text=f"{files:,}")
        self._stat_cards[1].configure(text=f"{dirs:,}")
        self._stat_cards[2].configure(text=f"{len(result['extensions'])}")
        self._stat_cards[3].configure(text=f"{denied:,}")
        # Hero
        self._draw_hero(total, disk["used"], disk["total"], disk["free"])
        # Record in history
        try:
            self._record_history(result)
        except Exception:
            pass
        # Stage the heavy fills so the UI stays responsive
        self._stage_fill(result, denied)

    def _stage_fill(self, result, denied):
        steps = [
            (40, lambda: self._fill_tree(result["root"])),
            (10, lambda: self._fill_extensions(result["extensions"],
                                               result["total_bytes"])),
            (10, lambda: self._fill_top(result["largest_files"])),
            (10, lambda: self._focus_node(result["root"])),
            (10, lambda: self._redraw_top_folders()),
        ]
        # No more denied-dialog — denial info is already in the status bar.

        def step(i=0):
            if i >= len(steps):
                return
            delay, fn = steps[i]
            try:
                fn()
            except Exception as e:
                print("stage step error:", e)
            self.root.after(delay, lambda: step(i + 1))
        self.root.after(20, step)

    def _show_denied_dialog(self, result, denied):
        files = result["total_files"]
        dirs = result.get("total_dirs", 0)
        total = result["total_bytes"]
        disk = result["disk"]
        gap = max(disk["used"] - total, 0)
        self._info(
            "Scan complete — some items skipped",
            f"Scanned {files:,} files in {dirs:,} folders.\n"
            f"Measured: {human_size(total)}\n"
            f"Disk used: {human_size(disk['used'])} of "
            f"{human_size(disk['total'])}\n"
            f"Unaccounted: {human_size(gap)}\n\n"
            f"{denied:,} items were inaccessible — typically:\n"
            "  • System Volume Information, $Recycle.Bin\n"
            "  • Other users' profile folders\n"
            "  • Locked files (pagefile.sys, hiberfil.sys)\n\n"
            "Run as Administrator to include them."
        )

    def _redraw_hero(self):
        """Repaint the hero card from the cached values (used on resize)."""
        if not hasattr(self, "_hero_cache"):
            return
        c = self._hero_cache
        self._draw_hero(c["scanned"], c["used"], c["total"], c["free"])

    def _draw_hero(self, scanned, disk_used, disk_total, disk_free):
        # Cache for later redraws
        self._hero_cache = {"scanned": scanned, "used": disk_used,
                            "total": disk_total, "free": disk_free}
        t = self.theme
        # Big number
        self.hero_size_lbl.configure(
            text=human_size(disk_used) if disk_total else human_size(scanned))
        self.hero_label.configure(
            text=f"Disk usage  ({self.scan_root})")
        if disk_total:
            self.hero_sub_lbl.configure(
                text=f"used of {human_size(disk_total)}  ·  "
                     f"free {human_size(disk_free)}")
        else:
            self.hero_sub_lbl.configure(text=f"scanned: {human_size(scanned)}")
        # Bar
        self.hero_canvas.update_idletasks()
        cw = self.hero_canvas.winfo_width() or 600
        ch = 20
        self.hero_canvas.delete("all")
        # Track
        self.hero_canvas.create_rectangle(0, 0, cw, ch,
                                          fill=t["panel_alt"], outline="")
        total_for = disk_total if disk_total else max(scanned, 1)
        gap = max(disk_used - scanned, 0) if disk_total else 0
        seg_scan = scanned / total_for * cw
        seg_gap  = gap / total_for * cw
        if seg_scan > 0:
            self.hero_canvas.create_rectangle(0, 0, seg_scan, ch,
                                              fill=t["accent"], outline="")
        if seg_gap > 1:
            self.hero_canvas.create_rectangle(seg_scan, 0,
                                              seg_scan + seg_gap, ch,
                                              fill=t["warn"], outline="")
        # Legend
        for w in self.hero_legend.winfo_children():
            w.destroy()
        def dot(parent, color, text):
            wrap = tk.Frame(parent, bg=t["panel"])
            wrap.pack(side="left", padx=(0, 20))
            cv = tk.Canvas(wrap, width=10, height=10, bg=t["panel"],
                           highlightthickness=0)
            cv.create_oval(0, 0, 10, 10, fill=color, outline="")
            cv.pack(side="left", padx=(0, 6), pady=4)
            tk.Label(wrap, text=text, bg=t["panel"], fg=t["muted"],
                     font=UI_FONT).pack(side="left")
        dot(self.hero_legend, t["accent"], f"Scanned {human_size(scanned)}")
        if disk_total:
            dot(self.hero_legend, t["warn"], f"Inaccessible {human_size(gap)}")
            dot(self.hero_legend, t["panel_alt"], f"Free {human_size(disk_free)}")

    # ----- Tree filling -----
    def _fill_tree(self, root_node):
        self.tree.delete(*self.tree.get_children())
        self._tree_node_map.clear()
        self._node_iid_map.clear()
        self._expanded_iids.clear()
        iid = self._insert_node("", root_node, parent_size=root_node.size or 1)
        if root_node.children:
            self.tree.insert(iid, "end", text="…")
        self.tree.item(iid, open=True)
        self._expand_iid(iid, root_node)
        # Auto-expand the top-3 largest immediate folders one extra
        # level so the user sees actual files/subfolders out of the
        # box, not just a list of top-level directory names.
        big_kids = [c for c in root_node.children
                    if c.is_dir and c.path not in ("(free)", "(gap)")]
        big_kids.sort(key=lambda n: n.size, reverse=True)
        for big in big_kids[:3]:
            big_iid = self._node_iid_map.get(big)
            if not big_iid:
                continue
            try:
                self.tree.item(big_iid, open=True)
                self._expand_iid(big_iid, big)
            except tk.TclError:
                pass

    def _insert_node(self, parent_iid, node, parent_size, index=0):
        synth = node.path in ("(free)", "(gap)")
        if synth:
            icon = "🟢 " if node.path == "(free)" else "🔒 "
        else:
            icon = "📁 " if node.is_dir else "📄 "
        pct = (node.size / parent_size * 100) if parent_size > 0 else 0
        zebra = "odd" if index % 2 else "even"

        # Files / folders columns: only meaningful on real folders.
        if synth:
            files_v = ""
            folders_v = ""
        elif node.is_dir:
            files_v = f"{node.file_count:,}"
            folders_v = f"{node.folder_count:,}"
        else:
            files_v = ""
            folders_v = ""

        # Allocated column (cluster-rounded disk footprint)
        if synth or not node.allocated:
            allocated_v = ""
        else:
            allocated_v = human_size(node.allocated)

        # Modified column
        if synth or not node.mtime:
            modified_v = ""
        else:
            modified_v = time.strftime("%Y-%m-%d %H:%M",
                                       time.localtime(node.mtime))

        # Accessed column
        if synth or not node.atime:
            accessed_v = ""
        else:
            accessed_v = time.strftime("%Y-%m-%d %H:%M",
                                       time.localtime(node.atime))

        # Owner column starts as cache hit or "…" placeholder; the
        # actual lookup is dispatched lazily on a background thread.
        if synth:
            owner_v = ""
        else:
            owner_v = self._owner_cache.get(node.path, "…")

        iid = self.tree.insert(
            parent_iid, "end",
            text=icon + node.name,
            values=(human_size(node.size),
                    allocated_v,
                    files_v,
                    folders_v,
                    bar_string(pct, 12),
                    f"{pct:.1f}%",
                    modified_v,
                    accessed_v,
                    owner_v),
            tags=(zebra, "synth" if synth else "real"),
        )
        self._tree_node_map[iid] = node
        self._node_iid_map[node] = iid
        if not synth and owner_v == "…":
            self._queue_owner_lookup(iid, node.path)
        return iid

    def _queue_owner_lookup(self, iid, path):
        """Resolve the file owner on a worker thread so the UI never
        blocks on per-file Windows API calls. The result is posted
        back to the main thread via root.after()."""
        if not hasattr(self, "_owner_thread_pool"):
            # Single background thread + queue: keeps API calls
            # serialised, which is friendlier to slow remote shares.
            import queue as _q
            self._owner_queue = _q.Queue()
            self._owner_thread_pool = True
            def _worker():
                while True:
                    item = self._owner_queue.get()
                    if item is None:
                        return
                    w_iid, w_path = item
                    owner = self._owner_cache.get(w_path)
                    if owner is None:
                        try:
                            owner = get_file_owner(w_path) or ""
                        except Exception:
                            owner = ""
                        # LRU-style cap: session-long walks of huge
                        # trees would grow the cache without bound.
                        if len(self._owner_cache) > 10000:
                            try:
                                # Drop the oldest ~1000 entries
                                keys = list(self._owner_cache.keys())[:1000]
                                for k in keys:
                                    self._owner_cache.pop(k, None)
                            except Exception:
                                pass
                        self._owner_cache[w_path] = owner
                    # Schedule UI update on main thread
                    try:
                        self.root.after(0, self._apply_owner,
                                        w_iid, owner)
                    except Exception:
                        pass
            t = threading.Thread(target=_worker, daemon=True,
                                 name="OwnerLookup")
            t.start()
        try:
            self._owner_queue.put_nowait((iid, path))
        except Exception:
            pass

    def _apply_owner(self, iid, owner):
        """Update the Owner column for `iid` once the lookup finishes."""
        try:
            current_values = list(self.tree.item(iid, "values"))
            if not current_values:
                return
            # Owner is the LAST column (index 8) in the new layout
            if len(current_values) >= 9:
                current_values[8] = owner or "—"
                self.tree.item(iid, values=current_values)
        except tk.TclError:
            # Row no longer exists (panel rebuilt). Ignore.
            pass

    def _on_tree_open(self, _):
        iid = self.tree.focus()
        if iid in self._expanded_iids:
            return
        node = self._tree_node_map.get(iid)
        if not node:
            return
        self._expand_iid(iid, node)

    def _expand_iid(self, iid, node):
        for c in self.tree.get_children(iid):
            self.tree.delete(c)
        parent_size = max(node.size, 1)
        idx = 0
        for child in node.children:
            # Synthetic [Free space] / [Inaccessible] always visible
            is_synth = child.path in ("(free)", "(gap)")
            if not is_synth and not self._node_matches_adv(child):
                # If a folder fails the filter itself but any descendant
                # might pass, keep it for navigation. We can't know cheaply,
                # so just show dirs and filter files.
                if not child.is_dir:
                    continue
            cid = self._insert_node(iid, child, parent_size, index=idx)
            idx += 1
            if child.is_dir and child.children:
                self.tree.insert(cid, "end", text="…")
        self._expanded_iids.add(iid)

    def _apply_filter(self):
        # Parse the filter input into our advanced-filter dict.
        self._parse_filter(self.filter_var.get())
        # Collapse + re-expand the tree so the new filter takes effect.
        for iid in list(self._expanded_iids):
            self._expanded_iids.discard(iid)
            try:
                self.tree.item(iid, open=False)
            except tk.TclError:
                pass
        roots = self.tree.get_children("")
        if roots:
            self.tree.item(roots[0], open=True)
            node = self._tree_node_map.get(roots[0])
            if node:
                self._expand_iid(roots[0], node)

    def _parse_filter(self, txt):
        """Parse filter input:
           plain text → name contains
           *.ext      → ext match
           size:>10mb / size:<1gb
           age:>30d / age:<7d / age:>1y
        """
        adv = {"name": None, "min": 0, "max": float("inf"),
               "age_max_days": None, "age_min_days": None,
               "ext": None}
        if not txt:
            self._adv = adv
            return
        UNITS = {"b": 1, "kb": 1024, "mb": 1024**2, "gb": 1024**3,
                 "tb": 1024**4}
        import re as _re
        parts = [t for t in txt.split() if t]
        name_parts = []
        for part in parts:
            low = part.lower()
            if low.startswith("size:"):
                m = _re.match(r"size:([<>]=?)(\d+(?:\.\d+)?)(b|kb|mb|gb|tb)?",
                              low)
                if m:
                    op, val, unit = m.groups()
                    val = float(val) * UNITS.get(unit or "b", 1)
                    if ">" in op:
                        adv["min"] = max(adv["min"], val)
                    else:
                        adv["max"] = min(adv["max"], val)
                continue
            if low.startswith("age:"):
                m = _re.match(r"age:([<>]=?)(\d+)(d|w|mo|y)?", low)
                if m:
                    op, val, unit = m.groups()
                    val = int(val)
                    days = {"d": val, "w": val * 7,
                            "mo": val * 30, "y": val * 365,
                            None: val}.get(unit, val)
                    if ">" in op:
                        adv["age_min_days"] = days
                    else:
                        adv["age_max_days"] = days
                continue
            if low.startswith("*.") or low.startswith("."):
                ext = "." + low.lstrip("*").lstrip(".")
                adv["ext"] = ext
                continue
            name_parts.append(low)
        if name_parts:
            adv["name"] = " ".join(name_parts)
        self._adv = adv

    def _node_matches_adv(self, node):
        """Return True if node passes the current advanced-filter rules."""
        adv = self._adv
        if adv["name"] and adv["name"] not in node.name.lower():
            return False
        if not node.is_dir:  # size/age/ext filters only apply to files
            if node.size < adv["min"] or node.size > adv["max"]:
                return False
            if adv["ext"] and not node.name.lower().endswith(adv["ext"]):
                return False
            if adv["age_max_days"] is not None and node.mtime:
                age = (time.time() - node.mtime) / 86400
                if age > adv["age_max_days"]:
                    return False
            if adv["age_min_days"] is not None and node.mtime:
                age = (time.time() - node.mtime) / 86400
                if age < adv["age_min_days"]:
                    return False
        return True

    def _on_tree_select(self, _):
        if self._suppress_select:
            return
        iid = self.tree.focus()
        node = self._tree_node_map.get(iid)
        if node and node.is_dir and node is not self.current_node:
            self._focus_node(node, from_tree=True)

    def _focus_node(self, node, from_tree=False):
        self.current_node = node
        self._refresh_breadcrumb()
        items = [c for c in node.children if c.size > 0]
        if hasattr(self, "treemap"):
            self.treemap.set_data(items[:60])
        if hasattr(self, "bar_chart"):
            self.bar_chart.set_data([(c.name, c.size) for c in items[:14]])
        if hasattr(self, "pie"):
            pie_items = [(c.name, c.size) for c in items[:8]]
            if len(items) > 8:
                other = sum(c.size for c in items[8:])
                if other > 0:
                    pie_items.append(("Other", other))
            self.pie.set_data(pie_items)
        # Push selection back into tree only when not from tree click
        if not from_tree:
            iid = self._node_iid_map.get(node)
            if iid:
                self._suppress_select = True
                try:
                    self._ensure_visible(iid)
                    self.tree.see(iid)
                    self.tree.selection_set(iid)
                finally:
                    self._suppress_select = False

    def _ensure_visible(self, iid):
        parent = self.tree.parent(iid)
        chain = []
        while parent:
            chain.append(parent)
            parent = self.tree.parent(parent)
        for p in reversed(chain):
            if p not in self._expanded_iids:
                node = self._tree_node_map.get(p)
                if node:
                    self._expand_iid(p, node)
            self.tree.item(p, open=True)

    def _on_treemap_click(self, node):
        if node.path in ("(free)", "(gap)"):
            return
        if node.is_dir and node.children:
            self._focus_node(node)

    def _tree_up(self):
        if self.current_node and self.current_node.parent:
            self._focus_node(self.current_node.parent)

    def _tree_home(self):
        if self.scan_result:
            self._focus_node(self.scan_result["root"])

    def _sort_tree(self, col):
        if not self.scan_result:
            return
        root = self.scan_result["root"]
        cur_col, desc = self._sort_state.get(self.tree, (None, True))
        desc = not desc if cur_col == col else True
        self._sort_state[self.tree] = (col, desc)
        if col in ("size", "percent"):
            root.children.sort(key=lambda n: n.size, reverse=desc)
        elif col == "allocated":
            root.children.sort(key=lambda n: n.allocated or 0, reverse=desc)
        elif col == "#0":
            root.children.sort(key=lambda n: n.name.lower(), reverse=desc)
        elif col == "files":
            root.children.sort(key=lambda n: n.file_count, reverse=desc)
        elif col == "folders":
            root.children.sort(key=lambda n: n.folder_count, reverse=desc)
        elif col == "modified":
            root.children.sort(key=lambda n: n.mtime or 0, reverse=desc)
        elif col == "accessed":
            root.children.sort(key=lambda n: n.atime or 0, reverse=desc)
        else:
            return
        self._fill_tree(root)

    EXT_CATEGORIES = [
        ("Images",       "🖼", "#10B981",
         {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".svg",
          ".tiff", ".tif", ".ico", ".heic", ".raw", ".cr2", ".nef"}),
        ("Videos",       "🎬", "#EF4444",
         {".mp4", ".mov", ".avi", ".mkv", ".wmv", ".flv", ".webm",
          ".m4v", ".mpg", ".mpeg", ".3gp", ".ts"}),
        ("Audio",        "🎵", "#8B5CF6",
         {".mp3", ".wav", ".flac", ".aac", ".ogg", ".m4a", ".wma",
          ".opus", ".aiff", ".ape"}),
        ("Documents",    "📄", "#3B82F6",
         {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
          ".txt", ".rtf", ".odt", ".ods", ".csv", ".md", ".epub"}),
        ("Code",         "⚙",  "#F59E0B",
         {".py", ".js", ".ts", ".jsx", ".tsx", ".html", ".css",
          ".java", ".c", ".cpp", ".h", ".hpp", ".cs", ".go",
          ".rs", ".rb", ".php", ".swift", ".kt", ".sh", ".sql",
          ".xml", ".json", ".yml", ".yaml", ".toml"}),
        ("Archives",     "📦", "#A855F7",
         {".zip", ".rar", ".7z", ".tar", ".gz", ".bz2", ".xz",
          ".iso", ".dmg"}),
        ("Executables",  "⚡", "#0EA5E9",
         {".exe", ".msi", ".dll", ".so", ".app", ".apk", ".deb",
          ".rpm", ".bat", ".cmd", ".com", ".sys"}),
        ("Other",        "❔", "#94A3B8", set()),
    ]

    @classmethod
    def _categorize(cls, ext):
        e = ext.lower()
        for name, icon, color, exts in cls.EXT_CATEGORIES:
            if e in exts:
                return name, icon, color
        return "Other", "❔", "#94A3B8"

    def _fill_extensions(self, exts, total):
        # Aggregate per category
        cat_totals = {n: [0, 0] for n, _, _, _ in self.EXT_CATEGORIES}
        for ext, sz, cnt in exts:
            cat, _, _ = self._categorize(ext)
            cat_totals[cat][0] += sz
            cat_totals[cat][1] += cnt
        # Rebuild category cards
        if hasattr(self, "cat_row"):
            for w in self.cat_row.winfo_children():
                w.destroy()
            t = self.theme
            for name, icon, color, _ in self.EXT_CATEGORIES:
                sz, cnt = cat_totals[name]
                pct = (sz / total * 100) if total > 0 else 0
                card = tk.Frame(self.cat_row, bg=t["panel"],
                                highlightthickness=1,
                                highlightbackground=t["border"])
                card.pack(side="left", fill="both", expand=True,
                          padx=4 if name != "Images" else (0, 4))
                inner = tk.Frame(card, bg=t["panel"])
                inner.pack(fill="x", padx=12, pady=10)
                badge = tk.Label(inner, text=icon, bg=color, fg="white",
                                 font=("Segoe UI Emoji", 12),
                                 padx=8, pady=2)
                badge.pack(anchor="w")
                tk.Label(inner, text=name, bg=t["panel"], fg=t["muted"],
                         font=UI_FONT).pack(anchor="w", pady=(8, 0))
                tk.Label(inner, text=human_size(sz),
                         bg=t["panel"], fg=t["fg"],
                         font=("Segoe UI Semibold", 13)
                         ).pack(anchor="w")
                tk.Label(inner,
                         text=f"{cnt:,} files  ·  {pct:.1f}%",
                         bg=t["panel"], fg=t["muted"],
                         font=("Segoe UI", 8)).pack(anchor="w", pady=(2, 0))
        # Fill table with category column
        self.ext_tree.delete(*self.ext_tree.get_children())
        # Configure tags for color dots per category
        for name, _, color, _ in self.EXT_CATEGORIES:
            self.ext_tree.tag_configure(f"cat_{name}", foreground=color)
        for ext, sz, cnt in exts:
            cat, icon, color = self._categorize(ext)
            pct = (sz / total * 100) if total > 0 else 0
            self.ext_tree.insert("", "end",
                                 values=(f"●  {ext}", cat, human_size(sz),
                                         f"{pct:.2f}%", f"{cnt:,}"),
                                 tags=(f"cat_{cat}",))

    def _fill_top(self, files):
        self.top_tree.delete(*self.top_tree.get_children())
        now = time.time()
        for i, (sz, path, mt) in enumerate(files, 1):
            name = os.path.basename(path) or path
            emoji = self._file_emoji(path)
            age = format_age(now - mt) + " ago"
            tag = "zebra" if i % 2 else "largerow"
            self.top_tree.insert("", "end",
                                 values=(str(i),
                                         f"{emoji}  {name}",
                                         human_size(sz),
                                         f"🕒  {age}",
                                         path),
                                 tags=(tag,))
        # Update KPI cards
        if hasattr(self, "_large_kpi") and self.scan_result:
            sr = self.scan_result
            total_b = sr.get("total_bytes", 0)
            n_files = sr.get("total_files", 0)
            n_dirs = sr.get("total_dirs", 0)
            avg = total_b / n_files if n_files else 0
            self._large_kpi["Total Size"].configure(text=human_size(total_b))
            self._large_kpi["Files"].configure(text=f"{n_files:,}")
            self._large_kpi["Folders"].configure(text=f"{n_dirs:,}")
            self._large_kpi["Average File Size"].configure(
                text=human_size(avg))

    # ----- Duplicates -----
    def _find_duplicates(self):
        # Always switch to the Duplicates panel so the user sees the
        # status, even when something is wrong.
        self._select_panel("duplicates")
        if self.scan_thread and self.scan_thread.is_alive():
            self.dup_status_var.set(
                "Scan still running — wait until it finishes, then try again.")
            return
        if not self.scan_result:
            self.dup_status_var.set(
                "Run a scan first (New Scan in the header).")
            return
        if self.dup_thread and self.dup_thread.is_alive():
            self.dup_status_var.set("A duplicate search is already running…")
            return
        size_groups = self.scan_result.get("size_groups", {})
        candidates = {sz: ps for sz, ps in size_groups.items() if len(ps) > 1}
        thr_mb = ScanWorker.DUP_THRESHOLD // (1024 * 1024)
        if not candidates:
            self.dup_status_var.set(
                f"No duplicate candidates — there are no two files of the "
                f"same size ≥ {thr_mb} MB in the scan. Try lowering the "
                f"threshold in Settings, or scan a different folder.")
            return
        self.dup_tree.delete(*self.dup_tree.get_children())
        self.dup_stop = threading.Event()
        total_files = sum(len(p) for p in candidates.values())
        self.dup_status_var.set(
            f"Checking {total_files:,} files across {len(candidates):,} "
            f"size groups, computing SHA-1 hashes…")
        self.status_icon.configure(fg=self.theme["warn"])
        # Show the live progress card with cancel button
        try:
            self.dup_progress_bar["value"] = 0
            self.dup_progress_pct.configure(text="0%")
            self.dup_progress_detail.configure(
                text=f"Preparing {len(candidates):,} size groups…")
            self.dup_progress_frame.pack(fill="x")
        except (AttributeError, tk.TclError):
            pass
        self.dup_thread = DuplicateWorker(candidates, self.result_queue,
                                          self.dup_stop)
        self.dup_thread.start()

    def _cancel_duplicates(self):
        """User clicked Cancel during a duplicate search."""
        if self.dup_thread and self.dup_thread.is_alive():
            self.dup_stop.set()
            self.dup_status_var.set(
                "Cancelling duplicate search…")
            try:
                self.dup_progress_detail.configure(
                    text="Waiting for current group to finish…")
            except Exception:
                pass

    def _on_dup_done(self, payload):
        # Accept both new dict payload {"groups", "skipped_cloud"} and
        # the legacy list-of-tuples form, for backward compatibility.
        if isinstance(payload, dict):
            groups = payload.get("groups", [])
            skipped_cloud = payload.get("skipped_cloud", 0)
        else:
            groups = payload
            skipped_cloud = 0
        # v4.3: expose groups so the "Duplicate clusters by folder"
        # menu item has data to aggregate
        try:
            self._dup_result = list(groups)
        except Exception:
            self._dup_result = None
        self.dup_tree.delete(*self.dup_tree.get_children())
        self.status_icon.configure(fg=self.theme["success"])
        cloud_note = ""
        if skipped_cloud:
            cloud_note = (f"  ·  {skipped_cloud:,} cloud placeholder(s) "
                          f"skipped (not downloaded)")
        if not groups:
            self.dup_status_var.set(
                "Search complete, no duplicates found." + cloud_note)
            return
        total_waste = 0
        for sz, h, paths in groups:
            waste = sz * (len(paths) - 1)
            total_waste += waste
            gid = self.dup_tree.insert(
                "", "end",
                text=f"⚠  {len(paths)} copies  ·  hash {h[:10]}…",
                values=(human_size(sz), len(paths), human_size(waste)),
                open=False, tags=("group",))
            for p in paths:
                self.dup_tree.insert(gid, "end", text=p,
                                     values=("", "", ""), tags=("file",))
        self.dup_status_var.set(
            f"{len(groups)} duplicate group(s)  ·  potential savings: "
            f"{human_size(total_waste)}" + cloud_note)

    # ----- Export -----
    def _export_csv(self):
        if not self.scan_result:
            self._warn("Run a scan first",
                                "Pick a drive or folder and click New Scan, then come back.")
            return
        p = filedialog.asksaveasfilename(
            title="Save CSV report", defaultextension=".csv",
            filetypes=[("CSV", "*.csv"), ("All files", "*.*")],
            initialfile="capacitra_report.csv")
        if not p:
            return
        try:
            with open(p, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["Path", "Type", "Size (bytes)", "Size",
                            "Last modified"])
                self._walk_for_csv(self.scan_result["root"], w)
            self.status_var.set(f"CSV saved: {p}")
            if self._ask("Done",
                                   f"CSV saved:\n{p}\n\nOpen now?"):
                open_path(p)
        except Exception as e:
            self._error("Error", str(e))

    def _walk_for_csv(self, node, writer):
        kind = "Folder" if node.is_dir else "File"
        mt = (time.strftime("%Y-%m-%d %H:%M", time.localtime(node.mtime))
              if node.mtime else "")
        writer.writerow([_sanitize_cell(node.path), kind, node.size, human_size(node.size), _sanitize_cell(mt)])
        for c in node.children:
            self._walk_for_csv(c, writer)

    def _export_pdf(self):
        if not self.scan_result:
            self._warn("Run a scan first",
                                "Pick a drive or folder and click New Scan, then come back.")
            return
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import (SimpleDocTemplate, Paragraph,
                                            Spacer, Table, TableStyle)
            from reportlab.lib import colors
        except ImportError:
            self._info(
                "PDF export needs reportlab",
                "PDF export requires the 'reportlab' package.\n\n"
                "Install it with:\n    pip install reportlab\n\n"
                "Meanwhile, HTML export gives a printable browser report.")
            return
        path = filedialog.asksaveasfilename(
            title="Save PDF report", defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf"), ("All files", "*.*")],
            initialfile="capacitra_report.pdf")
        if not path:
            return
        try:
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.lib.enums import TA_LEFT, TA_RIGHT
            from reportlab.platypus import (PageBreak, KeepTogether)
            from reportlab.platypus.flowables import Flowable

            r = self.scan_result
            root = r["root"]
            total = r["total_bytes"]
            files = r["total_files"]
            dirs = r.get("total_dirs", 0)
            denied = r.get("denied", 0)
            disk = r.get("disk", {"total": 0, "used": 0, "free": 0})

            # ----- Brand palette -----
            C_NAVY    = colors.HexColor("#1E3A8A")
            C_BLUE    = colors.HexColor("#2563EB")
            C_CYAN    = colors.HexColor("#0EA5E9")
            C_GREEN   = colors.HexColor("#16A34A")
            C_ORANGE  = colors.HexColor("#EA580C")
            C_MUTED   = colors.HexColor("#64748B")
            C_FG      = colors.HexColor("#0F172A")
            C_PANEL   = colors.HexColor("#FFFFFF")
            C_BG      = colors.HexColor("#F8FAFC")
            C_BORDER  = colors.HexColor("#E2E8F0")
            C_HEADBG  = colors.HexColor("#EFF4FB")

            # ----- Inline branded header flowable (hex logo + name) -----
            class BrandHeader(Flowable):
                def __init__(self, width, scan_root, when):
                    Flowable.__init__(self)
                    self.width = width
                    self.height = 56
                    self.scan_root = scan_root
                    self.when = when

                def draw(self):
                    c = self.canv
                    # Gradient-ish navy band
                    c.setFillColor(C_NAVY)
                    c.rect(0, 0, self.width, self.height, fill=1, stroke=0)
                    # Logo on the left: hexagon + folder + bars
                    import math
                    cx, cy, R = 28, 28, 18
                    pts = []
                    for i in range(6):
                        ang = math.radians(-90 + 60 * i)
                        pts.append((cx + R * math.cos(ang),
                                    cy + R * math.sin(ang)))
                    c.setFillColor(C_BLUE)
                    p = c.beginPath()
                    p.moveTo(*pts[0])
                    for pt in pts[1:]:
                        p.lineTo(*pt)
                    p.close()
                    c.drawPath(p, fill=1, stroke=0)
                    # Folder body
                    c.setFillColor(C_CYAN)
                    fx, fy, fw, fh = cx - 11, cy - 7, 22, 14
                    p = c.beginPath()
                    p.moveTo(fx, fy + fh)
                    p.lineTo(fx + fw, fy + fh)
                    p.lineTo(fx + fw, fy + 2)
                    p.lineTo(fx + fw * 0.55, fy + 2)
                    p.lineTo(fx + fw * 0.45, fy)
                    p.lineTo(fx, fy)
                    p.close()
                    c.drawPath(p, fill=1, stroke=0)
                    # 3 ascending bars
                    c.setFillColor(colors.white)
                    bx = fx + 3
                    by = fy + 1
                    bw = (fw - 6) / 5
                    for i, h_bar in enumerate([3, 5, 8]):
                        x = bx + (i * 1.6 + 0.5) * bw
                        c.rect(x, by, bw, h_bar, fill=1, stroke=0)
                    # Title text
                    c.setFillColor(colors.white)
                    c.setFont("Helvetica-Bold", 18)
                    c.drawString(60, 32, APP_NAME)
                    c.setFont("Helvetica", 9)
                    c.setFillColor(colors.HexColor("#A7C5FF"))
                    c.drawString(60, 18, APP_TAGLINE)
                    # Right-aligned scan info
                    c.setFont("Helvetica", 8)
                    c.setFillColor(colors.HexColor("#CBD5E1"))
                    c.drawRightString(self.width - 10, 32,
                                      self.scan_root[:90])
                    c.drawRightString(self.width - 10, 18, self.when)

            # ----- KPI band: 4 coloured boxes -----
            class KpiRow(Flowable):
                def __init__(self, width, items):
                    Flowable.__init__(self)
                    self.width = width
                    self.items = items
                    self.height = 50

                def draw(self):
                    c = self.canv
                    n = len(self.items)
                    gap = 6
                    bw = (self.width - gap * (n - 1)) / n
                    for i, (label, value, col) in enumerate(self.items):
                        x = i * (bw + gap)
                        c.setFillColor(col)
                        c.roundRect(x, 0, bw, self.height, 4,
                                    fill=1, stroke=0)
                        c.setFillColor(colors.white)
                        c.setFont("Helvetica-Bold", 14)
                        c.drawString(x + 12, 22, value)
                        c.setFont("Helvetica", 7)
                        c.setFillColor(colors.HexColor("#E2E8F0"))
                        c.drawString(x + 12, 11, label.upper())

            page_w = A4[0] - 72  # 36pt margins on each side

            doc = SimpleDocTemplate(path, pagesize=A4,
                                    leftMargin=36, rightMargin=36,
                                    topMargin=36, bottomMargin=46,
                                    title=f"{APP_NAME} Report",
                                    author=APP_NAME,
                                    subject="Storage capacity report",
                                    creator=APP_NAME)
            styles = getSampleStyleSheet()
            h2 = ParagraphStyle("h2", parent=styles["Normal"],
                                fontName="Helvetica-Bold",
                                fontSize=12, textColor=C_NAVY,
                                spaceBefore=14, spaceAfter=8)
            body = ParagraphStyle("body", parent=styles["Normal"],
                                  fontName="Helvetica",
                                  fontSize=9, textColor=C_FG,
                                  leading=13)
            muted = ParagraphStyle("muted", parent=body,
                                   textColor=C_MUTED, fontSize=8)

            story = []
            when = time.strftime("%B %d, %Y · %H:%M")
            story.append(BrandHeader(page_w, self.scan_root, when))
            story.append(Spacer(1, 14))

            # KPI band
            disk_pct = (total / disk["total"] * 100) if disk.get("total") else 0
            story.append(KpiRow(page_w, [
                ("Scanned size", human_size(total), C_BLUE),
                ("Files", f"{files:,}", C_CYAN),
                ("Folders", f"{dirs:,}", C_GREEN),
                ("Inaccessible", f"{denied:,}", C_ORANGE),
            ]))
            story.append(Spacer(1, 6))

            # Disk usage summary
            story.append(Paragraph("Disk usage", h2))
            disk_data = [
                ["Disk capacity", human_size(disk.get("total", 0))],
                ["Used (everything on disk)", human_size(disk.get("used", 0))],
                ["Scanned (accessible)",
                 f"{human_size(total)}  ·  {disk_pct:.1f}% of disk"],
                ["Free space", human_size(disk.get("free", 0))],
                ["Inaccessible items", f"{denied:,}"],
            ]
            tbl = Table(disk_data, colWidths=[page_w * 0.55, page_w * 0.45])
            tbl.setStyle(TableStyle([
                ("FONTNAME",   (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE",   (0, 0), (-1, -1), 9.5),
                ("TEXTCOLOR",  (0, 0), (0, -1), C_MUTED),
                ("TEXTCOLOR",  (1, 0), (1, -1), C_FG),
                ("FONTNAME",   (1, 0), (1, -1), "Helvetica-Bold"),
                ("BACKGROUND", (0, 0), (-1, -1), C_PANEL),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [C_PANEL, C_BG]),
                ("LINEBELOW",  (0, 0), (-1, -2), 0.25, C_BORDER),
                ("LEFTPADDING",   (0, 0), (-1, -1), 12),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 12),
                ("TOPPADDING",    (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("BOX",        (0, 0), (-1, -1), 0.5, C_BORDER),
            ]))
            story.append(tbl)

            # Top folders
            story.append(Paragraph("Top 20 folders by size", h2))
            top_data = [["#", "Folder", "Size", "Share"]]
            denom = root.size or total or 1
            real = [c for c in root.children
                    if c.path not in ("(free)", "(gap)")]
            for i, c in enumerate(real[:20], start=1):
                pct = c.size / denom * 100
                top_data.append([
                    str(i),
                    c.name[:60],
                    human_size(c.size),
                    f"{pct:.1f}%",
                ])
            t2 = Table(top_data,
                       colWidths=[24, page_w * 0.58, 80, 60],
                       repeatRows=1)
            t2.setStyle(TableStyle([
                ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME",   (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE",   (0, 0), (-1, -1), 8.5),
                ("BACKGROUND", (0, 0), (-1, 0), C_NAVY),
                ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_PANEL, C_BG]),
                ("ALIGN",      (0, 0), (0, -1), "CENTER"),
                ("ALIGN",      (2, 0), (3, -1), "RIGHT"),
                ("LEFTPADDING",   (0, 0), (-1, -1), 7),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 7),
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LINEBELOW",     (0, 0), (-1, -1), 0.25, C_BORDER),
                ("BOX",        (0, 0), (-1, -1), 0.5, C_BORDER),
            ]))
            story.append(t2)

            # Top file types
            exts = r.get("extensions", [])
            if exts:
                story.append(Paragraph("Top 15 file types", h2))
                ft_data = [["Extension", "Total size", "Files", "Share"]]
                tdenom = total or 1
                for ext, sz, cnt in exts[:15]:
                    ft_data.append([ext, human_size(sz), f"{cnt:,}",
                                    f"{sz / tdenom * 100:.1f}%"])
                t3 = Table(ft_data,
                           colWidths=[page_w * 0.42, 100, 70, 60],
                           repeatRows=1)
                t3.setStyle(TableStyle([
                    ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME",   (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE",   (0, 0), (-1, -1), 8.5),
                    ("BACKGROUND", (0, 0), (-1, 0), C_NAVY),
                    ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_PANEL, C_BG]),
                    ("ALIGN",      (1, 0), (3, -1), "RIGHT"),
                    ("LEFTPADDING",   (0, 0), (-1, -1), 7),
                    ("RIGHTPADDING",  (0, 0), (-1, -1), 7),
                    ("TOPPADDING",    (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("LINEBELOW",     (0, 0), (-1, -1), 0.25, C_BORDER),
                    ("BOX",        (0, 0), (-1, -1), 0.5, C_BORDER),
                ]))
                story.append(t3)

            # Largest individual files
            big = r.get("largest_files", [])
            if big:
                story.append(Paragraph("Top 25 largest files", h2))
                bf_data = [["Size", "Path"]]
                for sz, p, _mt in big[:25]:
                    short = p if len(p) <= 90 else "…" + p[-87:]
                    bf_data.append([human_size(sz), short])
                t4 = Table(bf_data, colWidths=[80, page_w - 80],
                           repeatRows=1)
                t4.setStyle(TableStyle([
                    ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME",   (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE",   (0, 0), (-1, -1), 8),
                    ("BACKGROUND", (0, 0), (-1, 0), C_NAVY),
                    ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_PANEL, C_BG]),
                    ("ALIGN",      (0, 0), (0, -1), "RIGHT"),
                    ("LEFTPADDING",   (0, 0), (-1, -1), 7),
                    ("RIGHTPADDING",  (0, 0), (-1, -1), 7),
                    ("TOPPADDING",    (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("BOX",        (0, 0), (-1, -1), 0.5, C_BORDER),
                ]))
                story.append(t4)

            # Age cohort summary
            ages = r.get("age_buckets", [])
            if ages:
                story.append(Paragraph("Files by age", h2))
                age_data = [["Age range", "Files", "Total size"]]
                for label, lo, hi, cnt, sz in ages:
                    age_data.append([label, f"{cnt:,}", human_size(sz)])
                t5 = Table(age_data,
                           colWidths=[page_w * 0.50, 100, 110],
                           repeatRows=1)
                t5.setStyle(TableStyle([
                    ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME",   (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE",   (0, 0), (-1, -1), 9),
                    ("BACKGROUND", (0, 0), (-1, 0), C_NAVY),
                    ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_PANEL, C_BG]),
                    ("ALIGN",      (1, 0), (2, -1), "RIGHT"),
                    ("LEFTPADDING",   (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
                    ("TOPPADDING",    (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("BOX",        (0, 0), (-1, -1), 0.5, C_BORDER),
                ]))
                story.append(t5)

            story.append(Spacer(1, 18))
            story.append(Paragraph(
                f"Generated by <b>{APP_NAME}</b> v{APP_VERSION} · "
                f"{APP_TAGLINE} · <font color='#2563EB'>capacitra.com</font>",
                muted))

            # Page footer with brand + page number
            def _footer(canvas_obj, doc_obj):
                canvas_obj.saveState()
                canvas_obj.setStrokeColor(C_BORDER)
                canvas_obj.setLineWidth(0.5)
                canvas_obj.line(36, 30, A4[0] - 36, 30)
                canvas_obj.setFont("Helvetica", 8)
                canvas_obj.setFillColor(C_MUTED)
                canvas_obj.drawString(36, 18,
                                      f"{APP_NAME} · capacitra.com")
                canvas_obj.drawRightString(
                    A4[0] - 36, 18, f"Page {doc_obj.page}")
                canvas_obj.restoreState()

            doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
            self.status_var.set(f"PDF saved: {path}")
            if self._ask("Done", f"PDF saved:\n{path}\n\nOpen now?"):
                open_path(path)
        except Exception as e:
            self._error("Error", str(e))

    def _export_excel(self):
        if not self.scan_result:
            self._warn("Run a scan first",
                       "Pick a drive or folder and click New Scan, then come back.")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                         Border, Side, NamedStyle)
            from openpyxl.utils import get_column_letter
        except ImportError:
            self._info(
                "Excel export needs openpyxl",
                "Excel export requires the 'openpyxl' package.\n\n"
                "Install it with:\n    pip install openpyxl\n\n"
                "Meanwhile CSV export works without any dependency.")
            return
        path = filedialog.asksaveasfilename(
            title="Save Excel report", defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
            initialfile="capacitra_report.xlsx")
        if not path:
            return
        try:
            r = self.scan_result
            root = r["root"]
            total = r["total_bytes"]
            files = r["total_files"]
            dirs = r.get("total_dirs", 0)
            denied = r.get("denied", 0)
            disk = r.get("disk", {"total": 0, "used": 0, "free": 0})
            disk_pct = (total / disk["total"] * 100) if disk.get("total") else 0

            wb = Workbook()
            wb.properties.creator = APP_NAME
            wb.properties.title = f"{APP_NAME} storage report"
            wb.properties.subject = "Disk capacity analysis"

            # Reusable styles
            NAVY  = "FF1E3A8A"; BLUE = "FF2563EB"; CYAN = "FF0EA5E9"
            GREEN = "FF16A34A"; ORG  = "FFEA580C"
            MUTED = "FF64748B"; FG = "FF0F172A"; BG = "FFF8FAFC"
            BORDER_CLR = "FFE2E8F0"
            HEAD_FONT  = Font(bold=True, color="FFFFFFFF", size=11,
                              name="Calibri")
            BODY_FONT  = Font(name="Calibri", size=10, color=FG)
            MUTED_FONT = Font(name="Calibri", size=10, color=MUTED)
            BIG_FONT   = Font(name="Calibri", size=16, color=FG, bold=True)
            HUGE_FONT  = Font(name="Calibri", size=22, color="FFFFFFFF",
                              bold=True)
            SUB_FONT   = Font(name="Calibri", size=10,
                              color="FFCBD5E1", italic=True)
            head_fill  = PatternFill("solid", fgColor=NAVY)
            zebra_fill = PatternFill("solid", fgColor=BG)
            kpi_fills  = [PatternFill("solid", fgColor=BLUE),
                          PatternFill("solid", fgColor=CYAN),
                          PatternFill("solid", fgColor=GREEN),
                          PatternFill("solid", fgColor=ORG)]
            thin = Side(border_style="thin", color=BORDER_CLR)
            box = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(horizontal="center", vertical="center")
            left = Alignment(horizontal="left", vertical="center")
            right = Alignment(horizontal="right", vertical="center")

            def style_header_row(ws, row, n_cols):
                for col in range(1, n_cols + 1):
                    c = ws.cell(row=row, column=col)
                    c.font = HEAD_FONT
                    c.fill = head_fill
                    c.alignment = left
                    c.border = box
                ws.row_dimensions[row].height = 22

            def style_data_zebra(ws, start_row, end_row, n_cols):
                for ri in range(start_row, end_row + 1):
                    for col in range(1, n_cols + 1):
                        c = ws.cell(row=ri, column=col)
                        c.font = BODY_FONT
                        c.border = box
                        if (ri - start_row) % 2 == 1:
                            c.fill = zebra_fill

            # ===== Sheet 1: Summary =====
            s = wb.active
            s.title = "Summary"
            s.sheet_view.showGridLines = False
            s.merge_cells("A1:F2")
            s["A1"] = APP_NAME
            s["A1"].font = HUGE_FONT
            s["A1"].fill = head_fill
            s["A1"].alignment = Alignment(horizontal="left",
                                          vertical="center", indent=2)
            s.merge_cells("A3:F3")
            s["A3"] = APP_TAGLINE + "  ·  capacitra.com"
            s["A3"].font = SUB_FONT
            s["A3"].fill = head_fill
            s["A3"].alignment = Alignment(horizontal="left",
                                          vertical="center", indent=2)
            s.row_dimensions[1].height = 24
            s.row_dimensions[2].height = 18
            s.row_dimensions[3].height = 20

            # Scan info
            s["A5"] = "Scanned root"
            s["B5"] = _sanitize_cell(self.scan_root)
            s["A6"] = "Generated"
            s["B6"] = time.strftime("%Y-%m-%d %H:%M")
            s["A7"] = "Report version"
            s["B7"] = f"{APP_NAME} v{APP_VERSION}"
            for row in (5, 6, 7):
                s.cell(row=row, column=1).font = MUTED_FONT
                s.cell(row=row, column=2).font = BODY_FONT

            # KPI band — 4 coloured boxes on row 9-10
            kpi_items = [
                ("SCANNED SIZE",   human_size(total)),
                ("FILES",          f"{files:,}"),
                ("FOLDERS",        f"{dirs:,}"),
                ("INACCESSIBLE",   f"{denied:,}"),
            ]
            for i, (label, value) in enumerate(kpi_items):
                col_lbl = i * 2 + 1
                col_val = col_lbl + 1
                cl = s.cell(row=9, column=col_lbl, value=label)
                cv = s.cell(row=10, column=col_lbl, value=value)
                cl.font = Font(name="Calibri", size=8, color="FFE2E8F0",
                               bold=True)
                cv.font = Font(name="Calibri", size=14, color="FFFFFFFF",
                               bold=True)
                cl.fill = kpi_fills[i]
                cv.fill = kpi_fills[i]
                cl.alignment = left
                cv.alignment = left
                s.cell(row=9, column=col_val).fill = kpi_fills[i]
                s.cell(row=10, column=col_val).fill = kpi_fills[i]
                s.merge_cells(start_row=9, start_column=col_lbl,
                              end_row=9, end_column=col_val)
                s.merge_cells(start_row=10, start_column=col_lbl,
                              end_row=10, end_column=col_val)
            s.row_dimensions[9].height = 18
            s.row_dimensions[10].height = 26

            # Disk usage table
            s["A12"] = "DISK USAGE"
            s["A12"].font = Font(name="Calibri", size=11, bold=True,
                                 color=NAVY)
            rows = [
                ("Disk capacity",        human_size(disk.get("total", 0))),
                ("Used (everything)",    human_size(disk.get("used", 0))),
                ("Scanned (accessible)",
                 f"{human_size(total)}  ·  {disk_pct:.1f}% of disk"),
                ("Free space",           human_size(disk.get("free", 0))),
                ("Inaccessible items",   f"{denied:,}"),
            ]
            start = 13
            for i, (k, v) in enumerate(rows):
                r0 = start + i
                s.cell(row=r0, column=1, value=k).font = MUTED_FONT
                s.cell(row=r0, column=2, value=v).font = Font(
                    name="Calibri", size=10, bold=True, color=FG)
                if i % 2 == 1:
                    for col in range(1, 3):
                        s.cell(row=r0, column=col).fill = zebra_fill
                for col in range(1, 3):
                    s.cell(row=r0, column=col).border = box

            # Column widths
            for col, w in [("A", 26), ("B", 36), ("C", 18), ("D", 12),
                           ("E", 12), ("F", 12)]:
                s.column_dimensions[col].width = w

            # ===== Sheet 2: Top folders =====
            f = wb.create_sheet("Top folders")
            f.sheet_view.showGridLines = False
            f.append(["Rank", "Folder", "Size (bytes)", "Size", "Share %"])
            style_header_row(f, 1, 5)
            denom = root.size or total or 1
            real = [c for c in root.children
                    if c.path not in ("(free)", "(gap)")]
            for i, c in enumerate(real[:200], start=1):
                f.append([i, _sanitize_cell(c.name), c.size, human_size(c.size),
                          round(c.size / denom * 100, 2)])
            style_data_zebra(f, 2, f.max_row, 5)
            for cell in f[1]:
                cell.alignment = center
            for col, w in [("A", 8), ("B", 56), ("C", 18), ("D", 16),
                           ("E", 12)]:
                f.column_dimensions[col].width = w
            f.freeze_panes = "A2"

            # ===== Sheet 3: File types =====
            ft = wb.create_sheet("File types")
            ft.sheet_view.showGridLines = False
            ft.append(["Extension", "Size (bytes)", "Size", "Files",
                       "Share %"])
            style_header_row(ft, 1, 5)
            tdenom = total or 1
            for ext, sz, cnt in r.get("extensions", [])[:200]:
                ft.append([_sanitize_cell(ext), sz, human_size(sz), cnt,
                           round(sz / tdenom * 100, 2)])
            style_data_zebra(ft, 2, ft.max_row, 5)
            for col, w in [("A", 18), ("B", 18), ("C", 16),
                           ("D", 12), ("E", 12)]:
                ft.column_dimensions[col].width = w
            ft.freeze_panes = "A2"

            # ===== Sheet 4: Largest files =====
            tf = wb.create_sheet("Largest files")
            tf.sheet_view.showGridLines = False
            tf.append(["Rank", "Size (bytes)", "Size", "Path"])
            style_header_row(tf, 1, 4)
            for i, (sz, p, _mt) in enumerate(
                    r.get("largest_files", [])[:500], start=1):
                tf.append([i, sz, human_size(sz), _sanitize_cell(p)])
            style_data_zebra(tf, 2, tf.max_row, 4)
            for col, w in [("A", 8), ("B", 18), ("C", 16), ("D", 100)]:
                tf.column_dimensions[col].width = w
            tf.freeze_panes = "A2"

            # ===== Sheet 5: Age cohorts =====
            ages = r.get("age_buckets", [])
            if ages:
                ag = wb.create_sheet("Age cohorts")
                ag.sheet_view.showGridLines = False
                ag.append(["Age range", "Files", "Total size (bytes)",
                           "Total size"])
                style_header_row(ag, 1, 4)
                for label, lo, hi, cnt, sz in ages:
                    ag.append([label, cnt, sz, human_size(sz)])
                style_data_zebra(ag, 2, ag.max_row, 4)
                for col, w in [("A", 22), ("B", 14), ("C", 22),
                               ("D", 18)]:
                    ag.column_dimensions[col].width = w
                ag.freeze_panes = "A2"

            wb.save(path)
            self.status_var.set(f"Excel saved: {path}")
            if self._ask("Done",
                         f"Excel saved:\n{path}\n\nOpen now?"):
                open_path(path)
        except Exception as e:
            self._error("Error", str(e))

    def _export_html(self):
        if not self.scan_result:
            self._warn("Run a scan first",
                       "Pick a drive or folder and click New Scan, then come back.")
            return
        p = filedialog.asksaveasfilename(
            title="Save HTML report", defaultextension=".html",
            filetypes=[("HTML", "*.html"), ("All files", "*.*")],
            initialfile="capacitra_report.html")
        if not p:
            return
        try:
            self._write_html(p)
            self.status_var.set(f"HTML saved: {p}")
            if self._ask("Done", f"HTML saved:\n{p}\n\nOpen now?"):
                open_path(p)
        except Exception as e:
            self._error("Error", str(e))

    def _write_html(self, path):
        r = self.scan_result
        root = r["root"]
        total = r["total_bytes"]
        files = r["total_files"]
        dirs = r.get("total_dirs", 0)
        denied = r.get("denied", 0)
        disk = r.get("disk", {"total": 0, "used": 0, "free": 0})
        disk_pct = (total / disk["total"] * 100) if disk.get("total") else 0
        free_pct = (disk["free"] / disk["total"] * 100) if disk.get("total") else 0
        gap = max(0, disk.get("used", 0) - total)
        gap_pct = (gap / disk["total"] * 100) if disk.get("total") else 0

        esc = html_lib.escape

        # ----- Inline SVG hex logo -----
        logo_svg = (
            '<svg viewBox="0 0 64 64" width="36" height="36" '
            'xmlns="http://www.w3.org/2000/svg" aria-hidden="true">'
            '<defs>'
            '<linearGradient id="hexG" x1="0" y1="0" x2="1" y2="1">'
            '<stop offset="0%" stop-color="#3B82F6"/>'
            '<stop offset="100%" stop-color="#1E3A8A"/>'
            '</linearGradient></defs>'
            '<polygon points="32,2 60,17 60,47 32,62 4,47 4,17" '
            'fill="url(#hexG)" stroke="#3B82F6" stroke-width="1.5"/>'
            '<polygon points="32,8 55,20 55,44 32,56 9,44 9,20" '
            'fill="#1D4ED8" opacity="0.85"/>'
            '<path d="M14 27 L26 27 L29 30 L48 30 L48 47 L14 47 Z" '
            'fill="#0EA5E9"/>'
            '<rect x="20" y="40" width="3" height="5" fill="#FFF"/>'
            '<rect x="26" y="37" width="3" height="8" fill="#FFF"/>'
            '<rect x="32" y="33" width="3" height="12" fill="#FFF"/>'
            '<rect x="38" y="38" width="3" height="7" fill="#FFF"/>'
            '</svg>')

        # Top 20 folders
        denom = root.size or total or 1
        real = [c for c in root.children
                if c.path not in ("(free)", "(gap)")]
        top_rows = []
        for i, c in enumerate(real[:20], start=1):
            pct = c.size / denom * 100
            top_rows.append(
                f"<tr><td class='num muted'>{i}</td>"
                f"<td><strong>{esc(c.name)}</strong></td>"
                f"<td class='num'>{human_size(c.size)}</td>"
                f"<td class='num'>{pct:.1f}%</td>"
                f"<td><div class='bar'><div class='bar-fill' "
                f"style='width:{pct:.1f}%'></div></div></td></tr>")
        rows_top = "".join(top_rows) or (
            "<tr><td colspan='5' class='muted'>No folders scanned.</td></tr>")

        # File types
        tdenom = total or 1
        ext_rows = []
        for ext, sz, cnt in r.get("extensions", [])[:20]:
            pct = sz / tdenom * 100
            ext_rows.append(
                f"<tr><td><span class='chip'>{esc(ext)}</span></td>"
                f"<td class='num'>{human_size(sz)}</td>"
                f"<td class='num'>{cnt:,}</td>"
                f"<td class='num'>{pct:.1f}%</td></tr>")
        rows_ext = "".join(ext_rows) or (
            "<tr><td colspan='4' class='muted'>No file types.</td></tr>")

        # Top 50 largest files
        big_rows = []
        for sz, ppath, _mt in r.get("largest_files", [])[:50]:
            big_rows.append(
                f"<tr><td class='num'>{human_size(sz)}</td>"
                f"<td class='path'>{esc(ppath)}</td></tr>")
        rows_big = "".join(big_rows) or (
            "<tr><td colspan='2' class='muted'>No files.</td></tr>")

        # Age cohorts
        age_rows = []
        for label, lo, hi, cnt, sz in r.get("age_buckets", []):
            age_rows.append(
                f"<tr><td>{esc(label)}</td>"
                f"<td class='num'>{cnt:,}</td>"
                f"<td class='num'>{human_size(sz)}</td></tr>")
        rows_age = "".join(age_rows)

        # Disk usage stacked bar
        sp = disk_pct
        gp = gap_pct
        fp = free_pct
        disk_html = (
            "<section class='card'>"
            "<h2>Disk usage</h2>"
            f"<div class='sum'>Capacity {human_size(disk.get('total', 0))} "
            f"&nbsp;·&nbsp; Used {human_size(disk.get('used', 0))} "
            f"&nbsp;·&nbsp; Free {human_size(disk.get('free', 0))}</div>"
            "<div class='diskbar'>"
            f"<span style='background:#2563EB;width:{sp:.1f}%'></span>"
            f"<span style='background:#EA580C;width:{gp:.1f}%'></span>"
            f"<span style='background:#D1D5DB;width:{fp:.1f}%'></span>"
            "</div>"
            "<div class='legend'>"
            f"<span><span class='dot' style='background:#2563EB'></span>"
            f"Scanned {human_size(total)} ({sp:.1f}%)</span>"
            f"<span><span class='dot' style='background:#EA580C'></span>"
            f"Inaccessible {human_size(gap)} ({gp:.1f}%)</span>"
            f"<span><span class='dot' style='background:#D1D5DB'></span>"
            f"Free {human_size(disk.get('free', 0))} ({fp:.1f}%)</span>"
            "</div></section>")

        css = (
            ":root{--bg:#F8FAFC;--panel:#FFF;--fg:#0F172A;"
            "--muted:#64748B;--accent:#2563EB;--navy:#1E3A8A;"
            "--cyan:#0EA5E9;--border:#E2E8F0}"
            "*{box-sizing:border-box}"
            "body{margin:0;font:14px/1.55 'Segoe UI','Inter',system-ui,sans-serif;"
            "background:var(--bg);color:var(--fg)}"
            "header{background:linear-gradient(135deg,#1E3A8A 0%,#2563EB 60%,#0EA5E9 130%);"
            "color:#FFF;padding:30px 40px;display:flex;align-items:center;gap:18px}"
            "header svg{flex-shrink:0}"
            "header .brand h1{margin:0;font-size:24px;font-weight:700;letter-spacing:-.3px}"
            "header .brand .tag{opacity:.85;margin-top:2px;font-size:12px;"
            "text-transform:uppercase;letter-spacing:.8px}"
            "header .meta{margin-left:auto;text-align:right;font-size:12px;"
            "color:rgba(255,255,255,.85)}"
            "header .meta .path{font-family:'Consolas',monospace;"
            "opacity:.9;font-size:11px;margin-top:2px}"
            ".container{max-width:1200px;margin:0 auto;padding:28px}"
            ".kpis{display:grid;grid-template-columns:repeat(4,1fr);"
            "gap:14px;margin-bottom:22px}"
            ".kpi{background:var(--panel);border:1px solid var(--border);"
            "border-radius:14px;padding:18px 20px;"
            "box-shadow:0 1px 2px rgba(15,23,42,.04)}"
            ".kpi .lbl{color:var(--muted);font-size:10.5px;"
            "text-transform:uppercase;letter-spacing:.8px;font-weight:600}"
            ".kpi .val{font-size:24px;font-weight:700;margin-top:4px;"
            "letter-spacing:-.5px}"
            ".kpi.b1 .val{color:#2563EB}.kpi.b2 .val{color:#0EA5E9}"
            ".kpi.b3 .val{color:#16A34A}.kpi.b4 .val{color:#EA580C}"
            "section.card{background:var(--panel);border:1px solid var(--border);"
            "border-radius:14px;padding:22px;margin-bottom:18px;"
            "box-shadow:0 1px 2px rgba(15,23,42,.04)}"
            "section.card h2{margin:0 0 12px;font-size:16px;font-weight:700;"
            "color:var(--navy)}"
            ".sum{color:var(--muted);font-size:13px;margin-bottom:14px}"
            ".diskbar{display:flex;height:18px;border-radius:9px;"
            "overflow:hidden;background:#EEF1F5}"
            ".diskbar span{display:block;height:100%}"
            ".dot{display:inline-block;width:10px;height:10px;border-radius:50%;"
            "margin-right:6px;vertical-align:middle}"
            ".legend{margin-top:10px;font-size:12px;color:var(--muted);"
            "display:flex;flex-wrap:wrap;gap:18px}"
            "table{width:100%;border-collapse:separate;border-spacing:0}"
            "th,td{padding:10px 14px;border-bottom:1px solid var(--border);"
            "text-align:left;font-size:13px;vertical-align:middle}"
            "th{background:var(--navy);color:#FFF;font-weight:600;"
            "text-transform:uppercase;font-size:10.5px;letter-spacing:.6px}"
            "th:first-child{border-top-left-radius:8px}"
            "th:last-child{border-top-right-radius:8px}"
            "tbody tr:nth-child(even){background:var(--bg)}"
            "td.num{text-align:right;font-variant-numeric:tabular-nums}"
            "td.muted,.muted{color:var(--muted)}"
            "td.path{font-family:'Consolas',monospace;font-size:12px;"
            "color:#334155;word-break:break-all}"
            ".bar{height:8px;background:#EEF1F5;border-radius:99px;"
            "overflow:hidden;min-width:160px}"
            ".bar-fill{height:100%;background:linear-gradient(90deg,"
            "#1D4ED8,#0EA5E9);border-radius:99px}"
            ".chip{display:inline-block;background:#DBEAFE;color:#1E3A8A;"
            "padding:3px 10px;border-radius:99px;font-weight:600;"
            "font-size:11.5px;font-family:'Consolas',monospace}"
            "footer{text-align:center;color:var(--muted);"
            "padding:24px;font-size:12px;border-top:1px solid var(--border)}"
            "footer a{color:var(--accent);text-decoration:none}"
            "@media print{header,section.card{break-inside:avoid}}"
        )

        out = (
            "<!doctype html><html lang=\"en\"><head>"
            "<meta charset=\"utf-8\">"
            f"<title>{esc(APP_NAME)} report · {esc(self.scan_root)}</title>"
            "<meta name=\"viewport\" content=\"width=device-width,initial-scale=1\">"
            f"<style>{css}</style></head><body>"
            "<header>"
            f"{logo_svg}"
            "<div class='brand'>"
            f"<h1>{esc(APP_NAME)}</h1>"
            f"<div class='tag'>{esc(APP_TAGLINE)}</div>"
            "</div>"
            "<div class='meta'>"
            f"<div>Report generated {time.strftime('%B %d, %Y · %H:%M')}</div>"
            f"<div class='path'>{esc(self.scan_root)}</div>"
            "</div>"
            "</header>"
            "<div class='container'>"
            "<div class='kpis'>"
            f"<div class='kpi b1'><div class='lbl'>Scanned size</div>"
            f"<div class='val'>{human_size(total)}</div></div>"
            f"<div class='kpi b2'><div class='lbl'>Files</div>"
            f"<div class='val'>{files:,}</div></div>"
            f"<div class='kpi b3'><div class='lbl'>Folders</div>"
            f"<div class='val'>{dirs:,}</div></div>"
            f"<div class='kpi b4'><div class='lbl'>Inaccessible</div>"
            f"<div class='val'>{denied:,}</div></div>"
            "</div>"
            + disk_html +
            "<section class='card'><h2>Top 20 folders by size</h2>"
            "<table><thead><tr><th>#</th><th>Folder</th>"
            "<th class='num'>Size</th><th class='num'>Share</th>"
            "<th></th></tr></thead>"
            f"<tbody>{rows_top}</tbody></table></section>"
            "<section class='card'><h2>Top 20 file types</h2>"
            "<table><thead><tr><th>Extension</th>"
            "<th class='num'>Total size</th>"
            "<th class='num'>Files</th>"
            "<th class='num'>Share</th></tr></thead>"
            f"<tbody>{rows_ext}</tbody></table></section>"
            "<section class='card'><h2>Top 50 largest files</h2>"
            "<table><thead><tr><th class='num'>Size</th>"
            "<th>Path</th></tr></thead>"
            f"<tbody>{rows_big}</tbody></table></section>"
            + (("<section class='card'><h2>Files by age</h2>"
                "<table><thead><tr><th>Age range</th>"
                "<th class='num'>Files</th>"
                "<th class='num'>Total size</th></tr></thead>"
                f"<tbody>{rows_age}</tbody></table></section>")
               if rows_age else "")
            + "</div>"
            f"<footer>Generated by <strong>{esc(APP_NAME)}</strong> "
            f"v{esc(APP_VERSION)} · {esc(APP_TAGLINE)} · "
            f"<a href='https://capacitra.com'>capacitra.com</a></footer>"
            "</body></html>"
        )
        with open(path, "w", encoding="utf-8") as f:
            f.write(out)


def _make_app_icon():
    """Render the Capacitra hex+folder+bars logo as a PhotoImage."""
    try:
        import math as _m
        SIZE = 64
        img = tk.PhotoImage(width=SIZE, height=SIZE)

        def _putpx(x, y, color):
            if 0 <= x < SIZE and 0 <= y < SIZE:
                try:
                    img.put(color, (int(x), int(y)))
                except tk.TclError:
                    pass

        def _fill_poly(points, color):
            ys = [p[1] for p in points]
            y_min = max(0, int(min(ys)))
            y_max = min(SIZE - 1, int(max(ys)))
            n = len(points)
            for y in range(y_min, y_max + 1):
                xs = []
                for i in range(n):
                    p1, p2 = points[i], points[(i + 1) % n]
                    y1, y2 = p1[1], p2[1]
                    if (y1 <= y < y2) or (y2 <= y < y1):
                        if y2 != y1:
                            x = p1[0] + (y - y1) * (p2[0] - p1[0]) / (y2 - y1)
                            xs.append(x)
                xs.sort()
                for i in range(0, len(xs) - 1, 2):
                    x1, x2 = int(xs[i]), int(xs[i + 1])
                    for x in range(max(0, x1), min(SIZE, x2 + 1)):
                        _putpx(x, y, color)

        cx, cy, r = 32, 32, 28
        hex_outer = []
        for i in range(6):
            ang = _m.radians(-90 + 60 * i)
            hex_outer.append((cx + r * _m.cos(ang), cy + r * _m.sin(ang)))
        _fill_poly(hex_outer, "#1E3A8A")

        hex_inner = []
        for i in range(6):
            ang = _m.radians(-90 + 60 * i)
            hex_inner.append((cx + (r - 3) * _m.cos(ang),
                              cy + (r - 3) * _m.sin(ang)))
        _fill_poly(hex_inner, "#1D4ED8")

        fx, fy, fw, fh = 14, 22, 36, 22
        folder = [
            (fx,             fy),
            (fx + fw * 0.45, fy),
            (fx + fw * 0.55, fy + 3),
            (fx + fw,        fy + 3),
            (fx + fw,        fy + fh),
            (fx,             fy + fh),
        ]
        _fill_poly(folder, "#0EA5E9")

        bx = fx + 4
        by = fy + fh - 2
        bw = (fw - 8) / 5
        for i, h in enumerate([fh * 0.35, fh * 0.55, fh * 0.8]):
            x0 = bx + (i * 1.6 + 0.5) * bw
            for yy in range(int(by - h), int(by)):
                for xx in range(int(x0), int(x0 + bw)):
                    _putpx(xx, yy, "#FFFFFF")

        return img
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Headless CLI mode (v4.1)
#
# Usage:
#   Capacitra.exe --scan D:\ --export report.csv
#   Capacitra.exe --scan "C:\Users\me" --export report.json --quiet
#   Capacitra.exe --scan D:\ --export report.csv --exclude node_modules --exclude .git
#
# Exit codes:
#   0 = success, 1 = usage error, 2 = scan error, 3 = export error.
# ---------------------------------------------------------------------------

def _cli_walk_csv(node, writer):
    kind = "Folder" if node.is_dir else "File"
    mt = (time.strftime("%Y-%m-%d %H:%M", time.localtime(node.mtime))
          if node.mtime else "")
    writer.writerow([
        _sanitize_cell(node.path), kind,
        node.size, human_size(node.size),
        _sanitize_cell(mt),
    ])
    for c in node.children:
        _cli_walk_csv(c, writer)


def _cli_walk_json(node):
    return {
        "path": node.path,
        "name": node.name,
        "is_dir": bool(node.is_dir),
        "size": int(node.size or 0),
        "mtime": float(node.mtime or 0),
        "children": [_cli_walk_json(c) for c in node.children],
    }


def _run_cli(argv):
    _enable_hidpi()
    parser = argparse.ArgumentParser(
        prog="Capacitra",
        description="Capacitra headless disk-space scanner. "
                    "Runs a scan and writes a report without opening the GUI.",
    )
    parser.add_argument("--scan", metavar="PATH", required=True,
                        help="Root path to scan (e.g. D:\\ or C:\\Users\\me).")
    parser.add_argument("--export", metavar="FILE", required=True,
                        help="Output file. Format is inferred from the "
                             "extension: .csv or .json.")
    parser.add_argument("--exclude", action="append", default=[],
                        metavar="NAME",
                        help="Folder name to skip (repeatable, "
                             "case-insensitive). Example: --exclude node_modules")
    parser.add_argument("--quiet", action="store_true",
                        help="Suppress progress output (for cron / Task "
                             "Scheduler).")
    parser.add_argument("--version", action="version",
                        version=f"Capacitra {APP_VERSION}")

    try:
        args = parser.parse_args(argv)
    except SystemExit as e:
        # argparse already printed its own message
        return int(e.code) if e.code is not None else 1

    root_path = args.scan
    out_path = args.export
    ext = os.path.splitext(out_path)[1].lower()

    if ext not in (".csv", ".json"):
        sys.stderr.write(
            f"Capacitra: unsupported export format '{ext}'. "
            f"Use .csv or .json.\n"
        )
        return 1

    if not os.path.exists(root_path):
        sys.stderr.write(
            f"Capacitra: scan root does not exist: {root_path}\n"
        )
        return 1

    try:
        sys.setrecursionlimit(50000)
    except Exception:
        pass

    if not args.quiet:
        sys.stdout.write(f"Capacitra {APP_VERSION} scanning {root_path}...\n")
        sys.stdout.flush()

    q = queue.Queue()
    stop = threading.Event()
    worker = ScanWorker(root_path, q, stop, excludes=args.exclude)
    worker.start()

    result = None
    last_tick = time.time()
    while worker.is_alive() or not q.empty():
        try:
            msg = q.get(timeout=0.5)
        except queue.Empty:
            if not args.quiet and time.time() - last_tick > 2.0:
                sys.stdout.write(
                    f"  ... {worker.scanned_files:>10,} files, "
                    f"{human_size(worker.total_bytes)}\r"
                )
                sys.stdout.flush()
                last_tick = time.time()
            continue
        kind, data = msg
        if kind == "done":
            result = data
            break
        if kind == "error":
            sys.stderr.write(f"\nCapacitra: scan failed: {data}\n")
            return 2
        if kind == "cancelled":
            sys.stderr.write("\nCapacitra: scan cancelled\n")
            return 2

    if result is None:
        sys.stderr.write("\nCapacitra: scan produced no result\n")
        return 2

    if not args.quiet:
        sys.stdout.write("\n")
        sys.stdout.write(
            f"Scan done: {result['total_files']:,} files, "
            f"{result.get('total_dirs', 0):,} folders, "
            f"{human_size(result['total_bytes'])}\n"
        )

    try:
        if ext == ".csv":
            with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["Path", "Type", "Size (bytes)", "Size",
                            "Last modified"])
                _cli_walk_csv(result["root"], w)
        else:
            payload = {
                "capacitra_version": APP_VERSION,
                "generated": time.strftime("%Y-%m-%dT%H:%M:%S"),
                "scan_root": root_path,
                "total_files": result["total_files"],
                "total_dirs": result.get("total_dirs", 0),
                "total_bytes": result["total_bytes"],
                "denied": result.get("denied", 0),
                "extensions": [
                    {"ext": e, "size": sz, "count": cnt}
                    for (e, sz, cnt) in result.get("extensions", [])
                ],
                "largest_files": [
                    {"size": sz, "path": p,
                     "mtime": mt if isinstance(mt, (int, float)) else 0}
                    for (sz, p, mt) in result.get("largest_files", [])
                ],
                "age_buckets": [
                    {"label": b[0], "count": b[3], "bytes": b[4]}
                    for b in result.get("age_buckets", [])
                ],
                "tree": _cli_walk_json(result["root"]),
            }
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=2, ensure_ascii=False)
    except Exception as e:
        sys.stderr.write(f"Capacitra: export failed: {e}\n")
        return 3

    if not args.quiet:
        sys.stdout.write(f"Report written: {out_path}\n")
    return 0




def _enable_hidpi():
    """Enable Per-Monitor DPI awareness on Windows so tkinter renders
    crisply on 150 / 175 / 200% displays instead of bitmap-scaled.

    Tries Per-Monitor-V2 (Win 10 1607+), falls back to Per-Monitor
    (Win 8.1+), then System DPI (Vista+). Silent no-op otherwise."""
    if sys.platform != "win32":
        return
    try:
        from ctypes import windll, wintypes
        # Per-Monitor-V2 = -4 (Win10 1607+)
        try:
            windll.user32.SetProcessDpiAwarenessContext(-4)
            return
        except Exception:
            pass
        # Per-Monitor = 2 (Win 8.1+)
        try:
            windll.shcore.SetProcessDpiAwareness(2)
            return
        except Exception:
            pass
        # System DPI (Vista+)
        try:
            windll.user32.SetProcessDPIAware()
        except Exception:
            pass
    except Exception:
        pass

def main():
    _enable_hidpi()
    # Deep folder trees (WinSxS, node_modules) exceed the default
    # 1000-frame recursion limit. Raise well above realistic depths.
    try:
        sys.setrecursionlimit(50000)
    except Exception:
        pass
    root = tk.Tk()
    # Build the app immediately so the window appears in ~200 ms.
    # Defer the icon (pixel-by-pixel render, ~700 ms cost) to after
    # the first idle event so the user perceives a much faster boot.
    def _apply_icon_async():
        try:
            icon = _make_app_icon()
            if icon is not None:
                root.iconphoto(True, icon)
                root._capacitra_icon = icon
        except Exception:
            pass
    root.after(150, _apply_icon_async)
    CapacitraApp(root)
    root.mainloop()


if __name__ == "__main__":
    if len(sys.argv) > 1 and any(a.startswith("--") for a in sys.argv[1:]):
        raise SystemExit(_run_cli(sys.argv[1:]))
    main()
